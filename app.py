"""
app.py
Gestione Registrazioni SEG - Web App (Streamlit + Google Sheets)
"""

from datetime import datetime, date, timedelta
import io
import os
import re
import zipfile

import pandas as pd
import streamlit as st
from st_keyup import st_keyup
from streamlit_gsheets import GSheetsConnection

import gspread
from google.oauth2.service_account import Credentials
import pdfplumber
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# 1. CONFIGURAZIONE PAGINA (Deve essere la prima istruzione Streamlit)
# ==============================================================================
st.set_page_config(
    page_title="Gestione Registrazioni SEG",
    page_icon="📒",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ==============================================================================
# 2. CONFIGURAZIONE AUTENTICAZIONE GOOGLE OAUTH NATIVA (st.login())
# ==============================================================================
NOME_FOGLIO_UTENTI = "Utenti"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.readonly",
]

@st.cache_resource(show_spinner=False)
def get_client() -> gspread.Client:
    """Autentica il programma verso Google tramite l'account di servizio."""
    credenziali = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"], scopes=SCOPES
    )
    return gspread.authorize(credenziali)


@st.cache_resource(show_spinner=False)
def apri_foglio_dati():
    """Apre il foglio Google dati. Ritorna (workbook, errore)."""
    try:
        client = get_client()
        wb = client.open_by_key(st.secrets["sheet_id"])
        return wb, None
    except gspread.exceptions.APIError:
        email_sa = st.secrets["gcp_service_account"]["client_email"]
        return None, (
            "Impossibile aprire il foglio dati. Controlla che sia stato "
            f"condiviso (come Editor) con:\n`{email_sa}`"
        )
    except Exception as e:
        return None, f"Errore durante il collegamento: {e}"
# ─────────────────────────────────────────────────────────────────


# ==============================================================================
# 3. PANNELLO DI AUTENTICAZIONE GOOGLE
# ==============================================================================

def sola_lettura() -> bool:
    """Ritorna True se l'utente corrente ha accesso in sola lettura (ruolo 'utente')."""
    return st.session_state.get("ruolo") == "utente"


def verifica_utente_foglio(email_cercata):
    """Ritorna (autorizzato, ruolo, errore). 'errore' è valorizzato solo se la
    lettura del foglio Utenti è fallita (non se l'email semplicemente non c'è)."""
    wb, err = apri_foglio_dati()
    if err:
        return False, None, err
    try:
        ws = wb.worksheet(NOME_FOGLIO_UTENTI)
        valori = ws.get_all_values()
        for riga in valori[1:]:
            if len(riga) >= 4:
                email_foglio = riga[2].strip().lower()
                ruolo_foglio = riga[3].strip()
                if email_foglio == email_cercata:
                    return True, ruolo_foglio, None
    except Exception as e:
        return False, None, f"Errore durante la lettura del foglio «{NOME_FOGLIO_UTENTI}»: {e}"
    return False, None, None


# Se l'utente non ha ancora effettuato il login con Google (st.login() gestisce
# lui la sessione persistente per 30 giorni, niente cookie manager manuale)
if not st.user.is_logged_in:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.title("🔒 Accesso Riservato")
        st.subheader("Gestione Registrazioni SEG")
        st.write("Accedi utilizzando il tuo account Google autorizzato.")
        if st.button("🔑 Accedi con Google", use_container_width=True, type="primary"):
            st.login()
    st.stop()

# Da qui in poi st.user.email è affidabile (autenticato da Google)
email_autenticata = (st.user.email or "").strip().lower()

if st.session_state.get("email_logged") != email_autenticata:
    autorizzato, ruolo_trovato, errore_verifica = verifica_utente_foglio(email_autenticata)
    if autorizzato:
        st.session_state.email_logged = email_autenticata
        st.session_state.ruolo = ruolo_trovato.lower()
        if st.session_state.ruolo == "presenze":
            st.session_state.pagina = "presenze"
    elif errore_verifica:
        st.error(f"⚠️ Errore durante la verifica dell'accesso: {errore_verifica}")
        st.info("Potrebbe essere un problema temporaneo di connessione al foglio Google. "
                "Riprova tra qualche secondo.")
        st.stop()
    else:
        st.error(f"⚠️ L'account `{email_autenticata}` non è autorizzato ad accedere.")
        if st.button("🚪 Esci e riprova con un altro account"):
            st.session_state.pop("email_logged", None)
            st.session_state.pop("ruolo", None)
            st.logout()
        st.stop()


# ==============================================================================
# 4. AREA RISERVATA (DISPONIBILE SOLO A UTENTI AUTORIZZATI)
# ==============================================================================

# Barra laterale con dati utente e Logout
with st.sidebar:
    st.write("👤 Utente connesso:")
    st.write(f"📧 `{st.session_state.email_logged}`")
    
    # Recupera il ruolo salvato in sessione e lo mostra con l'iniziale maiuscola
    ruolo_utente = str(st.session_state.get("ruolo", "Non specificato")).capitalize()
    st.write(f"🏷️ **Ruolo:** `{ruolo_utente}`")

    if sola_lettura():
        st.caption("🔒 Modalità sola lettura: puoi consultare i dati ma non modificarli.")

    st.divider()  # Linea di separazione visiva
    
    if st.button("🚪 Logout", type="secondary", use_container_width=True):
        st.session_state.pop("ruolo", None)
        st.session_state.pop("email_logged", None)
        st.logout()

# ─────────────────────────────────────────────────────────────────
# COSTANTI E CONFIGURAZIONI DEL SISTEMA
# ─────────────────────────────────────────────────────────────────
# SCOPES è stato spostato in alto

NOME_FOGLIO_RISPOSTE = "Risposte del modulo 9"
RIGA_INTESTAZIONE_RISPOSTE = 9

NOME_FOGLIO_PRESENZE = "Presenze Adunanze"
RIGA_INTESTAZIONE_PRESENZE = 1
TIPI_ADUNANZA = ["Infrasettimanale", "Fine settimana"]

# ── Impostazioni configurabili (es. giorni in cui si tengono le adunanze) ──
NOME_FOGLIO_IMPOSTAZIONI = "Configurazioni"
RIGA_INTESTAZIONE_IMPOSTAZIONI = 1
CHIAVE_GIORNI_PER_TIPO = {
    "Infrasettimanale": "Giorni Adunanza Infrasettimanale",
    "Fine settimana": "Giorni Adunanza Fine Settimana",
}
GIORNI_SETTIMANA_IT = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
GIORNI_ADUNANZE_DEFAULT = {"Infrasettimanale": ["Giovedì"], "Fine settimana": ["Domenica"]}


@st.cache_data(ttl=300, show_spinner=False)
def leggi_giorni_adunanze_per_tipo(_workbook) -> dict:
    """Legge dal foglio 'Configurazioni' i giorni della settimana per
    ciasccun tipo di adunanza (Infrasettimanale/Fine settimana)."""
    risultato = {tipo: list(giorni) for tipo, giorni in GIORNI_ADUNANZE_DEFAULT.items()}
    if _workbook is None:
        return risultato
    try:
        ws = _workbook.worksheet(NOME_FOGLIO_IMPOSTAZIONI)
        valori = ws.get_all_values()
        righe = valori[RIGA_INTESTAZIONE_IMPOSTAZIONI:]
        for tipo, chiave in CHIAVE_GIORNI_PER_TIPO.items():
            for riga in righe:
                if len(riga) >= 2 and riga[0].strip().lower() == chiave.lower():
                    giorni = [g.strip() for g in riga[1].split(",") if g.strip()]
                    giorni_validi = [g for g in giorni if g in GIORNI_SETTIMANA_IT]
                    if giorni_validi:
                        risultato[tipo] = giorni_validi
                    break
    except Exception:
        pass
    return risultato


def giorni_adunanze_tutti(giorni_per_tipo: dict) -> list:
    """Unisce i giorni di tutti i tipi in un'unica lista (senza doppioni)."""
    tutti = []
    for giorni in giorni_per_tipo.values():
        for g in giorni:
            if g not in tutti:
                tutti.append(g)
    return tutti


def tipo_adunanza_del_giorno(nome_giorno: str, giorni_per_tipo: dict):
    """Ritorna il tipo di adunanza associato al giorno della settimana indicato."""
    for tipo, giorni in giorni_per_tipo.items():
        if nome_giorno in giorni:
            return tipo
    return None


def salva_giorni_adunanze_per_tipo(_workbook, giorni_per_tipo: dict):
    """Salva (o crea) le righe dei giorni adunanza nel foglio 'Configurazioni'."""
    try:
        ws = _workbook.worksheet(NOME_FOGLIO_IMPOSTAZIONI)
    except Exception:
        return False, (f"Il foglio «{NOME_FOGLIO_IMPOSTAZIONI}» non esiste ancora nel documento Google. "
                       f"Creane uno con questo nome esatto e metti «Chiave» in A1 e «Valore» in B1, poi riprova.")
    try:
        valori = ws.get_all_values()
        righe = valori[RIGA_INTESTAZIONE_IMPOSTAZIONI:]
        for tipo, chiave in CHIAVE_GIORNI_PER_TIPO.items():
            valore_str = ",".join(giorni_per_tipo.get(tipo, []))
            trovata = False
            for i, riga in enumerate(righe, start=RIGA_INTESTAZIONE_IMPOSTAZIONI + 1):
                if len(riga) >= 1 and riga[0].strip().lower() == chiave.lower():
                    ws.update_cell(i, 2, valore_str)
                    trovata = True
                    break
            if not trovata:
                ws.append_row([chiave, valore_str], value_input_option="USER_ENTERED")
        return True, None
    except Exception as e:
        return False, f"Errore durante il salvataggio: {e}"


def _prossima_data_valida_precedente(data_scelta, giorni_validi: list):
    """Se 'data_scelta' cade in uno dei giorni della settimana validi,
    ritorna (True, None). Altrimenti ritorna (False, data_proposta)."""
    nome_giorno = GIORNI_SETTIMANA_IT[data_scelta.weekday()]
    if nome_giorno in giorni_validi:
        return True, None
    for delta in range(1, 8):
        candidata = data_scelta - timedelta(days=delta)
        if GIORNI_SETTIMANA_IT[candidata.weekday()] in giorni_validi:
            return False, candidata
    return False, None

NOME_FOGLIO_ANAGRAFICA = "Anagrafica"
RIGA_INTESTAZIONE_ANAGRAFICA = 1

NOME_FOGLIO_TUTTI = "Tutti"
RIGA_INTESTAZIONE_TUTTI = 4
COL_TUTTI_NOME = 1
COL_TUTTI_MESE = 2
COL_TUTTI_TIPO_SERVIZIO = 3
COL_TUTTI_MINISTERO = 4
COL_TUTTI_ORE = 6
COL_TUTTI_CRED_ORE = 7
COL_TUTTI_STUDI = 8
COL_TUTTI_OSSERVAZIONI = 9

MESI_ITALIANI = {
    1: "Gennaio", 2: "Febbraio", 3: "Marzo", 4: "Aprile", 5: "Maggio", 6: "Giugno",
    7: "Luglio", 8: "Agosto", 9: "Settembre", 10: "Ottobre", 11: "Novembre", 12: "Dicembre",
}

OPZIONI_SESSO = ["Maschio", "Femmina"]
OPZIONI_INCARICO = ["(nessuno)", "Anziano", "Servitore di ministero"]
OPZIONI_TIPO = ["Proclamatore", "Pioniere Regolare", "Pioniere speciale", "Missionario sul campo"]
OPZIONI_HAI_SERVITO = ["Proclamatore", "Pioniere Ausiliario", "Pioniere Regolare",
                       "Pioniere Speciale", "Rappresentante sul campo"]
OPZIONI_ATTIVI_INATTIVI = ["A", "I", "TR"]

RIGA_INTESTAZIONE_UTENTI = 1
OPZIONI_RUOLO_UTENTE = ["Amministratore", "Utente", "Presenze"]
REGEX_EMAIL_VALIDA = re.compile(r"^[A-Za-z0-9_.+-]+@[A-Za-z0-9-]+\.[A-Za-z0-9.-]+$")
ETICHETTE_ATTIVI_INATTIVI = {"A": "Attivo", "I": "Inattivo", "TR": "Trasferito"}


def categoria_stato_proclamatore(valore: str) -> str:
    """Riconosce la categoria di stato da 'Attivi / Inattivi'."""
    v = (valore or "").strip().lower()
    if v.startswith("i"):
        return "I"
    if v.startswith("t"):
        return "TR"
    return "A"

NOME_CONGREGAZIONE = "Vibo Marina"

PERCORSO_MODULO_S21 = os.path.join(os.path.dirname(__file__), "S-21_s-Mlt_I.pdf")
S21_PAGE_W, S21_PAGE_H = 595.2, 841.9
S21_OFFSET_PANNELLO = 421.0

PERCORSO_MODULO_S88 = os.path.join(os.path.dirname(__file__), "S-88_I.pdf")
S88_PAGE_W, S88_PAGE_H = 595.2, 841.9

S21_ORDINE_MESI = ["Settembre", "Ottobre", "Novembre", "Dicembre", "Gennaio", "Febbraio",
                    "Marzo", "Aprile", "Maggio", "Giugno", "Luglio", "Agosto"]

S21_RIGHE = {
    "Settembre": (158.44, 171.09), "Ottobre": (176.17, 188.83), "Novembre": (193.90, 206.56),
    "Dicembre": (211.64, 224.29), "Gennaio": (229.37, 242.02), "Febbraio": (247.10, 259.75),
    "Marzo": (264.83, 277.48), "Aprile": (282.56, 295.21), "Maggio": (300.29, 312.94),
    "Giugno": (318.02, 330.68), "Luglio": (335.75, 348.41), "Agosto": (353.48, 366.14),
}
S21_TOTALE_RIGA = (371.9, 389.5)

S21_COL_MINISTERO = (130.28, 143.01)
S21_COL_AUSILIARIO = (271.52, 284.25)
S21_COL_STUDI = (172.0, 242.6)
S21_COL_ORE = (313.2, 383.8)
S21_COL_OSSERVAZIONI_X = 388.5

S21_BOX_SESSO_M = (384.18, 394.4, 53.12, 63.28)
S21_BOX_SESSO_F = (485.58, 495.8, 53.12, 63.28)
S21_BOX_ALTRE_PECORE = (384.18, 394.4, 67.65, 77.8)
S21_BOX_UNTO = (485.58, 495.8, 67.65, 77.8)
S21_BOX_ANZIANO = (17.03, 27.25, 82.08, 92.24)
S21_BOX_SERVITORE = (80.58, 90.8, 82.08, 92.24)
S21_BOX_PIONIERE_REGOLARE = (216.05, 226.27, 82.08, 92.24)
S21_BOX_PIONIERE_SPECIALE = (327.5, 337.72, 82.08, 92.24)
S21_BOX_MISSIONARIO = (436.37, 446.59, 82.08, 92.24)

S21_COL_ANNO_SERVIZIO = (17.5, 101.4)
S21_ANNO_LABEL_TOP = 145.5
S21_SPOSTAMENTO_RIGHE = 2.0
S21_SPOSTAMENTO_TESTATA = 1.8

S21_FONT_VALORI = 10.5
S21_FONT_CHECK_HEADER = 9.5
S21_FONT_TABELLA = 9.5
S21_FONT_CHECK_TABELLA = 10.5
S21_FONT_ETA = 8.5

S21_COLORE_ROSSO = (0.827, 0.125, 0.125)
S21_COLORE_NERO = (0, 0, 0)


# ─────────────────────────────────────────────────────────────────
# CONNESSIONE A GOOGLE
# ─────────────────────────────────────────────────────────────────
# Le funzioni get_client() e apri_foglio_dati() sono state spostate in alto

@st.cache_data(ttl=60, show_spinner=False)
def leggi_foglio_come_df(_workbook, nome_foglio: str, riga_intestazione: int = 1):
    """Legge un foglio (tab) del workbook e lo ritorna come DataFrame."""
    try:
        ws = _workbook.worksheet(nome_foglio)
    except gspread.WorksheetNotFound:
        nomi_disponibili = ", ".join(f"'{f.title}'" for f in _workbook.worksheets())
        return None, (
            f"Il foglio '{nome_foglio}' non esiste nel documento collegato. "
            f"Fogli disponibili: {nomi_disponibili}."
        )
    except Exception as e:
        return None, f"Errore durante la lettura del foglio: {e}"

    tutti_i_valori = ws.get_all_values()
    if len(tutti_i_valori) < riga_intestazione:
        return pd.DataFrame(), None

    intestazioni = tutti_i_valori[riga_intestazione - 1]
    righe_dati = tutti_i_valori[riga_intestazione:]

    intestazioni_pulite = []
    contatori = {}
    for i, nome in enumerate(intestazioni):
        nome = nome.strip() or f"Colonna {i + 1}"
        if nome in contatori:
            contatori[nome] += 1
            nome = f"{nome} ({contatori[nome]})"
        else:
            contatori[nome] = 0
        intestazioni_pulite.append(nome)

    righe_dati = [r for r in righe_dati if any(cella.strip() for cella in r)]

    df = pd.DataFrame(righe_dati, columns=intestazioni_pulite)
    return df, None


def trova_indice_colonna(intestazioni: list, parola_chiave: str):
    """Cerca la prima colonna la cui intestazione contiene 'parola_chiave'."""
    parola_chiave = parola_chiave.lower()
    for i, nome in enumerate(intestazioni):
        if parola_chiave in nome.lower():
            return i
    return None


def calcola_eta(data_str: str) -> str:
    """Calcola gli anni compiuti da una data in formato gg/mm/aaaa."""
    if not data_str:
        return ""
    try:
        d = datetime.strptime(data_str, "%d/%m/%Y")
    except ValueError:
        return ""
    oggi = datetime.now()
    anni = oggi.year - d.year - ((oggi.month, oggi.day) < (d.month, d.day))
    return str(anni)


def calcola_eta_dettagliata(data_str: str) -> str:
    """Calcola anni e mesi compiuti da una data in formato gg/mm/aaaa."""
    if not data_str:
        return ""
    try:
        d = datetime.strptime(data_str, "%d/%m/%Y")
    except ValueError:
        return ""
    oggi = datetime.now()
    anni = oggi.year - d.year
    mesi = oggi.month - d.month
    if oggi.day < d.day:
        mesi -= 1
    if mesi < 0:
        anni -= 1
        mesi += 12
    return f"{anni},{mesi}"


def opzioni_da_colonna(df: pd.DataFrame, nome_colonna: str) -> list:
    """Ritorna i valori unici (non vuoti) già presenti in una colonna del DataFrame."""
    if nome_colonna not in df.columns:
        return []
    valori = df[nome_colonna].astype(str).str.strip()
    valori = sorted({v for v in valori if v and v.lower() != "nan"})
    return valori


def a_float_it(s: str) -> float:
    """Converte una stringa numerica in formato italiano in float."""
    s = (s or "").strip()
    if not s:
        return 0.0
    try:
        return float(s.replace(".", "").replace(",", "."))
    except ValueError:
        return 0.0


def formatta_numero_it(v: float) -> str:
    """Formatta un numero come stringa in stile italiano: 42,00"""
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def formatta_mese_esteso(mese_anno: str) -> str:
    """Converte 'AAAA-MM' nel nome del mese per esteso."""
    try:
        _, mese = mese_anno.split("-")
        return MESI_ITALIANI.get(int(mese), mese_anno)
    except Exception:
        return mese_anno


@st.cache_data(ttl=60, show_spinner=False)
def leggi_foglio_tutti(_workbook):
    """Legge il foglio 'Tutti' leggendo le colonne per posizione."""
    try:
        ws = _workbook.worksheet(NOME_FOGLIO_TUTTI)
    except gspread.WorksheetNotFound:
        nomi_disponibili = ", ".join(f"'{f.title}'" for f in _workbook.worksheets())
        return None, (
            f"Il foglio '{NOME_FOGLIO_TUTTI}' non esiste nel documento collegato. "
            f"Fogli disponibili: {nomi_disponibili}."
        )
    except Exception as e:
        return None, f"Errore durante la lettura del foglio: {e}"

    tutti_i_valori = ws.get_all_values()
    righe_dati = tutti_i_valori[RIGA_INTESTAZIONE_TUTTI:]

    record = []
    ultima_colonna_utile = max(COL_TUTTI_NOME, COL_TUTTI_MESE, COL_TUTTI_TIPO_SERVIZIO,
                                COL_TUTTI_MINISTERO, COL_TUTTI_ORE, COL_TUTTI_CRED_ORE,
                                COL_TUTTI_STUDI, COL_TUTTI_OSSERVAZIONI)
    for i, r in enumerate(righe_dati):
        if len(r) <= ultima_colonna_utile:
            r = r + [""] * (ultima_colonna_utile + 1 - len(r))
        nome = r[COL_TUTTI_NOME].strip()
        if not nome:
            continue

        mese_anno = r[COL_TUTTI_MESE].strip()
        tipo_servizio = r[COL_TUTTI_TIPO_SERVIZIO].strip()
        ministero_testo = r[COL_TUTTI_MINISTERO].strip()
        ore_testo = r[COL_TUTTI_ORE].strip()
        cred_ore_testo = r[COL_TUTTI_CRED_ORE].strip()

        ha_partecipato = (ministero_testo.lower() in ("si", "sì")) or (a_float_it(ore_testo) > 0)
        pioniere_ausiliario = "pioniere ausiliario" in tipo_servizio.lower()

        riga_foglio = RIGA_INTESTAZIONE_TUTTI + 1 + i
        grezza = r[2:10]

        record.append({
            "Nome": nome,
            "Mese/Anno": mese_anno,
            "Anno di servizio": formatta_mese_esteso(mese_anno),
            "Ha partecipato al ministero": ha_partecipato,
            "Studi Biblici": r[COL_TUTTI_STUDI].strip(),
            "Pioniere ausiliario": pioniere_ausiliario,
            "Tipo Servizio": tipo_servizio,
            "Ore": ore_testo,
            "Cred. Ore": cred_ore_testo,
            "Osservazioni": r[COL_TUTTI_OSSERVAZIONI].strip(),
            "RigaFoglio": riga_foglio,
            "_grezza": grezza,
        })
    return pd.DataFrame(record), None


def anno_teocratico_di(mese_anno: str):
    """Dato un valore 'Mese/Anno' in formato AAAA-MM, ritorna l'anno teocratico."""
    try:
        anno_str, mese_str = mese_anno.split("-")
        anno, mese = int(anno_str), int(mese_str)
    except Exception:
        return None
    return anno if mese >= 9 else anno - 1


def _s21_y_da_top(top: float, offset: float = 0.0, alza: float = 8.0) -> float:
    return S21_PAGE_H - (top + offset + alza)


def _s21_y_da_bottom(bottom: float, offset: float = 0.0, alza: float = 1.5) -> float:
    return S21_PAGE_H - (bottom + offset - alza)


def _s21_centro_box(c: rl_canvas.Canvas, box: tuple, offset: float, testo: str = "X",
                     font_name: str = "Helvetica-Bold", font_size: float = 10.0, sposta: float = 0.0):
    x0, x1, top, bottom = box
    fattore_altezza_maiuscole = 0.717
    largo_testo = c.stringWidth(testo, font_name, font_size)
    x = (x0 + x1) / 2 - largo_testo / 2
    centro_verticale_top = (top + bottom) / 2 + offset
    baseline_top = centro_verticale_top + (font_size * fattore_altezza_maiuscole) / 2 + sposta
    y = S21_PAGE_H - baseline_top
    c.setFont(font_name, font_size)
    c.drawString(x, y, testo)


def _s21_testo_centrato_colonna(c: rl_canvas.Canvas, testo: str, col: tuple, top: float, bottom: float,
                                 offset: float, font_name: str = "Helvetica", font_size: float = 9.5,
                                 sposta: float = 0.0):
    x0, x1 = col
    largo_testo = c.stringWidth(testo, font_name, font_size)
    x = (x0 + x1) / 2 - largo_testo / 2
    c.setFont(font_name, font_size)
    c.drawString(x, _s21_y_da_top((top + bottom) / 2, offset, alza=font_size * 0.36 + sposta), testo)


def _s21_righe_anno_per_nome(df_tutti: pd.DataFrame, nome: str, anno_teocratico) -> dict:
    if anno_teocratico is None or df_tutti.empty:
        return {}
    righe_persona = df_tutti[df_tutti["Nome"].str.strip().str.lower() == nome.strip().lower()]
    righe_persona = righe_persona[righe_persona["Mese/Anno"].apply(anno_teocratico_di) == anno_teocratico]

    risultato = {}
    for _, r in righe_persona.iterrows():
        risultato[r["Anno di servizio"]] = {
            "ha_partecipato": bool(r["Ha partecipato al ministero"]),
            "pioniere_ausiliario": bool(r["Pioniere ausiliario"]),
            "studi": r["Studi Biblici"],
            "ore": r["Ore"],
            "cred_ore": r.get("Cred. Ore", ""),
            "osservazioni": r["Osservazioni"],
        }
    return risultato


def _s21_righe_anno_aggregate(df_tutti: pd.DataFrame, nomi: list, anno_teocratico,
                               etichetta_conteggio: str = "proclamatori") -> dict:
    if anno_teocratico is None or df_tutti.empty or not nomi:
        return {}
    nomi_lower = {n.strip().lower() for n in nomi if n and n.strip()}
    if not nomi_lower:
        return {}
    righe = df_tutti[df_tutti["Nome"].str.strip().str.lower().isin(nomi_lower)]
    righe = righe[righe["Mese/Anno"].apply(anno_teocratico_di) == anno_teocratico]
    if righe.empty:
        return {}

    risultato = {}
    for mese in S21_ORDINE_MESI:
        righe_mese = righe[righe["Anno di servizio"] == mese]
        if righe_mese.empty:
            continue
        conteggio_ministero = int(righe_mese["Ha partecipato al ministero"].astype(bool).sum())
        conteggio_ausiliario = int(righe_mese["Pioniere ausiliario"].astype(bool).sum())
        studi_tot = sum(a_float_it(v) for v in righe_mese["Studi Biblici"])
        ore_tot = sum(a_float_it(v) for v in righe_mese["Ore"])
        cred_tot = sum(a_float_it(v) for v in righe_mese["Cred. Ore"]) if "Cred. Ore" in righe_mese.columns else 0.0

        risultato[mese] = {
            "ha_partecipato": conteggio_ministero > 0,
            "pioniere_ausiliario": conteggio_ausiliario > 0,
            "studi": formatta_numero_it(studi_tot) if studi_tot else "",
            "ore": formatta_numero_it(ore_tot) if ore_tot else "",
            "cred_ore": formatta_numero_it(cred_tot) if cred_tot else "",
            "osservazioni": f"{conteggio_ministero} {etichetta_conteggio}" if conteggio_ministero else "",
        }
    return risultato


def _s21_righe_anno_aggregate_per_tipo(df_tutti: pd.DataFrame, anno_teocratico,
                                        parola_chiave_tipo: str, etichetta_conteggio: str) -> dict:
    if anno_teocratico is None or df_tutti.empty:
        return {}
    righe = df_tutti[df_tutti["Mese/Anno"].apply(anno_teocratico_di) == anno_teocratico]
    righe = righe[righe["Tipo Servizio"].str.lower().str.contains(parola_chiave_tipo, na=False, regex=True)]
    if righe.empty:
        return {}

    risultato = {}
    for mese in S21_ORDINE_MESI:
        righe_mese = righe[righe["Anno di servizio"] == mese]
        if righe_mese.empty:
            continue
        conteggio = len(righe_mese)
        conteggio_ministero = int(righe_mese["Ha partecipato al ministero"].astype(bool).sum())
        conteggio_ausiliario = int(righe_mese["Pioniere ausiliario"].astype(bool).sum())
        studi_tot = sum(a_float_it(v) for v in righe_mese["Studi Biblici"])
        ore_tot = sum(a_float_it(v) for v in righe_mese["Ore"])
        cred_tot = sum(a_float_it(v) for v in righe_mese["Cred. Ore"]) if "Cred. Ore" in righe_mese.columns else 0.0

        risultato[mese] = {
            "ha_partecipato": conteggio_ministero > 0,
            "pioniere_ausiliario": conteggio_ausiliario > 0,
            "studi": formatta_numero_it(studi_tot) if studi_tot else "",
            "ore": formatta_numero_it(ore_tot) if ore_tot else "",
            "cred_ore": formatta_numero_it(cred_tot) if cred_tot else "",
            "osservazioni": f"{conteggio} {etichetta_conteggio}" if conteggio else "",
        }
    return risultato


def _s21_dati_riepilogo(titolo: str) -> dict:
    return {
        "nome": titolo,
        "data_nascita": "",
        "data_battesimo": "",
        "sesso": "",
        "incarico": "",
        "tipo": "",
        "classe_spirituale": "",
    }


def _s21_nota_inattivo_dal(riga_anagrafica: dict) -> str:
    stato = categoria_stato_proclamatore(riga_anagrafica.get("Attivi / Inattivi", ""))
    if stato != "I":
        return ""
    data_inattivo = (riga_anagrafica.get("Inattivo dal") or riga_anagrafica.get("Dal") or "").strip()
    if not data_inattivo:
        return ""
    return f"Inattivo dal {data_inattivo}"


def _s21_con_nota_prima_riga(righe_anno: dict, nota: str) -> dict:
    if not nota:
        return righe_anno
    righe_anno = dict(righe_anno)
    prima_riga = dict(righe_anno.get("Settembre", {}))
    esistente = str(prima_riga.get("osservazioni") or "").strip()
    prima_riga["osservazioni"] = f"{nota} — {esistente}" if esistente else nota
    righe_anno["Settembre"] = prima_riga
    return righe_anno


def anni_teocratici_per_menu(df_tutti: pd.DataFrame) -> list:
    anni = set()
    if not df_tutti.empty and "Mese/Anno" in df_tutti.columns:
        anni |= {a for a in df_tutti["Mese/Anno"].apply(anno_teocratico_di) if a is not None}
    if not anni:
        oggi = datetime.now()
        anni = {oggi.year - 1, oggi.year}
    return sorted(anni, reverse=True)


def _s21_disegna_pannello(c: rl_canvas.Canvas, offset: float, dati: dict, righe_anno: dict,
                           anno_teocratico=None, mostra_equazione_crediti: bool = True):
    c.setFillColorRGB(*S21_COLORE_NERO)
    c.setFont("Helvetica", S21_FONT_VALORI)
    c.drawString(116, _s21_y_da_bottom(51.5, offset), dati.get("nome", ""))

    data_nascita_str = dati.get("data_nascita", "")
    c.drawString(104, _s21_y_da_bottom(65.9, offset), data_nascita_str)
    eta = calcola_eta_dettagliata(data_nascita_str)
    if eta:
        c.setFillColorRGB(*S21_COLORE_ROSSO)
        c.setFont("Helvetica-Bold", S21_FONT_ETA)
        larghezza_data = c.stringWidth(data_nascita_str, "Helvetica", S21_FONT_VALORI)
        c.drawString(104 + larghezza_data + 6, _s21_y_da_bottom(65.9, offset), f"({eta})")
        c.setFillColorRGB(*S21_COLORE_NERO)

    data_battesimo_str = dati.get("data_battesimo", "")
    c.setFont("Helvetica", S21_FONT_VALORI)
    c.drawString(125, _s21_y_da_bottom(80.4, offset), data_battesimo_str)
    eta_batt = calcola_eta_dettagliata(data_battesimo_str)
    if eta_batt:
        c.setFillColorRGB(*S21_COLORE_ROSSO)
        c.setFont("Helvetica-Bold", S21_FONT_ETA)
        larghezza_data_batt = c.stringWidth(data_battesimo_str, "Helvetica", S21_FONT_VALORI)
        c.drawString(125 + larghezza_data_batt + 6, _s21_y_da_bottom(80.4, offset), f"({eta_batt})")
        c.setFillColorRGB(*S21_COLORE_NERO)

    if dati.get("sesso") == "M":
        _s21_centro_box(c, S21_BOX_SESSO_M, offset, font_size=S21_FONT_CHECK_HEADER, sposta=S21_SPOSTAMENTO_TESTATA)
    elif dati.get("sesso") == "F":
        _s21_centro_box(c, S21_BOX_SESSO_F, offset, font_size=S21_FONT_CHECK_HEADER, sposta=S21_SPOSTAMENTO_TESTATA)

    classe_spirituale = (dati.get("classe_spirituale") or "").strip().upper()
    if classe_spirituale in ("U", "UNTO"):
        _s21_centro_box(c, S21_BOX_UNTO, offset, font_size=S21_FONT_CHECK_HEADER, sposta=S21_SPOSTAMENTO_TESTATA)
    elif classe_spirituale in ("AP", "A", "ALTRE PECORE"):
        _s21_centro_box(c, S21_BOX_ALTRE_PECORE, offset, font_size=S21_FONT_CHECK_HEADER, sposta=S21_SPOSTAMENTO_TESTATA)

    incarico = dati.get("incarico", "")
    tipo = dati.get("tipo", "")
    if incarico == "Anziano":
        _s21_centro_box(c, S21_BOX_ANZIANO, offset, font_size=S21_FONT_CHECK_HEADER, sposta=S21_SPOSTAMENTO_TESTATA)
    if incarico == "Servitore di ministero":
        _s21_centro_box(c, S21_BOX_SERVITORE, offset, font_size=S21_FONT_CHECK_HEADER, sposta=S21_SPOSTAMENTO_TESTATA)
    if tipo == "Pioniere Regolare":
        _s21_centro_box(c, S21_BOX_PIONIERE_REGOLARE, offset, font_size=S21_FONT_CHECK_HEADER, sposta=S21_SPOSTAMENTO_TESTATA)
    if tipo == "Pioniere speciale":
        _s21_centro_box(c, S21_BOX_PIONIERE_SPECIALE, offset, font_size=S21_FONT_CHECK_HEADER, sposta=S21_SPOSTAMENTO_TESTATA)
    if tipo == "Missionario sul campo":
        _s21_centro_box(c, S21_BOX_MISSIONARIO, offset, font_size=S21_FONT_CHECK_HEADER, sposta=S21_SPOSTAMENTO_TESTATA)

    c.setFillColorRGB(*S21_COLORE_NERO)

    if anno_teocratico is not None:
        etichetta_anno = f"{anno_teocratico}-{anno_teocratico + 1}"
        c.setFont("Helvetica-Bold", S21_FONT_TABELLA)
        largo_anno = c.stringWidth(etichetta_anno, "Helvetica-Bold", S21_FONT_TABELLA)
        x0_col, x1_col = S21_COL_ANNO_SERVIZIO
        x_anno = (x0_col + x1_col) / 2 - largo_anno / 2
        c.drawString(x_anno, _s21_y_da_top(S21_ANNO_LABEL_TOP, offset, alza=S21_FONT_TABELLA * 0.36),
                     etichetta_anno)

    totale_ore = 0.0
    totale_cred_ore = 0.0
    for mese in S21_ORDINE_MESI:
        riga = righe_anno.get(mese)
        if not riga:
            continue
        top, bottom = S21_RIGHE[mese]
        if riga.get("ha_partecipato") is True:
            _s21_centro_box(c, (*S21_COL_MINISTERO, top, bottom), offset,
                             font_size=S21_FONT_CHECK_TABELLA, sposta=S21_SPOSTAMENTO_RIGHE)
        if riga.get("pioniere_ausiliario") is True:
            _s21_centro_box(c, (*S21_COL_AUSILIARIO, top, bottom), offset,
                             font_size=S21_FONT_CHECK_TABELLA, sposta=S21_SPOSTAMENTO_RIGHE)

        studi_val = str(riga.get("studi") or "").strip()
        if studi_val and studi_val != "0":
            _s21_testo_centrato_colonna(c, studi_val, S21_COL_STUDI, top, bottom, offset,
                                         font_size=S21_FONT_TABELLA, sposta=S21_SPOSTAMENTO_RIGHE)

        ore_val = str(riga.get("ore") or "").strip()
        cred_ore_val = str(riga.get("cred_ore") or "").strip()
        if ore_val and ore_val != "0":
            _s21_testo_centrato_colonna(c, ore_val, S21_COL_ORE, top, bottom, offset,
                                         font_size=S21_FONT_TABELLA, sposta=S21_SPOSTAMENTO_RIGHE)
            totale_ore += a_float_it(ore_val)
        if cred_ore_val and cred_ore_val != "0":
            totale_cred_ore += a_float_it(cred_ore_val)

        osservazioni_val = str(riga.get("osservazioni") or "").strip()
        if osservazioni_val:
            c.setFont("Helvetica", S21_FONT_TABELLA)
            c.drawString(S21_COL_OSSERVAZIONI_X, _s21_y_da_top((top + bottom) / 2, offset,
                                                                 alza=S21_FONT_TABELLA * 0.36 + S21_SPOSTAMENTO_RIGHE),
                         osservazioni_val[:44])

    top_tot, bottom_tot = S21_TOTALE_RIGA
    if totale_ore:
        _s21_testo_centrato_colonna(c, formatta_numero_it(totale_ore), S21_COL_ORE, top_tot, bottom_tot, offset,
                                     font_name="Helvetica-Bold", font_size=S21_FONT_TABELLA,
                                     sposta=S21_SPOSTAMENTO_RIGHE)
    if totale_cred_ore and mostra_equazione_crediti:
        totale_finale = totale_ore + totale_cred_ore
        testo_crediti = (f"{formatta_numero_it(totale_ore)} + ({formatta_numero_it(totale_cred_ore)}) "
                          f"= {formatta_numero_it(totale_finale)}")
        c.setFont("Helvetica", S21_FONT_TABELLA)
        c.drawString(S21_COL_OSSERVAZIONI_X, _s21_y_da_top((top_tot + bottom_tot) / 2, offset,
                                                             alza=S21_FONT_TABELLA * 0.36 + S21_SPOSTAMENTO_RIGHE),
                     testo_crediti)


def _s21_dati_da_riga_anagrafica(riga: dict) -> dict:
    return {
        "nome": riga.get("Cognome e Nome", ""),
        "data_nascita": riga.get("Data Nascita", ""),
        "data_battesimo": riga.get("Data Battesimo", ""),
        "sesso": (riga.get("Sesso", "") or "").strip().upper()[:1],
        "incarico": riga.get("Incarico", ""),
        "tipo": riga.get("Tipo", ""),
        "classe_spirituale": riga.get("A/U", ""),
    }


def genera_pdf_s21_singolo(riga_anagrafica: dict, df_tutti: pd.DataFrame, anno_corrente: int) -> bytes:
    nome = riga_anagrafica.get("Cognome e Nome", "")
    dati = _s21_dati_da_riga_anagrafica(riga_anagrafica)

    anno_precedente = anno_corrente - 1

    righe_corrente = _s21_righe_anno_per_nome(df_tutti, nome, anno_corrente)
    righe_precedente = _s21_righe_anno_per_nome(df_tutti, nome, anno_precedente)

    nota_inattivo = _s21_nota_inattivo_dal(riga_anagrafica)
    righe_corrente = _s21_con_nota_prima_riga(righe_corrente, nota_inattivo)
    righe_precedente = _s21_con_nota_prima_riga(righe_precedente, nota_inattivo)

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=(S21_PAGE_W, S21_PAGE_H))
    _s21_disegna_pannello(c, 0.0, dati, righe_precedente, anno_teocratico=anno_precedente)
    _s21_disegna_pannello(c, S21_OFFSET_PANNELLO, dati, righe_corrente, anno_teocratico=anno_corrente)
    c.save()
    buf.seek(0)

    overlay_reader = PdfReader(buf)
    template_reader = PdfReader(PERCORSO_MODULO_S21)
    writer = PdfWriter()

    pagina = template_reader.pages[0]
    pagina.merge_page(overlay_reader.pages[0])
    writer.add_page(pagina)

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def genera_pdf_s21_multiplo(righe_anagrafica: list, df_tutti: pd.DataFrame, anno_corrente: int) -> bytes:
    writer = PdfWriter()
    template_reader = PdfReader(PERCORSO_MODULO_S21)

    for riga_anagrafica in righe_anagrafica:
        nome = riga_anagrafica.get("Cognome e Nome", "")
        dati = _s21_dati_da_riga_anagrafica(riga_anagrafica)

        anno_precedente = anno_corrente - 1

        righe_corrente = _s21_righe_anno_per_nome(df_tutti, nome, anno_corrente)
        righe_precedente = _s21_righe_anno_per_nome(df_tutti, nome, anno_precedente)

        nota_inattivo = _s21_nota_inattivo_dal(riga_anagrafica)
        righe_corrente = _s21_con_nota_prima_riga(righe_corrente, nota_inattivo)
        righe_precedente = _s21_con_nota_prima_riga(righe_precedente, nota_inattivo)

        buf = io.BytesIO()
        c = rl_canvas.Canvas(buf, pagesize=(S21_PAGE_W, S21_PAGE_H))
        _s21_disegna_pannello(c, 0.0, dati, righe_precedente, anno_teocratico=anno_precedente)
        _s21_disegna_pannello(c, S21_OFFSET_PANNELLO, dati, righe_corrente, anno_teocratico=anno_corrente)
        c.save()
        buf.seek(0)

        overlay_reader = PdfReader(buf)
        pagina = template_reader.pages[0]
        template_fresh = PdfReader(PERCORSO_MODULO_S21)
        pagina_overlay = template_fresh.pages[0]
        pagina_overlay.merge_page(overlay_reader.pages[0])
        writer.add_page(pagina_overlay)

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _s21_nome_file_sicuro(nome: str) -> str:
    nome_pulito = "".join(c for c in nome if c not in '\\/:*?"<>|').strip()
    return nome_pulito or "Senza_nome"


# ─────────────────────────────────────────────────────────────────
# IMPORTAZIONE S-21 RICEVUTA (da altra congregazione)
# ─────────────────────────────────────────────────────────────────

def _s21_ripulisci_data(testo: str) -> str:
    return re.sub(r"\s*\([\d,\.]+\)\s*$", "", testo or "").strip()


def _s21_estrai_dati_pdf(sorgente) -> dict:
    pannelli = []
    almeno_un_valore = False

    with pdfplumber.open(sorgente) as pdf:
        pagina = pdf.pages[0]

        def testo_in(x0: float, x1: float, top: float, bottom: float) -> str:
            nonlocal almeno_un_valore
            top_r = max(0.0, min(S21_PAGE_H, top))
            bottom_r = max(0.0, min(S21_PAGE_H, bottom))
            x0_r = max(0.0, min(S21_PAGE_W, x0))
            x1_r = max(0.0, min(S21_PAGE_W, x1))
            if top_r >= bottom_r or x0_r >= x1_r:
                return ""
            ritaglio = pagina.crop((x0_r, top_r, x1_r, bottom_r))
            testo_grezzo = (ritaglio.extract_text() or "").strip()
            testo = re.sub(r"\(cid:\d+\)", "", testo_grezzo).strip()
            if testo:
                almeno_un_valore = True
            return testo

        def casella_marcata(box: tuple, offset: float) -> bool:
            x0, x1, top, bottom = box
            return bool(testo_in(x0, x1, top + offset, bottom + offset))

        for offset in (0.0, S21_OFFSET_PANNELLO):
            nome = testo_in(116, 380, 51.5 - 9 + offset, 51.5 + 3 + offset)
            data_nascita = _s21_ripulisci_data(testo_in(104, 260, 65.9 - 9 + offset, 65.9 + 3 + offset))
            data_battesimo = _s21_ripulisci_data(testo_in(125, 260, 80.4 - 9 + offset, 80.4 + 3 + offset))
            anno_testo = testo_in(*S21_COL_ANNO_SERVIZIO, S21_ANNO_LABEL_TOP - 9 + offset,
                                   S21_ANNO_LABEL_TOP + 3 + offset)

            sesso = "M" if casella_marcata(S21_BOX_SESSO_M, offset) else (
                "F" if casella_marcata(S21_BOX_SESSO_F, offset) else "")
            classe_spirituale = "Unto" if casella_marcata(S21_BOX_UNTO, offset) else (
                "Altre pecore" if casella_marcata(S21_BOX_ALTRE_PECORE, offset) else "")
            incarico = "Anziano" if casella_marcata(S21_BOX_ANZIANO, offset) else (
                "Servitore di ministero" if casella_marcata(S21_BOX_SERVITORE, offset) else "")
            if casella_marcata(S21_BOX_PIONIERE_REGOLARE, offset):
                tipo = "Pioniere Regolare"
            elif casella_marcata(S21_BOX_PIONIERE_SPECIALE, offset):
                tipo = "Pioniere speciale"
            elif casella_marcata(S21_BOX_MISSIONARIO, offset):
                tipo = "Missionario sul campo"
            else:
                tipo = ""

            mesi = {}
            for mese in S21_ORDINE_MESI:
                top, bottom = S21_RIGHE[mese]
                ministero = casella_marcata((*S21_COL_MINISTERO, top, bottom), offset)
                ausiliario = casella_marcata((*S21_COL_AUSILIARIO, top, bottom), offset)
                studi = testo_in(*S21_COL_STUDI, top + offset, bottom + offset)
                ore = testo_in(*S21_COL_ORE, top + offset, bottom + offset)
                osservazioni = testo_in(S21_COL_OSSERVAZIONI_X, S21_PAGE_W - 12, top + offset, bottom + offset)
                if ministero or ausiliario or studi or ore or osservazioni:
                    mesi[mese] = {
                        "ministero": ministero, "ausiliario": ausiliario,
                        "studi": studi, "ore": ore, "osservazioni": osservazioni,
                    }

            pannelli.append({
                "anno_testo": anno_testo, "nome": nome,
                "data_nascita": data_nascita, "data_battesimo": data_battesimo,
                "sesso": sesso, "classe_spirituale": classe_spirituale,
                "incarico": incarico, "tipo": tipo, "mesi": mesi,
            })

    return {"testo_rilevato": almeno_un_valore, "pannelli": pannelli}


def _s21_anno_da_testo(anno_testo: str):
    m = re.search(r"(20\d{2})", anno_testo or "")
    return int(m.group(1)) if m else None


def _s21_tipo_servizio_mese(tipo_panello: str, ministero: bool, ausiliario: bool) -> str:
    if tipo_panello == "Pioniere Regolare":
        return "Pioniere Regolare"
    if tipo_panello == "Pioniere speciale":
        return "Pioniere Speciale"
    if tipo_panello == "Missionario sul campo":
        return "Missionario sul campo"
    if ausiliario:
        return "Pioniere Ausiliario"
    return "Proclamatore"


def _s21_costruisci_righe_import(dati_estratti: dict, mesi_gia_presenti: set) -> list:
    righe = []
    for pannello in dati_estratti["pannelli"]:
        anno = _s21_anno_da_testo(pannello["anno_testo"])
        for mese in S21_ORDINE_MESI:
            dati_mese = pannello["mesi"].get(mese)
            if not dati_mese:
                continue
            if anno is not None:
                indice_mese = S21_ORDINE_MESI.index(mese)
                mese_num = 9 + indice_mese if indice_mese <= 3 else indice_mese - 3
                anno_calendario = anno if mese_num >= 9 else anno + 1
                mese_anno = f"{anno_calendario}-{mese_num:02d}"
            else:
                mese_anno = ""

            tipo_servizio = _s21_tipo_servizio_mese(pannello["tipo"], dati_mese["ministero"],
                                                     dati_mese["ausiliario"])
            righe.append({
                "Mese (dal modulo)": mese,
                "Anno servizio letto": pannello["anno_testo"] or "(non letto)",
                "Mese/Anno": mese_anno,
                "Tipo Servizio": tipo_servizio,
                "Ha partecipato al ministero": "Si" if dati_mese["ministero"] else "",
                "Ore": dati_mese["ore"],
                "Studi Biblici": dati_mese["studi"],
                "Osservazioni": dati_mese["osservazioni"],
                "gia_presente": bool(mese_anno) and mese_anno in mesi_gia_presenti,
                "Importa": bool(mese_anno) and mese_anno not in mesi_gia_presenti,
            })
    return righe


def aggiungi_riga_tutti(_workbook, nome: str, mese_anno: str, tipo_servizio: str,
                         ministero: str, ore: str, studi: str, osservazioni: str):
    try:
        ws = _workbook.worksheet(NOME_FOGLIO_TUTTI)
        riga = [""] * 10
        riga[COL_TUTTI_NOME] = nome
        riga[COL_TUTTI_MESE] = mese_anno
        riga[COL_TUTTI_TIPO_SERVIZIO] = tipo_servizio
        riga[COL_TUTTI_MINISTERO] = ministero
        riga[COL_TUTTI_ORE] = ore
        riga[COL_TUTTI_STUDI] = studi
        riga[COL_TUTTI_OSSERVAZIONI] = osservazioni
        ws.append_row(riga, value_input_option="USER_ENTERED")
        return True, None
    except Exception as e:
        return False, f"Errore durante il salvataggio: {e}"


def _importa_s21_cerca_persona(df_anagrafica: pd.DataFrame, nome_letto: str) -> pd.DataFrame:
    if df_anagrafica.empty or not nome_letto:
        return df_anagrafica.iloc[0:0]
    parole = [p.lower() for p in nome_letto.split() if len(p) > 2]
    if not parole:
        return df_anagrafica.iloc[0:0]

    def punteggio(nome_anagrafica: str) -> int:
        n = nome_anagrafica.lower()
        return sum(1 for p in parole if p in n)

    df = df_anagrafica.copy()
    df["_punteggio"] = df["Cognome e Nome"].apply(punteggio)
    return df[df["_punteggio"] > 0].sort_values("_punteggio", ascending=False).drop(columns="_punteggio")


def _s21_anno_cartella_corrente() -> str:
    oggi = datetime.now()
    anno_teo = anno_teocratico_di(f"{oggi.year}-{oggi.month:02d}")
    return str(anno_teo + 1) if anno_teo is not None else str(oggi.year)


def _s21_cartella_per_riga(riga: dict) -> str:
    stato = categoria_stato_proclamatore(riga.get("Attivi / Inattivi", ""))
    if stato == "I":
        return "Inattivi"
    if stato == "TR":
        return "Trasferiti"

    tipo = (riga.get("Tipo") or "").strip()
    if tipo == "Pioniere Regolare":
        return "Attivi/Pionieri Regolari"
    if tipo == "Pioniere speciale":
        return "Attivi/Pionieri Speciali"
    if tipo == "Missionario sul campo":
        return "Attivi/Missionari sul campo"

    gruppo = (riga.get("Gruppo") or "").strip() or "(Senza gruppo)"
    return f"Attivi/Proclamatori/{gruppo}"


def genera_zip_s21(righe_anagrafica: list, df_tutti: pd.DataFrame, anno_corrente: int) -> bytes:
    anno_cartella = anno_corrente + 1
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for riga in righe_anagrafica:
            nome = (riga.get("Cognome e Nome") or "").strip()
            if not nome:
                continue
            pdf_bytes = genera_pdf_s21_singolo(riga, df_tutti, anno_corrente)
            sotto_cartella = _s21_cartella_per_riga(riga)
            nome_file = _s21_nome_file_sicuro(nome) + ".pdf"
            percorso = f"Anno {anno_cartella}/{sotto_cartella}/{nome_file}"
            zf.writestr(percorso, pdf_bytes)
    buf.seek(0)
    return buf.getvalue()


def genera_pdf_s21_riepilogo(titolo: str, nomi: list, df_tutti: pd.DataFrame, anno_corrente: int,
                              parola_chiave_tipo: str = None, etichetta_conteggio: str = "proclamatori") -> bytes:
    anno_precedente = anno_corrente - 1
    dati = _s21_dati_riepilogo(titolo)
    if parola_chiave_tipo:
        righe_corrente = _s21_righe_anno_aggregate_per_tipo(df_tutti, anno_corrente,
                                                              parola_chiave_tipo, etichetta_conteggio)
        righe_precedente = _s21_righe_anno_aggregate_per_tipo(df_tutti, anno_precedente,
                                                                parola_chiave_tipo, etichetta_conteggio)
    else:
        righe_corrente = _s21_righe_anno_aggregate(df_tutti, nomi, anno_corrente, etichetta_conteggio)
        righe_precedente = _s21_righe_anno_aggregate(df_tutti, nomi, anno_precedente, etichetta_conteggio)

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=(S21_PAGE_W, S21_PAGE_H))
    _s21_disegna_pannello(c, 0.0, dati, righe_precedente, anno_teocratico=anno_precedente,
                          mostra_equazione_crediti=False)
    _s21_disegna_pannello(c, S21_OFFSET_PANNELLO, dati, righe_corrente, anno_teocratico=anno_corrente,
                          mostra_equazione_crediti=False)
    c.save()
    buf.seek(0)

    overlay_reader = PdfReader(buf)
    template_reader = PdfReader(PERCORSO_MODULO_S21)
    writer = PdfWriter()
    pagina = template_reader.pages[0]
    pagina.merge_page(overlay_reader.pages[0])
    writer.add_page(pagina)

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def genera_zip_s21_completo(df: pd.DataFrame, df_tutti: pd.DataFrame, anno_corrente: int) -> bytes:
    anno_cartella = anno_corrente + 1
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for _, riga in df.iterrows():
            riga_dict = riga.to_dict()
            nome = (riga_dict.get("Cognome e Nome") or "").strip()
            if not nome:
                continue
            if categoria_stato_proclamatore(riga_dict.get("Attivi / Inattivi", "")) == "TR":
                continue
            pdf_bytes = genera_pdf_s21_singolo(riga_dict, df_tutti, anno_corrente)
            sotto_cartella = _s21_cartella_per_riga(riga_dict)
            nome_file = _s21_nome_file_sicuro(nome) + ".pdf"
            zf.writestr(f"Anno {anno_cartella}/{sotto_cartella}/{nome_file}", pdf_bytes)

        riepiloghi = [
            ("Tutti i proclamatori", "Riepilogo Tutti i Proclamatori.pdf", "proclamatore", "proclamatori"),
            ("Tutti i pionieri regolari", "Riepilogo Pionieri Regolari.pdf",
             "pioniere regolare", "pionieri regolari"),
            ("Tutti i pionieri speciali", "Riepilogo Pionieri Speciali.pdf",
             "pioniere speciale", "pionieri speciali"),
            ("Tutti i missionari sul campo", "Riepilogo Missionari sul Campo.pdf",
             "missionario|rappresentante", "missionari sul campo"),
            ("Tutti i pionieri ausiliari", "Riepilogo Pionieri Ausiliari.pdf",
             "pioniere ausiliario", "ausiliari"),
        ]
        for titolo, nome_file, parola_chiave_tipo, etichetta in riepiloghi:
            pdf_bytes = genera_pdf_s21_riepilogo(titolo, [], df_tutti, anno_corrente,
                                                  parola_chiave_tipo=parola_chiave_tipo,
                                                  etichetta_conteggio=etichetta)
            zf.writestr(f"Anno {anno_cartella}/{nome_file}", pdf_bytes)

    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────
# Pagina: RIEPILOGO ATTIVITÀ (report libero per sorveglianti di gruppo/categoria)
# ─────────────────────────────────────────────────────────────────
CATEGORIE_RIEPILOGO_ATTIVITA = {
    "Tutti": None,
    "Solo Proclamatori": "proclamatore",
    "Solo Pionieri Ausiliari": "pioniere ausiliario",
    "Solo Pionieri Regolari": "pioniere regolare",
    "Solo Pionieri Speciali": "pioniere speciale",
    "Solo Missionari sul campo": "missionario|rappresentante",
}


def _riepilogo_ultimo_mese_con_dati(df_tutti: pd.DataFrame):
    if df_tutti.empty or "Mese/Anno" not in df_tutti.columns:
        return None

    condizione_dati = (
        (df_tutti["Ha partecipato al ministero"] == True) |
        (df_tutti["Ha partecipato al ministero"].astype(str).str.lower().isin(["sì", "si", "1", "true"])) |
        (df_tutti["Ore"].notna() & (df_tutti["Ore"].astype(str).str.strip() != "")) |
        (df_tutti["Studi Biblici"].notna() & (df_tutti["Studi Biblici"].astype(str).str.strip() != ""))
    )

    df_con_dati = df_tutti[condizione_dati]

    if df_con_dati.empty:
        df_con_dati = df_tutti

    validi = []
    for m in df_con_dati["Mese/Anno"].dropna().unique():
        try:
            a, mm = str(m).strip().split("-")
            validi.append((int(a), int(mm)))
        except Exception:
            continue

    if not validi:
        return None

    return max(validi)


def _riepilogo_finestra_ultimi_n_mesi(anno_fine: int, mese_fine: int, n: int = 6) -> set:
    risultato = set()
    a, m = anno_fine, mese_fine
    for _ in range(n):
        risultato.add((a, m))
        m -= 1
        if m == 0:
            m = 12
            a -= 1
    return risultato


ETICHETTE_SINTETICO_CATEGORIA = {
    "Tutti": "Tutti i proclamatori",
    "Solo Proclamatori": "Proclamatori",
    "Solo Pionieri Ausiliari": "Pionieri Ausiliari",
    "Solo Pionieri Regolari": "Pionieri Regolari",
    "Solo Pionieri Speciali": "Pionieri Speciali",
    "Solo Missionari sul campo": "Missionari sul campo",
}


def _riepilogo_mesi_nel_periodo(df_tutti: pd.DataFrame, periodo: str) -> list:
    if df_tutti.empty or "Mese/Anno" not in df_tutti.columns:
        return []
    df = df_tutti.copy()
    if periodo == "12 mesi":
        oggi = datetime.now()
        anno_teo_corrente = anno_teocratico_di(f"{oggi.year}-{oggi.month:02d}")
        df = df[df["Mese/Anno"].apply(lambda m: anno_teocratico_di(m) == anno_teo_corrente)]
    elif periodo == "6 mesi":
        ultimo = _riepilogo_ultimo_mese_con_dati(df_tutti)
        if not ultimo:
            return []
        finestra = _riepilogo_finestra_ultimi_n_mesi(ultimo[0], ultimo[1], 6)

        def _dentro(mese_anno):
            try:
                a, m = str(mese_anno).split("-")
                return (int(a), int(m)) in finestra
            except Exception:
                return False

        df = df[df["Mese/Anno"].apply(_dentro)]
    return sorted({str(m) for m in df["Mese/Anno"].dropna().unique() if str(m).strip()})


def _riepilogo_etichetta_dati_estratti(df_tutti: pd.DataFrame, periodo: str):
    mesi = _riepilogo_mesi_nel_periodo(df_tutti, periodo)
    if not mesi:
        return None
    n = len(mesi)

    def _mese_anno_esteso(mese_anno: str) -> str:
        try:
            a, m = mese_anno.split("-")
            return f"{MESI_ITALIANI.get(int(m), m)} {a}"
        except Exception:
            return mese_anno

    primo, ultimo = _mese_anno_esteso(mesi[0]), _mese_anno_esteso(mesi[-1])
    if primo == ultimo:
        return f"Dati estratti (1 mese): {primo}"
    return f"Dati estratti ({n} mesi): {primo} – {ultimo}"


def _riepilogo_filtra_dati(df_tutti: pd.DataFrame, df_anagrafica: pd.DataFrame, periodo: str,
                            gruppo_scelto: str, categoria: str) -> pd.DataFrame:
    df = df_tutti.copy()
    if df.empty:
        return df

    if periodo == "12 mesi":
        oggi = datetime.now()
        anno_teo_corrente = anno_teocratico_di(f"{oggi.year}-{oggi.month:02d}")

        def _dentro_anno_corrente(mese_anno):
            return anno_teocratico_di(mese_anno) == anno_teo_corrente

        df = df[df["Mese/Anno"].apply(_dentro_anno_corrente)]
    elif periodo == "6 mesi":
        ultimo = _riepilogo_ultimo_mese_con_dati(df_tutti)
        if not ultimo:
            return df.iloc[0:0]
        finestra = _riepilogo_finestra_ultimi_n_mesi(ultimo[0], ultimo[1], 6)

        def _dentro_finestra(mese_anno):
            try:
                a, m = str(mese_anno).split("-")
                return (int(a), int(m)) in finestra
            except Exception:
                return False

        df = df[df["Mese/Anno"].apply(_dentro_finestra)]

    parola_categoria = CATEGORIE_RIEPILOGO_ATTIVITA.get(categoria)
    if parola_categoria:
        df = df[df["Tipo Servizio"].str.lower().str.contains(parola_categoria, na=False, regex=True)]

    if gruppo_scelto and gruppo_scelto != "Tutti i gruppi" and "Gruppo" in df_anagrafica.columns:
        nomi_gruppo = set(
            df_anagrafica.loc[df_anagrafica["Gruppo"].astype(str).str.strip() == gruppo_scelto, "Cognome e Nome"]
            .astype(str).str.strip()
        )
        df = df[df["Nome"].str.strip().isin(nomi_gruppo)]

    return df


def _riepilogo_costruisci_blocchi_dettagliato(df_filtrato: pd.DataFrame) -> list:
    if df_filtrato.empty:
        return []

    blocchi = []
    for nome, gruppo_df in df_filtrato.groupby("Nome"):
        gruppo_df = gruppo_df.sort_values("Mese/Anno")
        righe = []
        tot_ore = tot_cred = tot_studi = 0.0
        for _, r in gruppo_df.iterrows():
            ore_val = a_float_it(r.get("Ore", ""))
            cred_val = a_float_it(r.get("Cred. Ore", ""))
            studi_val = a_float_it(r.get("Studi Biblici", ""))
            tot_ore += ore_val
            tot_cred += cred_val
            tot_studi += studi_val
            righe.append({
                "mese_anno": r.get("Mese/Anno", ""),
                "tipo": r.get("Tipo Servizio", "") or "",
                "ministero": "Sì" if r.get("Ha partecipato al ministero") else "No",
                "ore": formatta_numero_it(ore_val) if ore_val else "",
                "crediti": formatta_numero_it(cred_val) if cred_val else "",
                "studi": formatta_numero_it(studi_val) if studi_val else "",
                "note": r.get("Osservazioni", "") or "",
            })
        n = len(righe)
        blocchi.append({
            "nome": nome,
            "righe": righe,
            "totale_ore": tot_ore,
            "totale_crediti": tot_cred,
            "totale_studi": tot_studi,
            "media_ore": tot_ore / n if n else 0.0,
            "media_crediti": tot_cred / n if n else 0.0,
            "media_studi": tot_studi / n if n else 0.0,
        })
    blocchi.sort(key=lambda b: b["nome"])
    return blocchi


def _riepilogo_costruisci_blocco_sintetico(df_filtrato: pd.DataFrame, categoria: str) -> list:
    if df_filtrato.empty:
        return []

    etichetta = ETICHETTE_SINTETICO_CATEGORIA.get(categoria, categoria)
    righe = []
    tot_ore = tot_cred = tot_studi = 0.0
    for mese_anno, gruppo_mese in df_filtrato.groupby("Mese/Anno"):
        ore_val = sum(a_float_it(v) for v in gruppo_mese.get("Ore", []))
        cred_val = sum(a_float_it(v) for v in gruppo_mese.get("Cred. Ore", []))
        studi_val = sum(a_float_it(v) for v in gruppo_mese.get("Studi Biblici", []))
        conteggio = len(gruppo_mese)
        tot_ore += ore_val
        tot_cred += cred_val
        tot_studi += studi_val
        righe.append({
            "mese_anno": mese_anno,
            "tipo": etichetta,
            "ministero": "",
            "ore": formatta_numero_it(ore_val) if ore_val else "",
            "crediti": formatta_numero_it(cred_val) if cred_val else "",
            "studi": formatta_numero_it(studi_val) if studi_val else "",
            "note": f"{conteggio} {etichetta.lower()}" if conteggio else "",
        })

    def _chiave_ordinamento(riga):
        try:
            a, m = str(riga["mese_anno"]).split("-")
            return (int(a), int(m))
        except Exception:
            return (0, 0)

    righe.sort(key=_chiave_ordinamento)
    n_mesi = len(righe)
    return [{
        "nome": etichetta,
        "righe": righe,
        "totale_ore": tot_ore,
        "totale_crediti": tot_cred,
        "totale_studi": tot_studi,
        "media_ore": tot_ore / n_mesi if n_mesi else 0.0,
        "media_crediti": tot_cred / n_mesi if n_mesi else 0.0,
        "media_studi": tot_studi / n_mesi if n_mesi else 0.0,
    }]


def _riepilogo_totali_generali_per_categoria(df_periodo_gruppo: pd.DataFrame) -> list:
    if df_periodo_gruppo.empty:
        return []

    risultati = []
    for chiave, parola in CATEGORIE_RIEPILOGO_ATTIVITA.items():
        if chiave == "Tutti" or not parola:
            continue
        df_cat = df_periodo_gruppo[df_periodo_gruppo["Tipo Servizio"].str.lower().str.contains(
            parola, na=False, regex=True)]
        if df_cat.empty:
            continue
        tot_ore = sum(a_float_it(v) for v in df_cat.get("Ore", []))
        tot_cred = sum(a_float_it(v) for v in df_cat.get("Cred. Ore", []))
        tot_studi = sum(a_float_it(v) for v in df_cat.get("Studi Biblici", []))
        n = len(df_cat)
        risultati.append({
            "categoria": ETICHETTE_SINTETICO_CATEGORIA.get(chiave, chiave),
            "totale_ore": tot_ore,
            "totale_crediti": tot_cred,
            "totale_studi": tot_studi,
            "media_ore": tot_ore / n if n else 0.0,
            "media_crediti": tot_cred / n if n else 0.0,
            "media_studi": tot_studi / n if n else 0.0,
        })
    return risultati

def _riepilogo_totali_per_categoria_e_gruppo(df_tutti: pd.DataFrame, df_anagrafica: pd.DataFrame,
                                              periodo: str) -> list:
    df_periodo = _riepilogo_filtra_dati(df_tutti, df_anagrafica, periodo, "Tutti i gruppi", "Tutti")
    if df_periodo.empty or "Gruppo" not in df_anagrafica.columns:
        return []

    n_mesi_periodo = len(_riepilogo_mesi_nel_periodo(df_tutti, periodo))

    nomi_per_gruppo = {}
    for _, riga in df_anagrafica.iterrows():
        nome = str(riga.get("Cognome e Nome", "")).strip()
        gruppo = str(riga.get("Gruppo", "")).strip()
        if nome and gruppo and gruppo.lower() != "trasferiti":
            nomi_per_gruppo.setdefault(gruppo, set()).add(nome)

    risultati = []
    for chiave, parola in CATEGORIE_RIEPILOGO_ATTIVITA.items():
        if chiave == "Tutti" or not parola:
            continue
        df_cat = df_periodo[df_periodo["Tipo Servizio"].str.lower().str.contains(parola, na=False, regex=True)]
        if df_cat.empty:
            continue
        righe_gruppi = []
        for gruppo in sorted(nomi_per_gruppo.keys()):
            df_cat_gruppo = df_cat[df_cat["Nome"].str.strip().isin(nomi_per_gruppo[gruppo])]
            if df_cat_gruppo.empty:
                continue
            tot_ore = sum(a_float_it(v) for v in df_cat_gruppo.get("Ore", []))
            tot_cred = sum(a_float_it(v) for v in df_cat_gruppo.get("Cred. Ore", []))
            tot_studi = sum(a_float_it(v) for v in df_cat_gruppo.get("Studi Biblici", []))
            n = len(df_cat_gruppo)
            righe_gruppi.append({
                "gruppo": gruppo,
                "n_proclamatori": n,
                "n_proclamatori_media": n / n_mesi_periodo if n_mesi_periodo else 0.0,
                "totale_ore": tot_ore, "totale_crediti": tot_cred, "totale_studi": tot_studi,
                "media_ore": tot_ore / n if n else 0.0,
                "media_crediti": tot_cred / n if n else 0.0,
                "media_studi": tot_studi / n if n else 0.0,
            })
        if righe_gruppi:
            risultati.append({"categoria": ETICHETTE_SINTETICO_CATEGORIA.get(chiave, chiave),
                               "gruppi": righe_gruppi, "n_mesi_periodo": n_mesi_periodo})
    return risultati


def genera_pdf_riepilogo_attivita(blocchi: list, etichetta_periodo: str, etichetta_categoria: str,
                                   etichetta_gruppo: str = None, etichetta_vista: str = "Dettagliato",
                                   totali_per_categoria: list = None,
                                   comparazione_gruppi: list = None,
                                   etichetta_dati_periodo: str = None) -> bytes:
    buf = io.BytesIO()
    if comparazione_gruppi is not None:
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=0.6 * cm, bottomMargin=0.6 * cm,
                                 leftMargin=0.6 * cm, rightMargin=0.6 * cm)
    else:
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                                 leftMargin=1.3 * cm, rightMargin=1.3 * cm)
    stili = getSampleStyleSheet()
    elementi = []
    stile_sottotitolo = stili["Normal"]
    if comparazione_gruppi is not None:
        stile_sottotitolo = ParagraphStyle("SottotitoloCentrato", parent=stili["Normal"], alignment=1)

    elementi.append(Paragraph("Attività dei proclamatori", stili["Title"]))
    sottotitolo = (f"Congregazione: {NOME_CONGREGAZIONE} · Periodo: {etichetta_periodo} · "
                   f"{etichetta_vista} - {etichetta_categoria}")
    if etichetta_gruppo:
        sottotitolo += f" · Gruppo: {etichetta_gruppo}"
    elementi.append(Paragraph(sottotitolo, stile_sottotitolo))
    if etichetta_dati_periodo:
        elementi.append(Paragraph(etichetta_dati_periodo, stile_sottotitolo))
    elementi.append(Spacer(1, 14))

    if totali_per_categoria is not None:
        if not totali_per_categoria:
            elementi.append(Paragraph("Nessun dato trovato per i filtri selezionati.", stili["Normal"]))
        else:
            dati_tabella = [["", "Categoria", "Ore", "Crediti", "Studi"]]
            for cat in totali_per_categoria:
                dati_tabella.append(["Totale", cat["categoria"], formatta_numero_it(cat["totale_ore"]),
                                      formatta_numero_it(cat["totale_crediti"]),
                                      formatta_numero_it(cat["totale_studi"])])
                dati_tabella.append(["Media", cat["categoria"], formatta_numero_it(cat["media_ore"]),
                                      formatta_numero_it(cat["media_crediti"]),
                                      formatta_numero_it(cat["media_studi"])])
            tabella = Table(dati_tabella, colWidths=[2.2 * cm, 4.5 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm])
            stile_righe = [
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (2, 0), (4, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B6FA8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ]
            for i in range(1, len(dati_tabella), 2):
                colore = colors.HexColor("#F2F2F2") if ((i - 1) // 2) % 2 == 0 else colors.HexColor("#DCEEF9")
                stile_righe.append(("BACKGROUND", (0, i), (-1, i + 1), colore))
                stile_righe.append(("GRID", (0, i), (-1, i + 1), 0.4, colors.grey))
            tabella.setStyle(TableStyle(stile_righe))
            elementi.append(tabella)
        doc.build(elementi)
        buf.seek(0)
        return buf.getvalue()

    if comparazione_gruppi is not None:
        if not comparazione_gruppi:
            elementi.append(Paragraph("Nessun dato trovato per i filtri selezionati.", stili["Normal"]))
        larghezza_colonne_tabella = [2.1 * cm, 4.6 * cm, 1.1 * cm, 2.0 * cm, 2.0 * cm, 2.0 * cm]
        larghezza_utile_pagina = A4[0] - doc.leftMargin - doc.rightMargin
        rientro_categoria = max(0.0, (larghezza_utile_pagina - sum(larghezza_colonne_tabella)) / 2)
        stile_categoria = ParagraphStyle("CategoriaCompatta", parent=stili["Heading3"],
                                          fontSize=9, spaceBefore=3, spaceAfter=1, alignment=0,
                                          leftIndent=rientro_categoria)
        intestazione = ["", "Gruppo", "N.", "Ore", "Crediti", "Studi"]
        for blocco_cat in comparazione_gruppi:
            elementi.append(Paragraph(f"<b>{blocco_cat['categoria']}:</b>", stile_categoria))
            dati_tabella = [intestazione]
            n_totale = 0
            totale_ore_fin = totale_cred_fin = totale_studi_fin = 0.0
            n_mesi_periodo = blocco_cat.get("n_mesi_periodo", 0)
            for g in blocco_cat["gruppi"]:
                dati_tabella.append(["Totale", f"Gruppo {g['gruppo']}", str(g["n_proclamatori"]),
                                      formatta_numero_it(g["totale_ore"]),
                                      formatta_numero_it(g["totale_crediti"]),
                                      formatta_numero_it(g["totale_studi"])])
                dati_tabella.append(["Media", f"Gruppo {g['gruppo']}",
                                      formatta_numero_it(g["n_proclamatori_media"]),
                                      formatta_numero_it(g["media_ore"]),
                                      formatta_numero_it(g["media_crediti"]),
                                      formatta_numero_it(g["media_studi"])])
                n_totale += g["n_proclamatori"]
                totale_ore_fin += g["totale_ore"]
                totale_cred_fin += g["totale_crediti"]
                totale_studi_fin += g["totale_studi"]
            riga_finale = len(dati_tabella)
            dati_tabella.append(["Totale finale", "", str(n_totale), formatta_numero_it(totale_ore_fin),
                                  formatta_numero_it(totale_cred_fin), formatta_numero_it(totale_studi_fin)])
            dati_tabella.append(["Media finale", "",
                                  formatta_numero_it(n_totale / n_mesi_periodo if n_mesi_periodo else 0.0),
                                  formatta_numero_it(totale_ore_fin / n_totale if n_totale else 0.0),
                                  formatta_numero_it(totale_cred_fin / n_totale if n_totale else 0.0),
                                  formatta_numero_it(totale_studi_fin / n_totale if n_totale else 0.0)])
            tabella = Table(dati_tabella, colWidths=larghezza_colonne_tabella)
            stile_righe = [
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (2, 0), (5, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 1),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B6FA8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, riga_finale), (-1, -1), colors.HexColor("#FFE9B3")),
                ("GRID", (0, riga_finale), (-1, -1), 0.4, colors.grey),
            ]
            for i in range(1, riga_finale, 2):
                colore = colors.HexColor("#F2F2F2") if ((i - 1) // 2) % 2 == 0 else colors.HexColor("#DCEEF9")
                stile_righe.append(("BACKGROUND", (0, i), (-1, i + 1), colore))
                stile_righe.append(("GRID", (0, i), (-1, i + 1), 0.4, colors.grey))
            tabella.setStyle(TableStyle(stile_righe))
            tabella.hAlign = "CENTER"
            elementi.append(KeepTogether([tabella, Spacer(1, 3)]))
        doc.build(elementi)
        buf.seek(0)
        return buf.getvalue()

    if not blocchi:
        elementi.append(Paragraph("Nessun dato trovato per i filtri selezionati.", stili["Normal"]))

    intestazione = ["Mese", "Servizio", "Ministero", "Ore", "Crediti", "Studi", "Note"]
    larghezze = [1.9 * cm, 3.0 * cm, 1.7 * cm, 1.6 * cm, 1.6 * cm, 1.6 * cm, 4.7 * cm]

    for blocco in blocchi:
        dati_tabella = [intestazione]
        for r in blocco["righe"]:
            dati_tabella.append([r["mese_anno"], r["tipo"], r["ministero"], r["ore"], r["crediti"],
                                  r["studi"], r["note"]])
        dati_tabella.append(["Totale", "", "", formatta_numero_it(blocco["totale_ore"]),
                              formatta_numero_it(blocco["totale_crediti"]),
                              formatta_numero_it(blocco["totale_studi"]), ""])
        dati_tabella.append(["Media", "", "", formatta_numero_it(blocco["media_ore"]),
                              formatta_numero_it(blocco["media_crediti"]),
                              formatta_numero_it(blocco["media_studi"]), ""])

        tabella = Table(dati_tabella, colWidths=larghezze, repeatRows=1)
        tabella.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B6FA8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, -2), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -2), (-1, -1), colors.HexColor("#F2F2F2")),
            ("GRID", (0, -2), (-1, -1), 0.4, colors.grey),
            ("LINEBEFORE", (1, -2), (1, -1), 0.6, colors.HexColor("#F2F2F2")),
            ("LINEAFTER", (1, -2), (1, -1), 0.6, colors.HexColor("#F2F2F2")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -3), [colors.white, colors.HexColor("#DCEEF9")]),
            ("ALIGN", (3, 0), (5, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        blocco_pdf = [Paragraph(f"<b>{blocco['nome']}</b>", stili["Heading3"]), tabella, Spacer(1, 16)]
        elementi.append(KeepTogether(blocco_pdf))

    doc.build(elementi)
    buf.seek(0)
    return buf.getvalue()


def prossimo_id_anagrafica(df: pd.DataFrame) -> int:
    if "ID" not in df.columns or df.empty:
        return 1
    numeri = pd.to_numeric(df["ID"], errors="coerce").dropna()
    return int(numeri.max()) + 1 if not numeri.empty else 1


def salva_riga_foglio(_workbook, nome_foglio: str, riga_intestazione: int,
                       valori: dict, riga_da_aggiornare: int = None):
    try:
        ws = _workbook.worksheet(nome_foglio)
        intestazioni = ws.row_values(riga_intestazione)
        riga_completa = [valori.get(nome, "") for nome in intestazioni]

        if riga_da_aggiornare is None:
            ws.append_row(riga_completa, value_input_option="USER_ENTERED")
        else:
            ultima_colonna = gspread.utils.rowcol_to_a1(1, len(intestazioni)).rstrip("0123456789")
            intervallo = f"A{riga_da_aggiornare}:{ultima_colonna}{riga_da_aggiornare}"
            ws.update(intervallo, [riga_completa], value_input_option="USER_ENTERED")
        return True, None
    except Exception as e:
        return False, f"Errore durante il salvataggio: {e}"


def elimina_riga_foglio(_workbook, nome_foglio: str, riga_da_eliminare: int):
    try:
        ws = _workbook.worksheet(nome_foglio)
        ws.delete_rows(riga_da_eliminare)
        return True, None
    except Exception as e:
        return False, f"Errore durante l'eliminazione: {e}"


def salva_riga_tutti(_workbook, riga_foglio: int, nuova_grezza: list):
    try:
        ws = _workbook.worksheet(NOME_FOGLIO_TUTTI)
        ws.update(f"C{riga_foglio}:J{riga_foglio}", [nuova_grezza], value_input_option="USER_ENTERED")
        return True, None
    except Exception as e:
        return False, f"Errore durante il salvataggio: {e}"


def salva_riga_anagrafica(_workbook, valori: dict, riga_da_aggiornare: int = None):
    return salva_riga_foglio(_workbook, NOME_FOGLIO_ANAGRAFICA, RIGA_INTESTAZIONE_ANAGRAFICA,
                              valori, riga_da_aggiornare)


# ─────────────────────────────────────────────────────────────────
# CONNESSIONE (navigazione tramite le card)
# ─────────────────────────────────────────────────────────────────
if "pagina" not in st.session_state:
    st.session_state.pagina = "home"


def vai_a(pagina: str):
    st.session_state.pagina = pagina


def vai_a_home_reset_riepilogo():
    st.session_state["riepilogo_expander_aperto"] = False
    vai_a("home")


def vai_a_home_reset_importa_s21():
    for chiave in ("importa_s21_dati", "importa_s21_chiave_file",
                   "importa_s21_persona_scelta", "s21_form_manuale_aperto"):
        st.session_state.pop(chiave, None)
    vai_a("home")


workbook, errore = apri_foglio_dati()
collegato = workbook is not None


# ─────────────────────────────────────────────────────────────────
# Pagina: per il controllo dell'Anno Teocratico nei Promemoria
# ─────────────────────────────────────────────────────────────────

def _normalizza_data_report(valore) -> str:
    """Converte un valore di 'Mese/Anno' nel formato canonico 'AAAA-MM',
    indipendentemente da come è stato scritto in origine. Gestisce:
    - 'AAAA-MM' (es. '2026-06')
    - 'AAAA-MM-GG' (es. '2026-06-01')
    - 'GG/MM/AAAA' (es. '01/06/2026')
    - 'MM/AAAA' (es. '06/2026')
    - oggetti data/datetime/Timestamp
    Restituisce None se il valore non è riconoscibile. Usata SOLO per il
    controllo dei rapporti mancanti in Home (nome diverso dalla
    'normalizza_mese_anno' usata in Importa S-21, per non avere due
    funzioni con lo stesso nome e comportamento diverso nello stesso file)."""
    if valore is None or (isinstance(valore, float) and pd.isna(valore)):
        return None
    if hasattr(valore, "year") and hasattr(valore, "month"):
        return f"{valore.year}-{valore.month:02d}"

    v = str(valore).strip()
    if not v or v.lower() == "nan":
        return None

    if "-" in v:
        parti = v.split("-")
        if len(parti) in (2, 3):
            try:
                a, m = int(parti[0]), int(parti[1])
                if 1 <= m <= 12:
                    return f"{a}-{m:02d}"
            except Exception:
                pass

    if "/" in v:
        parti = v.split("/")
        if len(parti) == 3:
            try:
                g, m, a = int(parti[0]), int(parti[1]), int(parti[2])
                if 1 <= m <= 12:
                    return f"{a}-{m:02d}"
            except Exception:
                pass
        elif len(parti) == 2:
            try:
                m, a = int(parti[0]), int(parti[1])
                if 1 <= m <= 12:
                    return f"{a}-{m:02d}"
            except Exception:
                pass

    return None


def trova_ultimo_mese_consegnato_foglio_tutti(df_tutti: pd.DataFrame) -> str:
    """Individua l'ultimo mese ('AAAA-MM') effettivamente presente nel
    foglio Tutti, indipendentemente dal formato con cui la data è scritta."""
    if df_tutti.empty or "Mese/Anno" not in df_tutti.columns:
        return None

    validi = []
    for m in df_tutti["Mese/Anno"].dropna().unique():
        m_norm = _normalizza_data_report(m)
        if m_norm:
            a, mm = m_norm.split("-")
            validi.append((int(a), int(mm), m_norm))

    if not validi:
        return None

    validi.sort(key=lambda x: (x[0], x[1]))
    return validi[-1][2]


def genera_mesi_anno_teocratico_fino_a_ultimo(ultimo_mese_str: str) -> list:
    """Genera la lista dei mesi ('AAAA-MM') da Settembre dell'anno
    teocratico corrente fino all'ultimo mese presente in archivio."""
    if not ultimo_mese_str:
        return []

    try:
        a_lim, m_lim = map(int, ultimo_mese_str.split("-"))
    except Exception:
        return []

    anno_inizio_teo = a_lim if m_lim >= 9 else a_lim - 1

    mesi_dovuti = []

    for m in range(9, 13):
        if (anno_inizio_teo < a_lim) or (anno_inizio_teo == a_lim and m <= m_lim):
            mesi_dovuti.append(f"{anno_inizio_teo}-{m:02d}")

    anno_fine_teo = anno_inizio_teo + 1
    for m in range(1, 9):
        if (anno_fine_teo < a_lim) or (anno_fine_teo == a_lim and m <= m_lim):
            mesi_dovuti.append(f"{anno_fine_teo}-{m:02d}")

    return mesi_dovuti


def trova_colonna_stato_robusta(df_anagrafica: pd.DataFrame):
    """Individua la colonna di stato (Attivi/Inattivi) tra i candidati che
    contengono 'attiv'/'stato'/'a/i' nel nome, scegliendo quella che si
    comporta realmente come una colonna di stato: pochi valori distinti
    e brevi (es. 'A', 'I', 'ATTIVO'), NON una colonna di date o testo
    libero (es. 'Data attività') che conterrebbe comunque 'attiv'."""
    candidati = []
    for col in df_anagrafica.columns:
        c_str = str(col).strip().lower()
        if "attiv" in c_str or c_str == "a/i" or "stato" in c_str:
            candidati.append(col)

    if not candidati:
        return None

    migliore = None
    punteggio_migliore = -1
    for col in candidati:
        vals = df_anagrafica[col].dropna().astype(str).str.strip()
        if vals.empty:
            continue
        n_unici = vals.nunique()
        lunghezza_media = vals.str.len().mean()
        if n_unici <= 6 and lunghezza_media <= 12:
            punteggio = 100 - n_unici - lunghezza_media
            if punteggio > punteggio_migliore:
                punteggio_migliore = punteggio
                migliore = col

    return migliore


def calcola_stato_rapporti_completo(df_tutti: pd.DataFrame, df_anagrafica: pd.DataFrame):
    """Verifica che ogni proclamatore attivo in Anagrafica abbia un
    rapporto nel foglio Tutti per CIASCUN mese da Settembre dell'anno
    teocratico corrente fino all'ultimo mese registrato. Le date in
    'Mese/Anno' vengono normalizzate indipendentemente dal formato in cui
    sono scritte. Restituisce esattamente 3 elementi:
    (totale_presenti, totale_dovuti, dettaglio_mancanti)"""
    if df_anagrafica.empty or df_tutti.empty:
        return None, None, []

    # 1. Colonna Nome in Anagrafica
    col_nome_ana = None
    for col in df_anagrafica.columns:
        c_str = str(col).strip().lower()
        if "cognome" in c_str or "nome" in c_str or "proclamatore" in c_str:
            col_nome_ana = col
            break

    if not col_nome_ana:
        return None, None, []

    # 2. Colonna Stato (versione robusta)
    col_stato = trova_colonna_stato_robusta(df_anagrafica)

    if col_stato:
        vals_stato = df_anagrafica[col_stato].astype(str).str.strip().str.upper()
        maschera_attivi = vals_stato.str.startswith("A") | (vals_stato == "SI") | (vals_stato == "TRUE")
    else:
        maschera_attivi = pd.Series(True, index=df_anagrafica.index)

    proclamatori_attivi = set(
        df_anagrafica.loc[maschera_attivi, col_nome_ana].dropna().astype(str).str.strip().unique()
    )
    proclamatori_attivi = {p for p in proclamatori_attivi if p and p.lower() not in ["cognome e nome", "nome", "none", "nan"]}

    n_attivi = len(proclamatori_attivi)
    if n_attivi == 0:
        return 0, 0, []

    # 3. Ultimo mese e mesi attesi
    ultimo_mese = trova_ultimo_mese_consegnato_foglio_tutti(df_tutti)
    mesi_attesi = genera_mesi_anno_teocratico_fino_a_ultimo(ultimo_mese)

    if not mesi_attesi:
        return 0, 0, []

    # 4. Mappa rapporti presenti (date normalizzate)
    col_nome_tutti = None
    for col in df_tutti.columns:
        c_str = str(col).strip().lower()
        if "cognome" in c_str or "nome" in c_str or "proclamatore" in c_str:
            col_nome_tutti = col
            break

    rapporti_presenti = {}
    if col_nome_tutti and "Mese/Anno" in df_tutti.columns:
        for _, riga in df_tutti.iterrows():
            nome = str(riga.get(col_nome_tutti, "")).strip()
            mese = _normalizza_data_report(riga.get("Mese/Anno", ""))
            if nome and mese:
                rapporti_presenti.setdefault(nome, set()).add(mese)

    dettaglio_mancanti = []
    totale_rapporti_dovuti = n_attivi * len(mesi_attesi)
    totale_rapporti_presenti = 0

    for nome in sorted(proclamatori_attivi):
        mesi_inseriti = rapporti_presenti.get(nome, set())
        for m in mesi_attesi:
            if m in mesi_inseriti:
                totale_rapporti_presenti += 1
            else:
                dettaglio_mancanti.append(f"{m} {nome}")

    return totale_rapporti_presenti, totale_rapporti_dovuti, dettaglio_mancanti

# ─────────────────────────────────────────────────────────────────
# PAGINA: HOME (Tab "🏠 Home" con card promemoria responsive + card originali)
# ─────────────────────────────────────────────────────────────────

def mostra_home():
    ora_ora = datetime.now().strftime('%d/%m/%Y %H:%M')
    st.markdown(
        f"""
        <div style="margin-bottom: 12px;">
            <h3 style="font-size: 1.25rem; font-weight: 700; margin: 0; padding: 0;">📒 Gestione Registrazioni SEG</h3>
            <p style="font-size: 0.8rem; color: #6b7280; margin: 2px 0 0 0; padding: 0;">
                Ultimo aggiornamento: {ora_ora}
            </p>
        </div>
        """,
        unsafe_allow_html=True
    )

    st.markdown("""
    <style>
        .custom-card-header {
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 12px;
            margin-bottom: 8px;
            width: 100%;
        }
        .custom-card-title-group {
            display: flex;
            align-items: center;
            gap: 10px;
            flex: 1;
            min-width: 0;
        }
        .custom-card-title {
            font-size: 1.1rem !important;
            font-weight: 700 !important;
            line-height: 1.3 !important;
            margin: 0 !important;
            padding: 0 !important;
        }
        .custom-icon-box {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            width: 38px;
            height: 38px;
            min-width: 38px;
            border-radius: 10px;
            font-size: 1.3rem;
            flex-shrink: 0;
        }
        .bg-orange { background: rgba(249, 115, 22, 0.15); border: 1px solid rgba(249, 115, 22, 0.4); }
        .bg-blue   { background: rgba(59, 130, 246, 0.15); border: 1px solid rgba(59, 130, 246, 0.4); }
        .bg-green  { background: rgba(16, 185, 129, 0.15); border: 1px solid rgba(16, 185, 129, 0.4); }
        .bg-purple { background: rgba(168, 85, 247, 0.15); border: 1px solid rgba(168, 85, 247, 0.4); }
        .bg-cyan   { background: rgba(6, 182, 212, 0.15); border: 1px solid rgba(6, 182, 212, 0.4); }
        .bg-amber  { background: rgba(245, 158, 11, 0.15); border: 1px solid rgba(245, 158, 11, 0.4); }
        .bg-slate  { background: rgba(100, 116, 139, 0.15); border: 1px solid rgba(100, 116, 139, 0.4); }

        .hud-badge {
            padding: 4px 10px;
            border-radius: 12px;
            font-size: 0.8rem;
            font-weight: 600;
            white-space: nowrap;
            display: inline-block;
        }
        .hud-green {
            background: rgba(16, 185, 129, 0.15);
            color: #10b981;
            border: 1px solid rgba(16, 185, 129, 0.3);
        }
        .hud-yellow {
            background: rgba(245, 158, 11, 0.15);
            color: #f59e0b;
            border: 1px solid rgba(245, 158, 11, 0.3);
        }
        .hud-red {
            background: rgba(239, 68, 68, 0.15);
            color: #ef4444;
            border: 1px solid rgba(239, 68, 68, 0.3);
        }

        .stTabs [data-baseweb="tab-list"] {
            gap: 6px;
            border-bottom: 1px solid rgba(128,128,128,0.3);
        }
        .stTabs [data-baseweb="tab"] {
            font-weight: 600;
            font-size: 0.95rem;
            color: var(--text-color);
            opacity: 0.65;
            padding: 10px 6px;
        }
        .stTabs [aria-selected="true"] {
            color: #2E7D32 !important;
            opacity: 1 !important;
            font-weight: 700 !important;
        }
        .stTabs [data-baseweb="tab-highlight"] {
            background-color: #2E7D32 !important;
            height: 3px !important;
            border-radius: 2px;
        }

        div[class*="st-key-card_"] {
            position: relative !important;
            box-shadow: 3px 5px 14px rgba(0,0,0,0.18);
            transition: border-color 0.15s ease, box-shadow 0.15s ease, transform 0.15s ease;
            cursor: pointer;
        }
        div[class*="st-key-card_"]:hover {
            border-color: #2E7D32 !important;
            box-shadow: 4px 7px 18px rgba(0,0,0,0.24);
            transform: translateY(-2px);
        }

        div[class*="st-key-card_"] .custom-card-header,
        div[class*="st-key-card_"] [data-testid="stCaptionContainer"] {
            pointer-events: none !important;
        }

        div[class*="st-key-card_"] div[data-testid="stElementContainer"]:has(div[data-testid="stButton"]) {
            position: absolute !important;
            inset: 0 !important;
            width: 100% !important;
            height: 100% !important;
            margin: 0 !important;
            padding: 0 !important;
            z-index: 5 !important;
        }

        div[class*="st-key-card_"] div[data-testid="stButton"] {
            width: 100% !important;
            height: 100% !important;
            margin: 0 !important;
            padding: 0 !important;
        }

        div[class*="st-key-card_"] div[data-testid="stButton"] button {
            width: 100% !important;
            height: 100% !important;
            opacity: 0 !important;
            background: transparent !important;
            border: none !important;
            cursor: pointer !important;
            margin: 0 !important;
            padding: 0 !important;
        }

        div[class*="st-key-card_"] div[data-testid="stButton"] button:disabled {
            cursor: not-allowed !important;
        }

        .postit-card {
            width: 92%;
            max-width: 900px;
            margin: 8px auto 24px auto;
            background: linear-gradient(135deg, #fff9c4, #fff3a0);
            border-radius: 4px 14px 4px 14px;
            padding: 22px clamp(20px, 4vw, 40px);
            box-shadow: 3px 5px 14px rgba(0,0,0,0.18);
            transform: rotate(-0.8deg);
        }
        .postit-titolo {
            font-size: clamp(1.05rem, 1.6vw, 1.3rem);
            font-weight: 700;
            color: #5c4a00;
            margin: 0 0 12px 0;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        .postit-lista {
            display: block;
        }
        .promemoria-riga {
            display: flex;
            align-items: flex-start;
            gap: 10px;
            padding: 7px 0;
            border-bottom: 1px dashed rgba(0,0,0,0.15);
        }
        .promemoria-riga:last-child {
            border-bottom: none;
        }
        .dot {
            width: 11px;
            height: 11px;
            border-radius: 50%;
            margin-top: 4px;
            flex-shrink: 0;
        }
        .dot-green  { background: #16a34a; }
        .dot-yellow { background: #eab308; }
        .dot-red    { background: #dc2626; }
        .dot-grey   { background: #9ca3af; }
        .promemoria-testo {
            font-size: 0.92rem;
            color: #4a3f00;
            line-height: 1.35;
        }

        @media (min-width: 900px) {
            .postit-lista {
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 2px 32px;
            }
            .promemoria-testo {
                font-size: 0.98rem;
            }
        }
    </style>
    """, unsafe_allow_html=True)

    badge_rapporti = ""
    badge_anagrafica = ""

    n_completi_anagrafica = None
    n_incompleti_anagrafica = None

    df_anagrafica_home = pd.DataFrame()
    df_tutti_home = pd.DataFrame()

    esito_presenze_adunanza = None
    nomi_mancanti_rapporto_mese = None  # None = non calcolabile; lista = proclamatori attivi senza rapporto del mese

    if collegato:
        try:
            df_anagrafica_home, err_ana_home = leggi_foglio_come_df(
                workbook, NOME_FOGLIO_ANAGRAFICA, RIGA_INTESTAZIONE_ANAGRAFICA)
            df_risposte_home, err_risp_home = leggi_foglio_come_df(
                workbook, NOME_FOGLIO_RISPOSTE, RIGA_INTESTAZIONE_RISPOSTE)
            df_tutti_home, err_tutti_home = leggi_foglio_come_df(
                workbook, "Tutti", 1)

            if not err_ana_home and not df_anagrafica_home.empty:
                if not err_risp_home:
                    if "Attivi / Inattivi" in df_anagrafica_home.columns:
                        categorie_home = df_anagrafica_home["Attivi / Inattivi"].apply(categoria_stato_proclamatore)
                    else:
                        categorie_home = pd.Series(["A"] * len(df_anagrafica_home), index=df_anagrafica_home.index)

                    nomi_attivi_home = set(
                        df_anagrafica_home.loc[categorie_home == "A", "Cognome e Nome"].astype(str).str.strip()
                    )
                    conteggio_attivi_home = len(nomi_attivi_home)

                    if "Cognome e Nome" in df_risposte_home.columns:
                        nomi_consegnati_home = set(
                            df_risposte_home["Cognome e Nome"].astype(str).str.strip()
                        ) & nomi_attivi_home
                    else:
                        nomi_consegnati_home = set()

                    conteggio_consegnati_home = len(nomi_consegnati_home)
                    completo = conteggio_attivi_home > 0 and conteggio_consegnati_home >= conteggio_attivi_home
                    cls_badge = "hud-green" if completo else "hud-red"
                    badge_rapporti = f'<span class="hud-badge {cls_badge}">{conteggio_consegnati_home} / {conteggio_attivi_home}</span>'

                    # Proclamatori attivi che non compaiono ancora in "Risposte del modulo 9" (colonna B) per questo mese
                    nomi_mancanti_rapporto_mese = sorted(nomi_attivi_home - nomi_consegnati_home)

                colonne_obbligatorie = ["ID", "Cognome e Nome", "Data Nascita", "Sesso", "Tipo", "A/U", "Gruppo", "Attivi / Inattivi"]
                if "Attivi / Inattivi" in df_anagrafica_home.columns:
                    df_attivi_ana = df_anagrafica_home[
                        df_anagrafica_home["Attivi / Inattivi"].astype(str).str.strip().str.upper().str.startswith("A")
                    ].copy()

                    def riga_completa_home(riga):
                        for col in colonne_obbligatorie:
                            if col in riga:
                                val = str(riga[col]).strip()
                                if not val or val.lower() in ["none", "nan", "null"]:
                                    return False
                            else:
                                return False
                        return True

                    if not df_attivi_ana.empty:
                        esiti = df_attivi_ana.apply(riga_completa_home, axis=1)
                        tot_comp = esiti.sum()
                        tot_incomp = len(df_attivi_ana) - tot_comp

                        b_list = []
                        if tot_comp > 0:
                            b_list.append(f'<span class="hud-badge hud-green"> Completi: {tot_comp}</span>')
                        if tot_incomp > 0:
                            b_list.append(f'<span class="hud-badge hud-yellow"> Incompleti: {tot_incomp}</span>')

                        badge_anagrafica = " ".join(b_list)

                        n_completi_anagrafica = int(tot_comp)
                        n_incompleti_anagrafica = int(tot_incomp)

            try:
                df_config, err_cfg = leggi_foglio_come_df(workbook, "Configurazioni", 1)
                df_presenze, err_pres = leggi_foglio_come_df(workbook, "Presenze Adunanze", 1)

                giorni_map = {0: "Lunedì", 1: "Martedì", 2: "Mercoledì", 3: "Giovedì", 4: "Venerdì", 5: "Sabato", 6: "Domenica"}

                giorni_adunanza = []
                if not err_cfg and not df_config.empty:
                    for col in df_config.columns:
                        vals = df_config[col].dropna().astype(str).str.strip().tolist()
                        for v in vals:
                            if v in giorni_map.values():
                                giorni_adunanza.append(v)

                if not giorni_adunanza:
                    giorni_adunanza = ["Giovedì", "Domenica"]

                date_registrate = set()
                if not err_pres and not df_presenze.empty:
                    col_data = df_presenze.columns[0]
                    for d in df_presenze[col_data].dropna():
                        d_str = str(d).strip()
                        try:
                            dt_parsed = pd.to_datetime(d_str, dayfirst=True).date()
                            date_registrate.add(dt_parsed)
                        except Exception:
                            pass

                oggi = date.today()
                check_date = oggi - timedelta(days=1)
                ultima_adunanza_data = None
                giorno_nome = ""

                for _ in range(7):
                    g_nome = giorni_map[check_date.weekday()]
                    if g_nome in giorni_adunanza:
                        ultima_adunanza_data = check_date
                        giorno_nome = g_nome
                        break
                    check_date -= timedelta(days=1)

                if ultima_adunanza_data:
                    if ultima_adunanza_data in date_registrate:
                        esito_presenze_adunanza = True
                    else:
                        d_fmt = ultima_adunanza_data.strftime("%d/%m/%Y")
                        esito_presenze_adunanza = (False, d_fmt, giorno_nome)
            except Exception:
                esito_presenze_adunanza = None

        except Exception:
            badge_rapporti = ""
            badge_anagrafica = ""

    promemoria = []

    # ─────────────────────────────────────────────────────────────────
    # Segnalazione: Rapporto del mese corrente (da Anagrafica attivi + foglio
    # "Risposte del modulo 9", colonna B) — non ancora consegnato
    # ─────────────────────────────────────────────────────────────────
    n_mancanti_mese = len(nomi_mancanti_rapporto_mese) if nomi_mancanti_rapporto_mese else 0
    if n_mancanti_mese > 0:
        dot_cls_mese = "dot-yellow" if n_mancanti_mese < 5 else "dot-red"
        if n_mancanti_mese == 1:
            testo_mese = "1 proclamatore non ha ancora consegnato il rapporto di questo mese."
        else:
            testo_mese = f"{n_mancanti_mese} proclamatori non hanno ancora consegnato il rapporto di questo mese."
        promemoria.append((dot_cls_mese, testo_mese))
    # Se hanno consegnato tutti (o non ci sono proclamatori attivi), non si scrive nulla.

    # ─────────────────────────────────────────────────────────────────
    # Segnalazione 1: Rapporti dell'Anno Teocratico
    # ─────────────────────────────────────────────────────────────────
    n_consegnati_rapporti, n_dovuti_rapporti, mancanti_dettaglio = calcola_stato_rapporti_completo(
        df_tutti_home, df_anagrafica_home)

    if n_dovuti_rapporti is None:
        promemoria.append(("dot-grey", "Connettiti al foglio Google per vedere lo stato dei rapporti consegnati."))
    elif mancanti_dettaglio:
        n_persone_mancanti = len(mancanti_dettaglio)
        dot_cls = "dot-yellow" if n_persone_mancanti < 5 else "dot-red"
        testo = "Rapporti mancanti in archivio:<br>" + "<br>".join(mancanti_dettaglio)
        promemoria.append((dot_cls, testo))
    # Se non manca nulla (o non ci sono proclamatori attivi da controllare), non si scrive nulla.

    # Segnalazione 2: Anagrafiche
    if n_completi_anagrafica is None or (n_completi_anagrafica == 0 and n_incompleti_anagrafica == 0):
        promemoria.append(("dot-grey", "Nessuna anagrafica attiva da verificare al momento."))
    elif n_incompleti_anagrafica == 0:
        promemoria.append(("dot-green", "Tutte le anagrafiche sono complete."))
    elif n_incompleti_anagrafica < n_completi_anagrafica:
        promemoria.append(("dot-yellow", f"{n_incompleti_anagrafica} anagrafiche incomplete da controllare."))
    else:
        promemoria.append(("dot-red", f"{n_incompleti_anagrafica} anagrafiche incomplete su {n_completi_anagrafica + n_incompleti_anagrafica}: da sistemare con priorità."))

    # Segnalazione 3: Presenze Adunanze
    if esito_presenze_adunanza is True:
        promemoria.append(("dot-green", "Presenze dell'ultima adunanza registrate correttamente."))
    elif isinstance(esito_presenze_adunanza, tuple) and esito_presenze_adunanza[0] is False:
        _, data_mancante, giorno_mancante = esito_presenze_adunanza
        promemoria.append(("dot-red", f"Presenze adunanza non inserite per {giorno_mancante} {data_mancante}."))

    righe_html = "".join(
        f'<div class="promemoria-riga"><span class="dot {dot_cls}"></span>'
        f'<span class="promemoria-testo">{testo}</span></div>'
        for dot_cls, testo in promemoria
    )

    postit_html = f"""
    <div class="postit-card">
        <div class="postit-titolo">📌 Promemoria e Segnalazioni</div>
        <div class="postit-lista">
            {righe_html}
        </div>
    </div>
    """

    lista_impostazioni = [
        ("⚙️", "bg-slate",  "Impostazioni", "Configura i giorni delle adunanze e altre opzioni.", "impostazioni", ""),
    ]
    if st.session_state.get("ruolo") == "amministratore":
        lista_impostazioni.append(
            ("🔐", "bg-purple", "Accessi", "Gestisci chi può accedere all'app e con quale ruolo.", "utenti", "")
        )

    sezioni = {
        "📖 Rapporti": [
            ("📖", "bg-orange", "Rapporti consegnati", "Visualizza e modifica i rapporti di servizio consegnati.", "registrazioni", badge_rapporti),
            ("📚", "bg-blue",   "Storico rapporti", "Storico dei rapporti di servizio per Proclamatore.", "storico", ""),
            ("📇", "bg-cyan",   "Cartoline di registrazione", "Genera le cartoline S-21 per i Proclamatori scelti.", "cartoline", ""),
            ("🏢", "bg-amber",  "Rapporto per la Filiale", "Dati statistici mensili (tipo modulo S-10).", "filiale", ""),
            ("📊", "bg-blue",   "Riepilogo attività e statistiche", "Report su ore, studi e crediti per Proclamatore o per categoria.", "riepilogo_statistiche", ""),
            ("📥", "bg-orange", "Importa da S-21", "Importa ore/studi da una S-21 ricevuta (Proclamatore trasferito).", "importa_s21", ""),
        ],
        "🗂️ Anagrafiche": [
            ("🗂️", "bg-green",  "Anagrafiche", "Gestisci i dati dei Proclamatori.", "anagrafiche", badge_anagrafica),
            ("👥", "bg-purple", "Gruppi di servizio", "Abbina i Proclamatori a un sorvegliante di gruppo.", "gruppi", ""),
        ],
        "🙌 Adunanze": [
            ("🙌", "bg-green",  "Presenti alle adunanze", "Registra e monitora le presenze alle due adunanze.", "presenze", ""),
        ],
        "⚙️ Impostazioni": lista_impostazioni,
    }

    def mostra_griglia_card(lista_card):
        """Mostra le card di una tab in una griglia a 2 colonne."""
        for i in range(0, len(lista_card), 2):
            coppia = lista_card[i:i + 2]
            cols = st.columns(2)
            for col, (icon, bg_cls, titolo, desc, pagina, badge) in zip(cols, coppia):
                with col:
                    with st.container(key=f"card_{pagina}", border=True):
                        st.markdown(
                            f"""
                            <div class="custom-card-header">
                                <div class="custom-card-title-group">
                                    <span class="custom-icon-box {bg_cls}">{icon}</span>
                                    <span class="custom-card-title">{titolo}</span>
                                </div>
                                <div>{badge}</div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
                        st.caption(desc)

                        st.button(" ", key=f"nav_{pagina}", disabled=not collegato,
                                 on_click=vai_a, args=(pagina,), use_container_width=True)

    nomi_tab = ["🏠 Home"] + list(sezioni.keys())
    tabs = st.tabs(nomi_tab)

    with tabs[0]:
        st.markdown(postit_html, unsafe_allow_html=True)

    for tab, (nome_tab, lista_card) in zip(tabs[1:], sezioni.items()):
        with tab:
            mostra_griglia_card(lista_card)

# ─────────────────────────────────────────────────────────────────
# PAGINA: RAPPORTI CONSEGNATI
# ─────────────────────────────────────────────────────────────────
def _form_rapporto(df: pd.DataFrame, riga_esistente: dict, numero_riga_foglio: int,
                    chiave: str, chiave_stato_modifica: str = None):
    e = riga_esistente
    colonne_numeriche = {"ore", "cr. ore", "cr ore", "studi"}
    colonne_nascoste = {"video mostrati", "cognome e nome"}
    bloccato = sola_lettura()

    with st.form(f"form_rapporto_{chiave}", clear_on_submit=False):
        valori_inseriti = {}
        for colonna in df.columns:
            valore_attuale = e.get(colonna, "")
            chiave_norm = colonna.strip().lower()
            if chiave_norm in colonne_nascoste:
                valori_inseriti[colonna] = valore_attuale
                continue
            if "hai servito" in chiave_norm:
                opzioni = list(OPZIONI_HAI_SERVITO)
                if valore_attuale and valore_attuale not in opzioni:
                    opzioni = [valore_attuale] + opzioni
                indice = opzioni.index(valore_attuale) if valore_attuale in opzioni else 0
                valori_inseriti[colonna] = st.selectbox(colonna, opzioni, index=indice,
                                                        key=f"campo_{colonna}_{chiave}",
                                                        disabled=bloccato)
            elif chiave_norm in colonne_numeriche:
                try:
                    default_num = float(str(valore_attuale).replace(",", ".")) if valore_attuale else 0.0
                except ValueError:
                    default_num = 0.0
                valori_inseriti[colonna] = st.number_input(colonna, value=default_num, step=1.0,
                                                           key=f"campo_{colonna}_{chiave}",
                                                           disabled=bloccato)
            else:
                valori_inseriti[colonna] = st.text_input(colonna, value=str(valore_attuale),
                                                         key=f"campo_{colonna}_{chiave}",
                                                         disabled=bloccato)

        col_btn1, col_btn2 = st.columns([1, 4])
        with col_btn1:
            invia = st.form_submit_button("✔ Salva", type="primary", use_container_width=True,
                                          disabled=bloccato)
        with col_btn2:
            annulla = st.form_submit_button("Annulla", use_container_width=True)

    if annulla:
        if chiave_stato_modifica:
            st.session_state[chiave_stato_modifica] = None
        st.rerun()

    if invia:
        def _sembra_mese_anno(testo: str) -> bool:
            parti = testo.strip().split("-")
            return len(parti) == 2 and parti[0].isdigit() and len(parti[0]) == 4 and parti[1].isdigit()

        def _formatta_valore_salvato(v):
            if isinstance(v, str):
                if _sembra_mese_anno(v):
                    return "'" + v.strip()
                return v
            if isinstance(v, float):
                if v == int(v):
                    return str(int(v))
                return str(v).replace(".", ",")
            return str(v)

        valori_finali = {colonna: _formatta_valore_salvato(v) for colonna, v in valori_inseriti.items()}
        ok, err = salva_riga_foglio(workbook, NOME_FOGLIO_RISPOSTE, RIGA_INTESTAZIONE_RISPOSTE,
                                     valori_finali, riga_da_aggiornare=numero_riga_foglio)
        if ok:
            st.cache_data.clear()
            if chiave_stato_modifica:
                st.session_state[chiave_stato_modifica] = None
            st.success("✔ Salvato correttamente.")
            st.rerun()
        else:
            st.error(err)

def _form_modifica_rapporto_consegnato(dati_selezione: dict):
    df = dati_selezione["df"]
    riga_dict = dati_selezione["riga_dict"]
    numero_riga_foglio = dati_selezione["numero_riga_foglio"]
    nome = dati_selezione["nome"]

    st.title("Modifica rapporto")
    st.caption(f"{nome} (foglio «{NOME_FOGLIO_RISPOSTE}», riga {numero_riga_foglio})")

    _form_rapporto(df, riga_dict, numero_riga_foglio, chiave=str(numero_riga_foglio),
                    chiave_stato_modifica="rapporto_modifica_globale")


def mostra_registrazioni():
    st.title("Rapporti consegnati")
    st.button("🏠 Torna alla Home", key="home_da_registrazioni", use_container_width=True,
              on_click=vai_a, args=("home",))

    if not collegato:
        st.warning("⚠️  Nessun foglio dati collegato.")
        return

    if "rapporto_modifica_globale" not in st.session_state:
        st.session_state.rapporto_modifica_globale = None

    if st.session_state.rapporto_modifica_globale is not None:
        _form_modifica_rapporto_consegnato(st.session_state.rapporto_modifica_globale)
        return

    st.caption(f"Dati letti dal foglio «{NOME_FOGLIO_RISPOSTE}» (intestazione riga {RIGA_INTESTAZIONE_RISPOSTE}).")

    df_anagrafica, err_anagrafica = leggi_foglio_come_df(
        workbook, NOME_FOGLIO_ANAGRAFICA, RIGA_INTESTAZIONE_ANAGRAFICA)
    if err_anagrafica:
        st.error(err_anagrafica)
        return

    df, err = leggi_foglio_come_df(workbook, NOME_FOGLIO_RISPOSTE, RIGA_INTESTAZIONE_RISPOSTE)
    if err:
        st.error(err)
        return

    if df_anagrafica.empty or "Cognome e Nome" not in df_anagrafica.columns:
        st.info("Nessun Proclamatore trovato in Anagrafica.")
        return

    ricerca = st_keyup("🔍 Cerca per nome", placeholder="Digita per filtrare…", key="ricerca_dinamica")

    def e_attivo(valore: str) -> bool:
        return (valore or "").strip().lower().startswith("a")

    colonna_stato = "Attivi / Inattivi" if "Attivi / Inattivi" in df_anagrafica.columns else None
    colonna_gruppo = "Gruppo" if "Gruppo" in df_anagrafica.columns else None
    stato_per_nome = {}
    gruppo_per_nome = {}
    if colonna_stato or colonna_gruppo:
        for _, riga in df_anagrafica.iterrows():
            n = str(riga.get("Cognome e Nome", "")).strip()
            if not n:
                continue
            if colonna_stato:
                stato_per_nome[n] = riga.get(colonna_stato, "")
            if colonna_gruppo:
                gruppo_per_nome[n] = str(riga.get(colonna_gruppo, "")).strip()

    nomi = sorted(n for n in df_anagrafica["Cognome e Nome"].astype(str).str.strip().unique() if n)

    if colonna_stato:
        nomi = [n for n in nomi if e_attivo(stato_per_nome.get(n, ""))]

    testo_ricerca = ricerca.strip().lower()
    if testo_ricerca:
        nomi = [n for n in nomi if testo_ricerca in n.lower()]

    if not nomi:
        st.info("Nessun Proclamatore corrisponde alla ricerca.")
        return

    conteggi = {}
    if "Cognome e Nome" in df.columns:
        serie_nomi_df = df["Cognome e Nome"].astype(str).str.strip().str.lower()
        for nome in nomi:
            conteggi[nome] = (serie_nomi_df == nome.lower()).sum()
    else:
        for nome in nomi:
            conteggi[nome] = 0

    filtro_stato_rapporto = st.radio(
        "Filtro rapporti",
        ["Tutti", "🔴 Da consegnare", "🟢 Consegnati"],
        horizontal=True,
        label_visibility="collapsed",
        key="registrazioni_filtro_stato",
    )
    if filtro_stato_rapporto == "🔴 Da consegnare":
        nomi = [n for n in nomi if conteggi.get(n, 0) == 0]
    elif filtro_stato_rapporto == "🟢 Consegnati":
        nomi = [n for n in nomi if conteggi.get(n, 0) >= 1]

    if not nomi:
        st.info("Nessun Proclamatore corrisponde al filtro selezionato.")
        return

    gruppi = {}
    for n in nomi:
        g = gruppo_per_nome.get(n, "") or "(Senza gruppo)"
        gruppi.setdefault(g, []).append(n)
    for g in gruppi:
        gruppi[g].sort()

    def _riga_proclamatore_rapporto(nome: str):
        conteggio = conteggi.get(nome, 0)
        pallino = "🟢" if conteggio == 1 else "🟡" if conteggio >= 2 else "🔴"

        with st.expander(f"{pallino}  {nome}"):
            if "Cognome e Nome" not in df.columns:
                righe_persona = df.iloc[0:0]
            else:
                righe_persona = df[df["Cognome e Nome"].astype(str).str.strip().str.lower() == nome.lower()]

            if righe_persona.empty:
                st.caption("Nessun rapporto consegnato per questo mese.")
                return

            colonne_tabella = [c for c in df.columns if c.strip().lower() != "cognome e nome"]
            evento_tabella = st.dataframe(
                righe_persona[colonne_tabella],
                hide_index=True,
                use_container_width=True,
                on_select="rerun",
                selection_mode="single-row",
                key=f"tabella_rapp_{nome}",
            )

            righe_sel = evento_tabella.selection.rows if evento_tabella and evento_tabella.selection else []
            idx_originale = None
            if righe_sel:
                sel_idx = righe_sel[0]
                if sel_idx < len(righe_persona):
                    idx_originale = righe_persona.index[sel_idx]

            chiave_conferma_elim = f"rapp_elim_{nome}"

            col_mod, col_elim = st.columns(2)
            with col_mod:
                if st.button("✏️ Modifica riga selezionata", key=f"btn_mod_{nome}",
                            disabled=idx_originale is None, use_container_width=True):
                    numero_riga_foglio = RIGA_INTESTAZIONE_RISPOSTE + 1 + idx_originale
                    st.session_state.rapporto_modifica_globale = {
                        "df": df,
                        "riga_dict": df.loc[idx_originale].to_dict(),
                        "numero_riga_foglio": numero_riga_foglio,
                        "nome": nome,
                    }
                    st.rerun()
            with col_elim:
                if st.button("🗑️ Elimina riga selezionata", key=f"btn_elim_{nome}",
                            disabled=(idx_originale is None) or sola_lettura(), use_container_width=True):
                    st.session_state[chiave_conferma_elim] = True
                    st.rerun()

            if st.session_state.get(chiave_conferma_elim, False) and idx_originale is not None:
                numero_riga_foglio = RIGA_INTESTAZIONE_RISPOSTE + 1 + idx_originale
                st.warning("Confermi l'eliminazione di questo rapporto? L'operazione non è reversibile.")
                col_si, col_no = st.columns(2)
                with col_si:
                    if st.button("✔ Sì, elimina", key=f"btn_conf_si_{nome}",
                                   type="primary", use_container_width=True):
                        ok, err_elim = elimina_riga_foglio(workbook, NOME_FOGLIO_RISPOSTE, numero_riga_foglio)
                        if ok:
                            st.cache_data.clear()
                            st.session_state[chiave_conferma_elim] = False
                            st.success("✔ Rapporto eliminato.")
                            st.rerun()
                        else:
                            st.error(err_elim)
                with col_no:
                    if st.button("No, annulla", key=f"btn_conf_no_{nome}", use_container_width=True):
                        st.session_state[chiave_conferma_elim] = False
                        st.rerun()

            if len(righe_persona) > 1:
                st.divider()

    for gruppo in sorted(gruppi.keys()):
        if gruppi[gruppo]:
            st.markdown(f"#### 👤 {gruppo}")
            for nome in gruppi[gruppo]:
                _riga_proclamatore_rapporto(nome)
            st.divider()

# ─────────────────────────────────────────────────────────────────
# PAGINA: ANAGRAFICHE
# ─────────────────────────────────────────────────────────────────
def _form_anagrafica(df: pd.DataFrame, riga_esistente: dict = None, numero_riga_foglio: int = None,
                      chiave: str = "nuovo", modo_nuovo: bool = False, chiave_expander: str = None):
    e = riga_esistente or {}
    bloccato = sola_lettura()

    def parse_data(s):
        try:
            return datetime.strptime(s, "%d/%m/%Y").date()
        except Exception:
            return None

    with st.form(f"form_anagrafica_{chiave}", clear_on_submit=False):
        nome_cognome = st.text_input("Cognome e Nome *", value=e.get("Cognome e Nome", ""),
                                     key=f"nome_{chiave}", disabled=bloccato)

        eta_nascita = calcola_eta_dettagliata(e.get("Data Nascita", ""))
        if eta_nascita:
            st.markdown(f"**Data di nascita** &nbsp; "
                        f"<span style='color:#D32F2F'>(anni {eta_nascita})</span>",
                        unsafe_allow_html=True)
        else:
            st.markdown("**Data di nascita**")
        data_nascita = st.date_input("Data di nascita", value=parse_data(e.get("Data Nascita", "")),
                                     format="DD/MM/YYYY", min_value=datetime(1900, 1, 1),
                                     label_visibility="collapsed", key=f"data_nascita_{chiave}",
                                     disabled=bloccato)

        sesso_corrente = e.get("Sesso", "")
        sesso_default = ("Maschio" if sesso_corrente.upper().startswith("M")
                          else "Femmina" if sesso_corrente.upper().startswith("F") else "Maschio")
        sesso = st.selectbox("Sesso", OPZIONI_SESSO, index=OPZIONI_SESSO.index(sesso_default),
                              key=f"sesso_{chiave}", disabled=bloccato)

        eta_battesimo = calcola_eta_dettagliata(e.get("Data Battesimo", ""))
        if eta_battesimo:
            st.markdown(f"**Data del battesimo** &nbsp; "
                        f"<span style='color:#D32F2F'>(anni {eta_battesimo})</span>",
                        unsafe_allow_html=True)
        else:
            st.markdown("**Data del battesimo**")
        data_battesimo = st.date_input("Data del battesimo", value=parse_data(e.get("Data Battesimo", "")),
                                        format="DD/MM/YYYY", min_value=datetime(1900, 1, 1),
                                        label_visibility="collapsed", key=f"data_batt_{chiave}",
                                        disabled=bloccato)

        incarico_corrente = e.get("Incarico", "") or "(nessuno)"
        if incarico_corrente not in OPZIONI_INCARICO:
            incarico_corrente = "(nessuno)"
        incarico = st.selectbox("Incarico", OPZIONI_INCARICO,
                                 index=OPZIONI_INCARICO.index(incarico_corrente),
                                 key=f"incarico_{chiave}", disabled=bloccato)

        tipo_corrente = e.get("Tipo", "") or "Proclamatore"
        if tipo_corrente not in OPZIONI_TIPO:
            tipo_corrente = "Proclamatore"
        tipo = st.selectbox("Tipo di servizio", OPZIONI_TIPO,
                             index=OPZIONI_TIPO.index(tipo_corrente), key=f"tipo_{chiave}",
                             disabled=bloccato)
        pr_dal = None
        if tipo in ("Pioniere Regolare", "Pioniere speciale", "Missionario sul campo"):
            pr_dal = st.date_input(f"{tipo} dal", value=parse_data(e.get("PR dal", "")),
                                    format="DD/MM/YYYY", min_value=datetime(1900, 1, 1),
                                    key=f"pr_dal_{chiave}", disabled=bloccato)

        opzioni_gruppo = opzioni_da_colonna(df, "Gruppo")
        gruppo_corrente = e.get("Gruppo", "")
        elenco_gruppo = opzioni_gruppo + ["➕ Nuovo…"]
        if gruppo_corrente and gruppo_corrente not in elenco_gruppo:
            elenco_gruppo = [gruppo_corrente] + elenco_gruppo
        scelta_gruppo = st.selectbox("Gruppo", elenco_gruppo or ["➕ Nuovo…"],
                                      index=(elenco_gruppo.index(gruppo_corrente)
                                             if gruppo_corrente in elenco_gruppo else 0),
                                      key=f"gruppo_{chiave}", disabled=bloccato)
        if scelta_gruppo == "➕ Nuovo…":
            scelta_gruppo = st.text_input("Nome del nuovo gruppo", key=f"gruppo_nuovo_{chiave}",
                                          disabled=bloccato)

        opzioni_au = opzioni_da_colonna(df, "A/U")
        au_corrente = e.get("A/U", "")
        elenco_au = opzioni_au + ["➕ Nuovo…"]
        if au_corrente and au_corrente not in elenco_au:
            elenco_au = [au_corrente] + elenco_au
        scelta_au = st.selectbox("A/U", elenco_au or ["➕ Nuovo…"],
                                  index=(elenco_au.index(au_corrente) if au_corrente in elenco_au else 0),
                                  key=f"au_{chiave}", disabled=bloccato)
        if scelta_au == "➕ Nuovo…":
            scelta_au = st.text_input("Nuovo valore A/U", key=f"au_nuovo_{chiave}", disabled=bloccato)

        note = st.text_area("Note", value=e.get("Note", ""), height=100, key=f"note_{chiave}",
                            disabled=bloccato)

        st.divider()
        st.caption("Promemoria regolarità (da aggiornare quando manca il rapporto mensile)")
        irregolare = st.checkbox("Irregolare", value=e.get("Irregolare", "").strip().upper() in ("X", "SI", "SÌ"),
                                  key=f"irregolare_{chiave}", disabled=bloccato)
        irregolare_mesi = st.number_input("Irregolare da mesi", min_value=0, max_value=36, step=1,
                                           value=int(e.get("Irregolare da Mesi", 0) or 0),
                                           key=f"irregolare_mesi_{chiave}", disabled=bloccato)
        attivi_inattivi_corrente = e.get("Attivi / Inattivi", "A") or "A"
        if attivi_inattivi_corrente not in OPZIONI_ATTIVI_INATTIVI:
            attivi_inattivi_corrente = "A"
        etichetta_stato = st.selectbox("Stato", list(ETICHETTE_ATTIVI_INATTIVI.values()),
                                        index=OPZIONI_ATTIVI_INATTIVI.index(attivi_inattivi_corrente),
                                        key=f"stato_{chiave}", disabled=bloccato)
        attivi_inattivi = {v: k for k, v in ETICHETTE_ATTIVI_INATTIVI.items()}[etichetta_stato]
        dal = st.date_input("Inattivo Da", value=parse_data(e.get("Inattivo dal", "")),
                             format="DD/MM/YYYY", min_value=datetime(1900, 1, 1), key=f"dal_{chiave}",
                             disabled=bloccato)

        st.divider()

        opzioni_trasf = opzioni_da_colonna(df, "Trasf.")
        trasf_corrente = e.get("Trasf.", "")
        elenco_trasf = opzioni_trasf + ["➕ Nuovo…"]
        if trasf_corrente and trasf_corrente not in elenco_trasf:
            elenco_trasf = [trasf_corrente] + elenco_trasf
        scelta_trasf = st.selectbox("Trasf.", elenco_trasf or ["➕ Nuovo…"],
                                     index=(elenco_trasf.index(trasf_corrente)
                                            if trasf_corrente in elenco_trasf else 0),
                                     key=f"trasf_{chiave}", disabled=bloccato)
        if scelta_trasf == "➕ Nuovo…":
            scelta_trasf = st.text_input("Nuovo valore Trasf.", key=f"trasf_nuovo_{chiave}",
                                         disabled=bloccato)

        messaggio = st.text_input("Messaggio", value=e.get("Messaggio", ""), key=f"messaggio_{chiave}",
                                  disabled=bloccato)

        col_btn1, col_btn2 = st.columns([1, 4])
        with col_btn1:
            invia = st.form_submit_button("✔ Salva", use_container_width=True, type="primary",
                                          disabled=bloccato)
        with col_btn2:
            annulla = st.form_submit_button("Annulla", use_container_width=True)

    if annulla:
        if modo_nuovo:
            st.session_state.anagrafica_nuovo = False
        if chiave_expander:
            st.session_state[chiave_expander] = False
        st.rerun()

    if invia:
        if not nome_cognome.strip():
            st.error("Il campo 'Cognome e Nome' è obbligatorio.")
            return

        data_nascita_str = data_nascita.strftime("%d/%m/%Y") if data_nascita else ""
        data_battesimo_str = data_battesimo.strftime("%d/%m/%Y") if data_battesimo else ""

        valori = {
            "ID": e.get("ID") or str(prossimo_id_anagrafica(df)),
            "Cognome e Nome": nome_cognome.strip(),
            "Data Nascita": data_nascita_str,
            "Data Battesimo": data_battesimo_str,
            "Sesso": "M" if sesso == "Maschio" else "F",
            "Incarico": "" if incarico == "(nessuno)" else incarico,
            "Tipo": tipo,
            "PR dal": pr_dal.strftime("%d/%m/%Y") if pr_dal else "",
            "Gruppo": scelta_gruppo,
            "A/U": scelta_au,
            "Trasf.": scelta_trasf,
            "Messaggio": messaggio.strip(),
            "Note": note.strip(),
            "Irregolare": "X" if irregolare else "",
            "Irregolare da Mesi": str(irregolare_mesi) if irregolare else "",
            "Attivi / Inattivi": attivi_inattivi,
            "Inattivo dal": dal.strftime("%d/%m/%Y") if dal else "",
            "Anni Età": calcola_eta(data_nascita_str),
            "Anni Batt": calcola_eta(data_battesimo_str),
        }

        ok, err = salva_riga_anagrafica(workbook, valori, riga_da_aggiornare=numero_riga_foglio)
        if ok:
            st.cache_data.clear()
            if modo_nuovo:
                st.session_state.anagrafica_nuovo = False
            if chiave_expander:
                st.session_state[chiave_expander] = False
            st.success("✔ Salvato correttamente.")
            st.rerun()
        else:
            st.error(err)


def mostra_anagrafiche():
    st.title("Anagrafiche")
    st.button("🏠 Torna alla Home", key="home_da_anagrafiche", use_container_width=True,
              on_click=vai_a, args=("home",))
    st.caption(f"Dati letti dal foglio «{NOME_FOGLIO_ANAGRAFICA}».")

    if not collegato:
        st.warning("⚠️  Nessun foglio dati collegato.")
        return

    if "anagrafica_nuovo" not in st.session_state:
        st.session_state.anagrafica_nuovo = False

    df, err = leggi_foglio_come_df(workbook, NOME_FOGLIO_ANAGRAFICA, RIGA_INTESTAZIONE_ANAGRAFICA)
    if err:
        st.error(err)
        return

    if st.session_state.anagrafica_nuovo:
        st.subheader("➕ Nuovo Proclamatore")
        _form_anagrafica(df, chiave="nuovo", modo_nuovo=True)
        return

    if st.button("➕ Nuovo Proclamatore", use_container_width=True):
        st.session_state.anagrafica_nuovo = True
        st.rerun()

    ricerca = st.text_input("🔍 Cerca per nome, gruppo, tipo…", placeholder="Digita per filtrare…")

    if df.empty:
        st.info("Il foglio è collegato correttamente ma non contiene ancora Proclamatori.")
        return

    df_mostrato = df.reset_index(drop=True)
    if ricerca:
        maschera = df_mostrato.apply(
            lambda riga: riga.astype(str).str.contains(ricerca, case=False, na=False).any(), axis=1
        )
        df_mostrato = df_mostrato[maschera]

    colonne_obbligatorie = ["ID", "Cognome e Nome", "Data Nascita", "Sesso", "Tipo", "A/U", "Gruppo", "Attivi / Inattivi"]

    def riga_e_completa(riga):
        for col in colonne_obbligatorie:
            if col in riga:
                valore = str(riga[col]).strip()
                if not valore or valore.lower() in ["none", "nan", "null"]:
                    return False
            else:
                return False
        return True

    df_mostrato["__completo"] = df_mostrato.apply(riga_e_completa, axis=1)

    if "Attivi / Inattivi" in df_mostrato.columns:
        categorie = df_mostrato["Attivi / Inattivi"].apply(categoria_stato_proclamatore)
    else:
        categorie = pd.Series(["A"] * len(df_mostrato), index=df_mostrato.index)

    conteggio_a = int((categorie == "A").sum())
    conteggio_i = int((categorie == "I").sum())
    conteggio_tr = int((categorie == "TR").sum())

    conteggio_inc = int(((categorie == "A") & (~df_mostrato["__completo"])).sum())

    opzioni_stato = [
        f"🟢 Attivi ({conteggio_a})",
        f"🟡 Anagrafica incompleta ({conteggio_inc})",
        f"🔺 Inattivi ({conteggio_i})",
        f"↔️ Trasferiti ({conteggio_tr})",
    ]

    scelta_stato = st.radio("Stato", opzioni_stato, index=0, horizontal=True,
                             label_visibility="collapsed", key="anagrafica_filtro_stato")

    if scelta_stato == opzioni_stato[0]:
        df_mostrato = df_mostrato[categorie == "A"]
    elif scelta_stato == opzioni_stato[1]:
        df_mostrato = df_mostrato[(categorie == "A") & (~df_mostrato["__completo"])]
    elif scelta_stato == opzioni_stato[2]:
        df_mostrato = df_mostrato[categorie == "I"]
    elif scelta_stato == opzioni_stato[3]:
        df_mostrato = df_mostrato[categorie == "TR"]

    df_mostrato = df_mostrato.sort_values("Cognome e Nome") if "Cognome e Nome" in df_mostrato.columns else df_mostrato

    st.caption(f"{len(df_mostrato)} Proclamatori su {len(df)} totali. Tocca un nominativo per aprirne la scheda.")

    if "anagrafica_aperto" not in st.session_state:
        st.session_state.anagrafica_aperto = None

    st.markdown(
        """
        <style>
        div[data-testid="stButton"] > button[kind="secondary"] {
            justify-content: flex-start !important;
            text-align: left !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    for idx, riga in df_mostrato.iterrows():
        nome = riga.get("Cognome e Nome", "(senza nome)")
        numero_riga_foglio = RIGA_INTESTAZIONE_ANAGRAFICA + 1 + idx
        chiave_persona = str(riga.get("ID") or idx)
        aperto = (st.session_state.anagrafica_aperto == chiave_persona)
        freccia = "▼" if aperto else "▶"

        badge_stato = " 🟡" if not riga.get("__completo", True) else ""

        if st.button(f"{freccia}  {nome}{badge_stato}", key=f"btn_anagrafica_{chiave_persona}", use_container_width=True):
            st.session_state.anagrafica_aperto = None if aperto else chiave_persona
            st.rerun()

        if aperto:
            with st.container(border=True):
                _form_anagrafica(df, riga_esistente=riga.to_dict(), numero_riga_foglio=numero_riga_foglio,
                                  chiave=chiave_persona, modo_nuovo=False,
                                  chiave_expander="anagrafica_aperto")

# ─────────────────────────────────────────────────────────────────
# PAGINA: RIEPILOGO ATTIVITÀ E STATISTICHE
# ─────────────────────────────────────────────────────────────────
def mostra_riepilogo_attivita():
    st.title("📊 Riepilogo attività e statistiche")

    st.button("🏠", key="home_da_riepilogo", help="Torna alla Home", use_container_width=True,
              on_click=vai_a_home_reset_riepilogo)

    if not collegato:
        st.warning("⚠️  Nessun foglio dati collegato.")
        return

    df, err = leggi_foglio_come_df(workbook, NOME_FOGLIO_ANAGRAFICA, RIGA_INTESTAZIONE_ANAGRAFICA)
    if err:
        st.error(err)
        return
    if df.empty or "Cognome e Nome" not in df.columns:
        st.info("Nessun Proclamatore trovato in Anagrafica.")
        return

    df_tutti, err_tutti = leggi_foglio_tutti(workbook)
    if err_tutti:
        st.error(err_tutti)
        return

    anni_presenti = anni_teocratici_per_menu(df_tutti)
    anno_scelto = st.selectbox(
        "Seleziona anno teocratico",
        anni_presenti,
        format_func=lambda a: f"{a} – {a + 1} (set {a} → ago {a + 1})",
    )

    with st.expander("📊 Riepilogo attività", expanded=True, key="riepilogo_expander_aperto"):
        st.caption("Report libero (non la scheda S-21): un elenco con mese, tipo di servizio, ore, "
                   "crediti, studi e note per ciascun Proclamatore, con totali e medie. Utile da "
                   "spedire ai sorveglianti di gruppo.")

        periodo_scelto = st.radio("Periodo", ["12 mesi", "6 mesi"], horizontal=True,
                                   key="riepilogo_periodo")

        tipo_vista = st.radio("Tipo", ["Dettagliato", "Sintetico", "Sintetico compara gruppi"], horizontal=True,
                               key="riepilogo_tipo_vista",
                               help="Dettagliato: un blocco per ciascun Proclamatore. "
                                    "Sintetico: un unico blocco con i totali della categoria scelta. "
                                    "Sintetico compara gruppi: totali/medie per Gruppo, per ogni categoria.")

        gruppi_disponibili = ["Tutti i gruppi"]
        if "Gruppo" in df.columns:
            gruppi_disponibili += sorted({g.strip() for g in df["Gruppo"].astype(str) if g.strip()})
        gruppo_scelto = st.selectbox("Gruppo", gruppi_disponibili, key="riepilogo_gruppo",
                                      disabled=(tipo_vista == "Sintetico compara gruppi"))

        categoria_scelta = st.selectbox("Categoria", list(CATEGORIE_RIEPILOGO_ATTIVITA.keys()),
                                         key="riepilogo_categoria",
                                         disabled=(tipo_vista == "Sintetico compara gruppi"))

        if tipo_vista == "Sintetico compara gruppi":
            st.caption("Questa vista confronta tutti i gruppi in tutte le categorie: "
                       "Gruppo e Categoria scelti sopra vengono ignorati.")

        if st.button("📄 Crea PDF", key="riepilogo_crea_pdf", use_container_width=True):
            with st.spinner("Genero il riepilogo…"):
                etichetta_dati_periodo = _riepilogo_etichetta_dati_estratti(df_tutti, periodo_scelto)
                if tipo_vista == "Sintetico compara gruppi":
                    comparazione = _riepilogo_totali_per_categoria_e_gruppo(df_tutti, df, periodo_scelto)
                    trovato_qualcosa = bool(comparazione)
                    pdf_bytes = genera_pdf_riepilogo_attivita(
                        [], periodo_scelto, categoria_scelta, None,
                        etichetta_vista=tipo_vista, comparazione_gruppi=comparazione,
                        etichetta_dati_periodo=etichetta_dati_periodo,
                    )
                elif tipo_vista == "Sintetico" and categoria_scelta == "Tutti":
                    df_periodo_gruppo = _riepilogo_filtra_dati(df_tutti, df, periodo_scelto,
                                                                gruppo_scelto, "Tutti")
                    totali_categoria = _riepilogo_totali_generali_per_categoria(df_periodo_gruppo)
                    trovato_qualcosa = bool(totali_categoria)
                    pdf_bytes = genera_pdf_riepilogo_attivita(
                        [], periodo_scelto, categoria_scelta,
                        gruppo_scelto if gruppo_scelto != "Tutti i gruppi" else None,
                        etichetta_vista=tipo_vista, totali_per_categoria=totali_categoria,
                        etichetta_dati_periodo=etichetta_dati_periodo,
                    )
                else:
                    df_filtrato = _riepilogo_filtra_dati(df_tutti, df, periodo_scelto, gruppo_scelto,
                                                          categoria_scelta)
                    if tipo_vista == "Sintetico":
                        blocchi = _riepilogo_costruisci_blocco_sintetico(df_filtrato, categoria_scelta)
                    else:
                        blocchi = _riepilogo_costruisci_blocchi_dettagliato(df_filtrato)
                    trovato_qualcosa = bool(blocchi)
                    pdf_bytes = genera_pdf_riepilogo_attivita(
                        blocchi, periodo_scelto, categoria_scelta,
                        gruppo_scelto if gruppo_scelto != "Tutti i gruppi" else None,
                        etichetta_vista=tipo_vista, etichetta_dati_periodo=etichetta_dati_periodo,
                    )
            if not trovato_qualcosa:
                st.warning("Nessun dato trovato per i filtri selezionati — il PDF generato sarà vuoto.")
            st.session_state.riepilogo_pdf_pronto = pdf_bytes

        if st.session_state.get("riepilogo_pdf_pronto"):
            nome_file = "Riepilogo_Attivita"
            if gruppo_scelto != "Tutti i gruppi" and tipo_vista != "Sintetico compara gruppi":
                nome_file += f"_{_s21_nome_file_sicuro(gruppo_scelto)}"
            nome_file += f"_{tipo_vista.replace(' ', '_')}_{categoria_scelta.replace(' ', '_')}.pdf"

            def _chiudi_e_resetta_riepilogo():
                st.session_state.pop("riepilogo_pdf_pronto", None)
                for chiave in ("riepilogo_periodo", "riepilogo_tipo_vista",
                               "riepilogo_gruppo", "riepilogo_categoria"):
                    st.session_state.pop(chiave, None)

            st.download_button(
                "⬇️ Scarica Riepilogo attività (PDF)",
                data=st.session_state.riepilogo_pdf_pronto,
                file_name=nome_file,
                mime="application/pdf",
                key="download_riepilogo_attivita",
                use_container_width=True,
                on_click=_chiudi_e_resetta_riepilogo,
            )

    st.divider()


# ─────────────────────────────────────────────────────────────────
# PAGINA: CARTOLINE DI REGISTRAZIONE (S-21)
# ─────────────────────────────────────────────────────────────────
def mostra_cartoline_registrazione():
    st.title("📇 Cartoline di registrazione")
    contenitore_pulsanti = st.container()

    if not collegato:
        st.warning("⚠️  Nessun foglio dati collegato.")
        return

    if not os.path.exists(PERCORSO_MODULO_S21):
        st.error("Modulo S-21 non trovato: metti il file «S-21_s-Mlt_I.pdf» nella stessa cartella di app.py.")
        return

    df, err = leggi_foglio_come_df(workbook, NOME_FOGLIO_ANAGRAFICA, RIGA_INTESTAZIONE_ANAGRAFICA)
    if err:
        st.error(err)
        return
    if df.empty or "Cognome e Nome" not in df.columns:
        st.info("Nessun Proclamatore trovato in Anagrafica.")
        return

    df_tutti, err_tutti = leggi_foglio_tutti(workbook)
    if err_tutti:
        st.error(err_tutti)
        return

    anni_presenti = anni_teocratici_per_menu(df_tutti)
    anno_scelto = st.selectbox(
        "Seleziona anno teocratico",
        anni_presenti,
        format_func=lambda a: f"{a} – {a + 1} (set {a} → ago {a + 1})",
    )

    df_lista = df.reset_index(drop=True)
    if "Attivi / Inattivi" in df_lista.columns:
        categorie = df_lista["Attivi / Inattivi"].apply(categoria_stato_proclamatore)
        df_lista = df_lista[categorie.isin(["A", "I"])]
        stato_per_nome = dict(zip(df_lista["Cognome e Nome"].astype(str).str.strip(),
                                   categorie[df_lista.index]))
    else:
        stato_per_nome = {}
    df_lista = df_lista[df_lista["Cognome e Nome"].astype(str).str.strip() != ""]

    ricerca = st.text_input("Cerca per nome", placeholder="🔍 Cerca per nome…", label_visibility="collapsed")
    df_mostrato = df_lista
    if ricerca:
        df_mostrato = df_mostrato[df_mostrato["Cognome e Nome"].astype(str).str.contains(ricerca, case=False, na=False)]
    df_mostrato = df_mostrato.sort_values("Cognome e Nome")

    nomi_tutti = [str(n).strip() for n in df_lista["Cognome e Nome"] if str(n).strip()]
    nomi_visibili = [str(n).strip() for n in df_mostrato["Cognome e Nome"] if str(n).strip()]

    def _chiave_cb(nome: str) -> str:
        return f"cb_cartolina_{nome}"

    for nome in nomi_visibili:
        etichetta = f"🔺 {nome}" if stato_per_nome.get(nome) == "I" else nome
        st.checkbox(etichetta, key=_chiave_cb(nome))

    selezionati = [nome for nome in nomi_tutti if st.session_state.get(_chiave_cb(nome), False)]
    n_sel = len(selezionati)

    with contenitore_pulsanti:
        col_home, col_menu, col_vuota = st.columns([1, 1, 5])
        with col_home:
            st.button("🏠", key="home_da_cartoline", use_container_width=True,
                      help="Torna alla Home", on_click=vai_a, args=("home",))
        with col_menu:
            if st.button("⋯", key="toggle_menu_cartoline", use_container_width=True):
                st.session_state.cartoline_menu_aperto = not st.session_state.get(
                    "cartoline_menu_aperto", False)

        genera_tutti = genera_sel = False
        if st.session_state.get("cartoline_menu_aperto"):
            genera_tutti = st.button("🗂️ Crea tutti i PDF delle registrazioni", use_container_width=True)
            genera_sel = st.button(f"📄 Genera cartoline selezionate ({n_sel})",
                                    use_container_width=True, disabled=n_sel == 0)
            if genera_tutti or genera_sel:
                st.session_state.cartoline_menu_aperto = False

        if genera_tutti:
            with st.spinner("Genero il pacchetto completo…"):
                zip_completo = genera_zip_s21_completo(df, df_tutti, anno_scelto)
            st.session_state.cartoline_pacchetto_completo = zip_completo

        if genera_sel:
            righe_sel = [r.to_dict() for _, r in df.iterrows()
                         if str(r.get("Cognome e Nome", "")).strip() in selezionati]
            with st.spinner("Genero le cartoline…"):
                if len(righe_sel) == 1:
                    pdf_bytes = genera_pdf_s21_singolo(righe_sel[0], df_tutti, anno_scelto)
                    st.session_state.cartoline_pronto = ("pdf", pdf_bytes, righe_sel[0].get("Cognome e Nome", ""))
                else:
                    zip_bytes = genera_zip_s21(righe_sel, df_tutti, anno_scelto)
                    st.session_state.cartoline_pronto = ("zip", zip_bytes, anno_scelto)

        if st.session_state.get("cartoline_pacchetto_completo"):
            st.download_button(
                "⬇️ Scarica il pacchetto completo (ZIP)",
                data=st.session_state.cartoline_pacchetto_completo,
                file_name=f"Registrazioni_Complete_{anno_scelto + 1}.zip",
                mime="application/zip",
                key="download_pacchetto_completo",
                use_container_width=True,
                on_click=lambda: st.session_state.pop("cartoline_pacchetto_completo", None),
            )

        pronto = st.session_state.get("cartoline_pronto")
        if pronto:
            tipo, dati_file, extra = pronto
            if tipo == "pdf":
                st.download_button("⬇️ Scarica PDF", data=dati_file,
                                    file_name=f"{_s21_nome_file_sicuro(extra)}.pdf",
                                    mime="application/pdf", key="download_cartolina_pdf",
                                    use_container_width=True,
                                    on_click=lambda: st.session_state.pop("cartoline_pronto", None))
            else:
                st.download_button("⬇️ Scarica ZIP", data=dati_file,
                                    file_name=f"Schede_S21_{extra + 1}.zip",
                                    mime="application/zip", key="download_cartolina_zip",
                                    use_container_width=True,
                                    on_click=lambda: st.session_state.pop("cartoline_pronto", None))

    st.divider()


# ─────────────────────────────────────────────────────────────────
# PAGINA: GRUPPI DI SERVIZIO
# ─────────────────────────────────────────────────────────────────
ETICHETTE_STATO_GRUPPI = {"A": "🟢 Attivi", "I": "🔺 Inattivi", "TR": "↔️ Trasferiti"}


COLORI_GRUPPI = ["BDD7EE", "FBE0D0", "D8ECD2", "FCEDB6", "E4D6EC", "F5C6C6"]
COLORI_TESTATA_GRUPPI = ["9DC3E6", "F4B183", "A9D18E", "FFD966", "C9A0DC", "E8A0A0"]


def _gruppi_calcola_sigla(riga: dict) -> str:
    parti = []
    incarico = (riga.get("Incarico") or "").strip()
    tipo = (riga.get("Tipo") or "").strip()
    if incarico == "Anziano":
        parti.append("A")
    elif incarico == "Servitore di ministero":
        parti.append("SM")
    if tipo == "Pioniere Regolare":
        parti.append("PR")
    elif tipo == "Pioniere speciale":
        parti.append("PS")
    elif tipo == "Missionario sul campo":
        parti.append("M")
    return "/".join(parti)


def _gruppi_trova_assistente(df: pd.DataFrame, gruppo: str) -> str:
    for _, r in df.iterrows():
        g = str(r.get("Gruppo", "")).strip()
        note = str(r.get("Note", "")).strip().lower()
        if g == gruppo and "assistente gruppo di servizio" in note:
            return str(r.get("Cognome e Nome", "")).strip()
    return ""


def _gruppi_dati_filtrati(df: pd.DataFrame, includi_inattivi: bool = False):
    if "Attivi / Inattivi" in df.columns:
        categorie = df["Attivi / Inattivi"].apply(categoria_stato_proclamatore)
        df = df[categorie != "TR"] if includi_inattivi else df[categorie == "A"]
    else:
        categorie = pd.Series(["A"] * len(df), index=df.index)

    gruppi = {}
    for idx, riga in df.iterrows():
        nome = str(riga.get("Cognome e Nome", "")).strip()
        if not nome:
            continue
        g = str(riga.get("Gruppo", "")).strip()
        if not g:
            continue
        stato = categorie.loc[idx] if idx in categorie.index else "A"
        gruppi.setdefault(g, []).append({"nome": nome, "sigla": _gruppi_calcola_sigla(riga.to_dict()),
                                          "stato": stato})
    return df, gruppi


def _gruppi_ordina_membri(membri: list) -> list:
    return sorted(membri, key=lambda m: (0 if m.get("stato") != "I" else 1, m["nome"]))


def genera_excel_gruppi_servizio(df: pd.DataFrame, includi_inattivi: bool = False) -> bytes:
    df, gruppi = _gruppi_dati_filtrati(df, includi_inattivi=includi_inattivi)

    nomi_gruppi = sorted(gruppi.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Gruppi di servizio"

    bordo_sottile = Side(style="thin", color="999999")
    bordo = Border(left=bordo_sottile, right=bordo_sottile, top=bordo_sottile, bottom=bordo_sottile)

    blocco_colonne = 3
    gutter = 1
    riga_cursore = 1

    for indice_coppia in range(0, len(nomi_gruppi), 2):
        coppia = nomi_gruppi[indice_coppia:indice_coppia + 2]
        max_membri = max(len(gruppi[g]) for g in coppia)

        for posizione, nome_gruppo in enumerate(coppia):
            indice_colore = (indice_coppia // 2 + posizione) % len(COLORI_GRUPPI)
            colore_corpo = COLORI_GRUPPI[indice_colore]
            colore_testata = COLORI_TESTATA_GRUPPI[indice_colore]
            col_base = 1 + posizione * (blocco_colonne + gutter)
            col_num, col_nome, col_sigla = col_base, col_base + 1, col_base + 2

            r = riga_cursore
            ws.merge_cells(start_row=r, start_column=col_num, end_row=r, end_column=col_sigla)
            cella = ws.cell(row=r, column=col_num, value=f"Gruppo {nome_gruppo.upper()}")
            cella.font = Font(name="Arial", size=12, bold=True)
            cella.alignment = Alignment(horizontal="center")
            cella.fill = PatternFill("solid", fgColor=colore_testata)

            assistente = _gruppi_trova_assistente(df, nome_gruppo)
            for etichetta, valore, r_offset in (("Sorvegliante", nome_gruppo, 1), ("Assistente", assistente, 2)):
                rr = r + r_offset
                ws.merge_cells(start_row=rr, start_column=col_num, end_row=rr, end_column=col_nome)
                c1 = ws.cell(row=rr, column=col_num, value=valore)
                c1.font = Font(name="Arial", size=10, italic=True, bold=True, color="1F4E78")
                c1.fill = PatternFill("solid", fgColor=colore_corpo)
                c2 = ws.cell(row=rr, column=col_sigla, value=etichetta)
                c2.font = Font(name="Arial", size=10, bold=True)
                c2.alignment = Alignment(horizontal="right")
                c2.fill = PatternFill("solid", fgColor=colore_corpo)

            membri = _gruppi_ordina_membri(gruppi[nome_gruppo])
            for i in range(max_membri):
                rr = r + 3 + i
                cn = ws.cell(row=rr, column=col_num, value=i + 1 if i < len(membri) else "")
                cnome = ws.cell(row=rr, column=col_nome, value=membri[i]["nome"] if i < len(membri) else "")
                csigla = ws.cell(row=rr, column=col_sigla, value=membri[i]["sigla"] if i < len(membri) else "")
                colore_font = "FF0000" if i < len(membri) and membri[i].get("stato") == "I" else "000000"
                for c in (cn, cnome, csigla):
                    c.font = Font(name="Arial", size=10, color=colore_font)
                    c.fill = PatternFill("solid", fgColor=colore_corpo)
                    c.border = bordo
                cn.alignment = Alignment(horizontal="center")
                csigla.alignment = Alignment(horizontal="center")

            ws.column_dimensions[get_column_letter(col_num)].width = 5
            ws.column_dimensions[get_column_letter(col_nome)].width = 26
            ws.column_dimensions[get_column_letter(col_sigla)].width = 10

        riga_cursore += 3 + max_membri + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _gruppi_tabella_pdf(df: pd.DataFrame, nome_gruppo: str, membri: list,
                         colore_corpo: str, colore_testata: str, righe_totali: int = None):
    membri_ordinati = _gruppi_ordina_membri(membri)
    assistente = _gruppi_trova_assistente(df, nome_gruppo)
    n_righe = righe_totali if righe_totali is not None else len(membri_ordinati)

    dati = [[f"Gruppo {nome_gruppo.upper()}", "", ""],
            [nome_gruppo, "", "Sorvegliante"],
            [assistente, "", "Assistente"]]
    for i in range(n_righe):
        if i < len(membri_ordinati):
            m = membri_ordinati[i]
            dati.append([str(i + 1), m["nome"], m["sigla"]])
        else:
            dati.append(["", "", ""])

    larghezze = [1.1 * cm, 4.6 * cm, 1.8 * cm]
    t = Table(dati, colWidths=larghezze)
    stile = [
        ("SPAN", (0, 0), (2, 0)),
        ("SPAN", (0, 1), (1, 1)),
        ("SPAN", (0, 2), (1, 2)),
        ("BACKGROUND", (0, 0), (2, 0), colors.HexColor("#" + colore_testata)),
        ("FONTNAME", (0, 0), (2, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (2, 0), "CENTER"),
        ("BACKGROUND", (0, 1), (2, 2), colors.HexColor("#" + colore_corpo)),
        ("FONTNAME", (0, 1), (1, 2), "Helvetica-BoldOblique"),
        ("ALIGN", (2, 1), (2, 2), "RIGHT"),
        ("BACKGROUND", (0, 3), (2, -1), colors.HexColor("#" + colore_corpo)),
        ("GRID", (0, 3), (2, -1), 0.4, colors.grey),
        ("ALIGN", (0, 3), (0, -1), "CENTER"),
        ("ALIGN", (2, 3), (2, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (2, -1), 7),
        ("TOPPADDING", (0, 0), (2, -1), 1),
        ("BOTTOMPADDING", (0, 0), (2, -1), 1),
        ("VALIGN", (0, 0), (2, -1), "MIDDLE"),
    ]
    for i, m in enumerate(membri_ordinati):
        if m.get("stato") == "I":
            riga_tabella = 3 + i
            stile.append(("TEXTCOLOR", (0, riga_tabella), (2, riga_tabella), colors.red))
    t.setStyle(TableStyle(stile))
    return t


def genera_pdf_gruppi_servizio(df: pd.DataFrame, includi_inattivi: bool = False) -> bytes:
    df, gruppi = _gruppi_dati_filtrati(df, includi_inattivi=includi_inattivi)
    nomi_gruppi = sorted(gruppi.keys())

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=0.7 * cm, bottomMargin=0.7 * cm,
                             leftMargin=0.7 * cm, rightMargin=0.7 * cm)
    stili = getSampleStyleSheet()
    elementi = [Paragraph("Gruppi di servizio", stili["Title"]), Spacer(1, 8)]

    for indice in range(0, len(nomi_gruppi), 2):
        coppia = nomi_gruppi[indice:indice + 2]
        max_membri = max(len(gruppi[g]) for g in coppia)
        celle = []
        for posizione, nome_gruppo in enumerate(coppia):
            indice_colore = (indice // 2 + posizione) % len(COLORI_GRUPPI)
            celle.append(_gruppi_tabella_pdf(df, nome_gruppo, gruppi[nome_gruppo],
                                              COLORI_GRUPPI[indice_colore],
                                              COLORI_TESTATA_GRUPPI[indice_colore],
                                              righe_totali=max_membri))
        if len(celle) == 1:
            celle.append("")
        riga_esterna = Table([celle], colWidths=[9 * cm, 9 * cm])
        riga_esterna.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
        elementi.append(KeepTogether([riga_esterna, Spacer(1, 8)]))

    doc.build(elementi)
    buf.seek(0)
    return buf.getvalue()


def mostra_gruppi_servizio():
    st.title("👥 Gruppi di servizio")
    st.button("🏠 Torna alla Home", key="home_da_gruppi", use_container_width=True,
              on_click=vai_a, args=("home",))
    contenitore_associa = st.container()
    st.caption("Seleziona uno o più Proclamatori e abbinali a un sorvegliante di gruppo.")

    if not collegato:
        st.warning("⚠️  Nessun foglio dati collegato.")
        return

    df, err = leggi_foglio_come_df(workbook, NOME_FOGLIO_ANAGRAFICA, RIGA_INTESTAZIONE_ANAGRAFICA)
    if err:
        st.error(err)
        return
    if df.empty or "Cognome e Nome" not in df.columns:
        st.info("Nessun Proclamatore trovato in Anagrafica.")
        return

    df = df.reset_index(drop=True)

    formato_export = st.radio("Formato esportazione", ["Excel", "PDF", "PDF includi inattivi"],
                               horizontal=True, key="gruppi_formato_export")
    if st.button(f"📥 Esporta Gruppi di servizio ({formato_export})", key="esporta_gruppi",
                 use_container_width=True):
        if formato_export == "Excel":
            st.session_state.gruppi_export_pronto = ("xlsx", genera_excel_gruppi_servizio(df))
        elif formato_export == "PDF":
            st.session_state.gruppi_export_pronto = ("pdf", genera_pdf_gruppi_servizio(df))
        else:
            st.session_state.gruppi_export_pronto = (
                "pdf", genera_pdf_gruppi_servizio(df, includi_inattivi=True))

    if st.session_state.get("gruppi_export_pronto"):
        tipo_file, dati_file = st.session_state.gruppi_export_pronto
        if tipo_file == "xlsx":
            st.download_button(
                "⬇️ Scarica Gruppi di servizio.xlsx",
                data=dati_file,
                file_name="Gruppi_di_servizio.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_gruppi_excel",
                use_container_width=True,
                on_click=lambda: st.session_state.pop("gruppi_export_pronto", None),
            )
        else:
            st.download_button(
                "⬇️ Scarica Gruppi di servizio.pdf",
                data=dati_file,
                file_name="Gruppi_di_servizio.pdf",
                mime="application/pdf",
                key="download_gruppi_pdf",
                use_container_width=True,
                on_click=lambda: st.session_state.pop("gruppi_export_pronto", None),
            )

    if "Attivi / Inattivi" in df.columns:
        categorie = df["Attivi / Inattivi"].apply(categoria_stato_proclamatore)
    else:
        categorie = pd.Series(["A"] * len(df), index=df.index)

    stato_scelto = st.radio("Stato", ["🟢 Attivi", "🔺 Inattivi"], horizontal=True,
                             key="gruppi_stato_filtro")
    codice_stato = {v: k for k, v in ETICHETTE_STATO_GRUPPI.items()}[stato_scelto]

    def _chiave_cb(nome: str) -> str:
        return f"cb_gruppi_{nome}"

    if st.session_state.get("gruppi_stato_precedente") != codice_stato:
        for chiave in list(st.session_state.keys()):
            if chiave.startswith("cb_gruppi_"):
                st.session_state[chiave] = False
        st.session_state.gruppi_stato_precedente = codice_stato

    df_filtrato = df[categorie == codice_stato]
    df_filtrato = df_filtrato[df_filtrato["Cognome e Nome"].astype(str).str.strip() != ""]

    if df_filtrato.empty:
        st.info("Nessun Proclamatore in questa categoria.")
        return

    conteggi_per_gruppo = {}
    for idx, riga in df.iterrows():
        stato_riga = categorie.loc[idx]
        if stato_riga not in ("A", "I"):
            continue
        g = str(riga.get("Gruppo", "")).strip() or "(Senza gruppo)"
        conteggi_per_gruppo.setdefault(g, {"A": 0, "I": 0})
        conteggi_per_gruppo[g][stato_riga] += 1

    gruppi_vista = {}
    for _, riga in df_filtrato.iterrows():
        nome = str(riga.get("Cognome e Nome", "")).strip()
        g = str(riga.get("Gruppo", "")).strip() or "(Senza gruppo)"
        gruppi_vista.setdefault(g, []).append(nome)

    for g in sorted(gruppi_vista.keys()):
        conteggi = conteggi_per_gruppo.get(g, {"A": 0, "I": 0})
        st.markdown(f"#### 👤 {g} (Attivi {conteggi['A']} - Inattivi {conteggi['I']})")
        for nome in sorted(gruppi_vista[g]):
            st.checkbox(nome, key=_chiave_cb(nome))
        st.divider()

    nomi_filtrati = [str(n).strip() for n in df_filtrato["Cognome e Nome"] if str(n).strip()]
    selezionati = [nome for nome in nomi_filtrati if st.session_state.get(_chiave_cb(nome), False)]
    n_sel = len(selezionati)

    with contenitore_associa:
        if st.button(f"🔗 Associa al gruppo ({n_sel})", use_container_width=True,
                     disabled=(n_sel == 0 or sola_lettura())):
            st.session_state.gruppi_mostra_scelta = True

        if st.session_state.get("gruppi_mostra_scelta") and n_sel > 0:
            with st.container(border=True):
                st.caption(f"{n_sel} Proclamatori selezionati.")
                gruppi_esistenti = sorted({g.strip() for g in df["Gruppo"].astype(str) if g.strip()}) \
                    if "Gruppo" in df.columns else []
                opzioni = gruppi_esistenti + ["➕ Nuovo sorvegliante…"]
                scelta = st.selectbox("Sorvegliante di gruppo", opzioni, key="gruppi_scelta_sorvegliante")
                nuovo_nome_gruppo = ""
                if scelta == "➕ Nuovo sorvegliante…":
                    nuovo_nome_gruppo = st.text_input("Nome del nuovo sorvegliante", key="gruppi_nuovo_nome")

                col_abbina, col_annulla, col_elimina = st.columns(3)
                with col_abbina:
                    conferma_abbina = st.button("✔ Abbina", type="primary", use_container_width=True,
                                                key="gruppi_conferma_abbina")
                with col_annulla:
                    conferma_annulla = st.button("✖ Annulla", use_container_width=True,
                                                 key="gruppi_conferma_annulla")
                with col_elimina:
                    conferma_elimina = st.button("🗑️ Elimina", use_container_width=True,
                                                 key="gruppi_conferma_elimina")

                if conferma_annulla:
                    st.session_state.gruppi_mostra_scelta = False
                    for nome in selezionati:
                        st.session_state.pop(_chiave_cb(nome), None)
                    st.rerun()

                if conferma_elimina:
                    errori = []
                    with st.spinner("Rimuovo il gruppo dai Proclamatori selezionati…"):
                        for nome in selezionati:
                            idx_lista = df.index[df["Cognome e Nome"].astype(str).str.strip() == nome]
                            if len(idx_lista) == 0:
                                continue
                            idx = idx_lista[0]
                            numero_riga_foglio = RIGA_INTESTAZIONE_ANAGRAFICA + 1 + idx
                            valori = df.loc[idx].to_dict()
                            valori["Gruppo"] = ""
                            ok, err_salva = salva_riga_anagrafica(workbook, valori,
                                                                   riga_da_aggiornare=numero_riga_foglio)
                            if not ok:
                                errori.append(f"{nome}: {err_salva}")
                    for nome in selezionati:
                        st.session_state.pop(_chiave_cb(nome), None)
                    if errori:
                        st.error("Alcune rimozioni non sono riuscite:\n" + "\n".join(errori))
                    else:
                        st.cache_data.clear()
                        st.session_state.gruppi_mostra_scelta = False
                        st.success(f"✔ Gruppo rimosso per {n_sel} Proclamatori.")

                if conferma_abbina:
                    nome_gruppo_finale = nuovo_nome_gruppo.strip() if scelta == "➕ Nuovo sorvegliante…" else scelta
                    if not nome_gruppo_finale:
                        st.error("Indica il nome del sorvegliante di gruppo.")
                    else:
                        errori = []
                        with st.spinner("Aggiorno l'Anagrafica…"):
                            for nome in selezionati:
                                idx_lista = df.index[df["Cognome e Nome"].astype(str).str.strip() == nome]
                                if len(idx_lista) == 0:
                                    continue
                                idx = idx_lista[0]
                                numero_riga_foglio = RIGA_INTESTAZIONE_ANAGRAFICA + 1 + idx
                                valori = df.loc[idx].to_dict()
                                valori["Gruppo"] = nome_gruppo_finale
                                ok, err_salva = salva_riga_anagrafica(workbook, valori,
                                                                       riga_da_aggiornare=numero_riga_foglio)
                                if not ok:
                                    errori.append(f"{nome}: {err_salva}")
                        for nome in selezionati:
                            st.session_state.pop(_chiave_cb(nome), None)
                        if errori:
                            st.error("Alcuni abbinamenti non sono riusciti:\n" + "\n".join(errori))
                        else:
                            st.cache_data.clear()
                            for nome in selezionati:
                                st.session_state.pop(_chiave_cb(nome), None)
                            st.session_state.gruppi_mostra_scelta = False
                            st.success(f"✔ {n_sel} Proclamatori abbinati a «{nome_gruppo_finale}».")

# ─────────────────────────────────────────────────────────────────
# PAGINA: Presenti alle adunanze
# ─────────────────────────────────────────────────────────────────
S88_CAMPI_RECT = {
    "1-Attendance_1": (161.84, 694.2, 221.65, 713.04),
    "1-Attendance_10": (161.84, 515.52, 221.65, 534.36),
    "1-Attendance_11": (161.84, 495.72, 221.65, 514.56),
    "1-Attendance_12": (161.84, 475.8, 221.65, 494.64),
    "1-Attendance_2": (161.84, 674.28, 221.65, 693.12),
    "1-Attendance_3": (161.84, 654.36, 221.65, 673.2),
    "1-Attendance_4": (161.84, 634.56, 221.65, 653.4),
    "1-Attendance_5": (161.84, 614.64, 221.65, 633.48),
    "1-Attendance_6": (161.84, 594.84, 221.65, 613.68),
    "1-Attendance_7": (161.84, 575.04, 221.65, 593.88),
    "1-Attendance_8": (161.84, 555.24, 221.65, 574.08),
    "1-Attendance_9": (161.84, 535.32, 221.65, 554.16),
    "1-Average_1": (223.57, 694.2, 296.88, 713.04),
    "1-Average_10": (223.57, 515.52, 296.88, 534.36),
    "1-Average_11": (223.57, 495.72, 296.88, 514.56),
    "1-Average_12": (223.57, 475.8, 296.88, 494.64),
    "1-Average_2": (223.57, 674.28, 296.88, 693.12),
    "1-Average_3": (223.57, 654.36, 296.88, 673.2),
    "1-Average_4": (223.57, 634.56, 296.88, 653.4),
    "1-Average_5": (223.57, 614.64, 296.88, 633.48),
    "1-Average_6": (223.57, 594.84, 296.88, 613.68),
    "1-Average_7": (223.57, 575.04, 296.88, 593.88),
    "1-Average_8": (223.57, 555.24, 296.88, 574.08),
    "1-Average_9": (223.57, 535.32, 296.88, 554.16),
    "1-Average_Total": (223.57, 455.88, 296.88, 474.72),
    "1-Meeting_1": (99.22, 694.2, 159.66, 713.04),
    "1-Meeting_10": (99.22, 515.52, 159.66, 534.36),
    "1-Meeting_11": (99.22, 495.72, 159.66, 514.56),
    "1-Meeting_12": (99.22, 475.8, 159.66, 494.64),
    "1-Meeting_2": (99.22, 674.28, 159.66, 693.12),
    "1-Meeting_3": (99.22, 654.36, 159.66, 673.2),
    "1-Meeting_4": (99.22, 634.56, 159.66, 653.4),
    "1-Meeting_5": (99.22, 614.64, 159.66, 633.48),
    "1-Meeting_6": (99.22, 594.84, 159.66, 613.68),
    "1-Meeting_7": (99.22, 575.04, 159.66, 593.88),
    "1-Meeting_8": (99.22, 555.24, 159.66, 574.08),
    "1-Meeting_9": (99.22, 535.32, 159.66, 554.16),
    "2-Attendance_1": (441.97, 694.2, 502.53, 713.04),
    "2-Attendance_10": (441.97, 515.52, 502.53, 534.36),
    "2-Attendance_11": (441.97, 495.72, 502.53, 514.56),
    "2-Attendance_12": (441.97, 475.8, 502.53, 494.64),
    "2-Attendance_2": (441.97, 674.28, 502.53, 693.12),
    "2-Attendance_3": (441.97, 654.36, 502.53, 673.2),
    "2-Attendance_4": (441.97, 634.56, 502.53, 653.4),
    "2-Attendance_5": (441.97, 614.64, 502.53, 633.48),
    "2-Attendance_6": (441.97, 594.84, 502.53, 613.68),
    "2-Attendance_7": (441.97, 575.04, 502.53, 593.88),
    "2-Attendance_8": (441.97, 555.24, 502.53, 574.08),
    "2-Attendance_9": (441.97, 535.32, 502.53, 554.16),
    "2-Average_1": (503.83, 694.2, 577.68, 713.04),
    "2-Average_10": (503.83, 515.52, 577.68, 534.36),
    "2-Average_11": (503.83, 495.72, 577.68, 514.56),
    "2-Average_12": (503.83, 475.8, 577.68, 494.64),
    "2-Average_2": (503.83, 674.28, 577.68, 693.12),
    "2-Average_3": (503.83, 654.36, 577.68, 673.2),
    "2-Average_4": (503.83, 634.56, 577.68, 653.4),
    "2-Average_5": (503.83, 614.64, 577.68, 633.48),
    "2-Average_6": (503.83, 594.84, 577.68, 613.68),
    "2-Average_7": (503.83, 575.04, 577.68, 593.88),
    "2-Average_8": (503.83, 555.24, 577.68, 574.08),
    "2-Average_9": (503.83, 535.32, 577.68, 554.16),
    "2-Average_Total": (503.83, 455.88, 577.68, 474.72),
    "2-Meeting_1": (380.04, 694.2, 440.1, 713.04),
    "2-Meeting_10": (380.04, 515.52, 440.1, 534.36),
    "2-Meeting_11": (380.04, 495.72, 440.1, 514.56),
    "2-Meeting_12": (380.04, 475.8, 440.1, 494.64),
    "2-Meeting_2": (380.04, 674.28, 440.1, 693.12),
    "2-Meeting_3": (380.04, 654.36, 440.1, 673.2),
    "2-Meeting_4": (380.04, 634.56, 440.1, 653.4),
    "2-Meeting_5": (380.04, 614.64, 440.1, 633.48),
    "2-Meeting_6": (380.04, 594.84, 440.1, 613.68),
    "2-Meeting_7": (380.04, 575.04, 440.1, 593.88),
    "2-Meeting_8": (380.04, 555.24, 440.1, 574.08),
    "2-Meeting_9": (380.04, 535.32, 440.1, 554.16),
    "3-Attendance_1": (161.84, 345.24, 221.65, 364.08),
    "3-Attendance_10": (161.84, 166.56, 221.65, 185.4),
    "3-Attendance_11": (161.84, 146.76, 221.65, 165.6),
    "3-Attendance_12": (161.84, 126.96, 221.65, 145.8),
    "3-Attendance_2": (161.84, 325.32, 221.65, 344.16),
    "3-Attendance_3": (161.84, 305.52, 221.65, 324.36),
    "3-Attendance_4": (161.84, 285.6, 221.65, 304.44),
    "3-Attendance_5": (161.84, 265.68, 221.65, 284.52),
    "3-Attendance_6": (161.84, 245.88, 221.65, 264.72),
    "3-Attendance_7": (161.84, 226.08, 221.65, 244.92),
    "3-Attendance_8": (161.84, 206.28, 221.65, 225.12),
    "3-Attendance_9": (161.84, 186.36, 221.65, 205.2),
    "3-Average_1": (223.57, 345.24, 296.88, 364.08),
    "3-Average_10": (223.57, 166.56, 296.88, 185.4),
    "3-Average_11": (223.57, 146.76, 296.88, 165.6),
    "3-Average_12": (223.57, 126.96, 296.88, 145.8),
    "3-Average_2": (223.57, 325.32, 296.88, 344.16),
    "3-Average_3": (223.57, 305.52, 296.88, 324.36),
    "3-Average_4": (223.57, 285.6, 296.88, 304.44),
    "3-Average_5": (223.57, 265.68, 296.88, 284.52),
    "3-Average_6": (223.57, 245.88, 296.88, 264.72),
    "3-Average_7": (223.57, 226.08, 296.88, 244.92),
    "3-Average_8": (223.57, 206.28, 296.88, 225.12),
    "3-Average_9": (223.57, 186.36, 296.88, 205.2),
    "3-Average_Total": (223.57, 106.92, 296.88, 125.76),
    "3-Meeting_1": (99.22, 345.24, 159.66, 364.08),
    "3-Meeting_10": (99.22, 166.56, 159.66, 185.4),
    "3-Meeting_11": (99.22, 146.76, 159.66, 165.6),
    "3-Meeting_12": (99.22, 126.96, 159.66, 145.8),
    "3-Meeting_2": (99.22, 325.32, 159.66, 344.16),
    "3-Meeting_3": (99.22, 305.52, 159.66, 324.36),
    "3-Meeting_4": (99.22, 285.6, 159.66, 304.44),
    "3-Meeting_5": (99.22, 265.68, 159.66, 284.52),
    "3-Meeting_6": (99.22, 245.88, 159.66, 264.72),
    "3-Meeting_7": (99.22, 226.08, 159.66, 244.92),
    "3-Meeting_8": (99.22, 206.28, 159.66, 225.12),
    "3-Meeting_9": (99.22, 186.36, 159.66, 205.2),
    "4-Attendance_1": (441.97, 345.24, 502.53, 364.08),
    "4-Attendance_10": (441.97, 166.56, 502.53, 185.4),
    "4-Attendance_11": (441.97, 146.76, 502.53, 165.6),
    "4-Attendance_12": (441.97, 126.96, 502.53, 145.8),
    "4-Attendance_2": (441.97, 325.32, 502.53, 344.16),
    "4-Attendance_3": (441.97, 305.52, 502.53, 324.36),
    "4-Attendance_4": (441.97, 285.6, 502.53, 304.44),
    "4-Attendance_5": (441.97, 265.68, 502.53, 284.52),
    "4-Attendance_6": (441.97, 245.88, 502.53, 264.72),
    "4-Attendance_7": (441.97, 226.08, 502.53, 244.92),
    "4-Attendance_8": (441.97, 206.28, 502.53, 225.12),
    "4-Attendance_9": (441.97, 186.36, 502.53, 205.2),
    "4-Average_1": (503.83, 345.24, 577.68, 364.08),
    "4-Average_10": (503.83, 166.56, 577.68, 185.4),
    "4-Average_11": (503.83, 146.76, 577.68, 165.6),
    "4-Average_12": (503.83, 126.13, 577.68, 144.97),
    "4-Average_2": (503.83, 325.32, 577.68, 344.16),
    "4-Average_3": (503.83, 305.52, 577.68, 324.36),
    "4-Average_4": (503.83, 285.6, 577.68, 304.44),
    "4-Average_5": (503.83, 265.68, 577.68, 284.52),
    "4-Average_6": (503.83, 245.88, 577.68, 264.72),
    "4-Average_7": (503.83, 226.08, 577.68, 244.92),
    "4-Average_8": (503.83, 205.45, 577.68, 224.29),
    "4-Average_9": (503.83, 186.36, 577.68, 205.2),
    "4-Average_Total": (503.83, 106.92, 577.68, 125.76),
    "4-Meeting_1": (380.04, 345.24, 440.1, 364.08),
    "4-Meeting_10": (380.04, 166.56, 440.1, 185.4),
    "4-Meeting_11": (380.04, 146.76, 440.1, 165.6),
    "4-Meeting_12": (380.04, 126.96, 440.1, 145.8),
    "4-Meeting_2": (380.04, 325.32, 440.1, 693.12),
    "4-Meeting_3": (380.04, 305.52, 440.1, 324.36),
    "4-Meeting_4": (380.04, 285.6, 440.1, 304.44),
    "4-Meeting_5": (380.04, 265.68, 440.1, 284.52),
    "4-Meeting_6": (380.04, 245.88, 440.1, 264.72),
    "4-Meeting_7": (380.04, 226.08, 440.1, 244.92),
    "4-Meeting_8": (380.04, 206.28, 440.1, 225.12),
    "4-Meeting_9": (380.04, 186.36, 440.1, 205.2),
    "Service Year_1": (18.84, 714.12, 97.92, 735.84),
    "Service Year_2": (299.28, 714.12, 378.36, 735.84),
    "Service Year_3": (18.84, 365.16, 97.92, 386.76),
    "Service Year_4": (299.28, 365.16, 378.36, 386.76),
}

S88_MESI_ORDINE_SERVIZIO = [9, 10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8]


def _s88_testo_centrato(c: rl_canvas.Canvas, testo: str, rect: tuple,
                          font_name: str = "Helvetica", font_size: float = 9.0):
    x0, y0, x1, y1 = rect
    largo_testo = c.stringWidth(testo, font_name, font_size)
    x = (x0 + x1) / 2 - largo_testo / 2
    fattore_altezza = 0.32
    y = (y0 + y1) / 2 - (font_size * fattore_altezza)
    c.setFont(font_name, font_size)
    c.drawString(x, y, testo)


def _s88_calcola_dati(df_presenze: pd.DataFrame, tipo_adunanza: str, anno_teocratico: int) -> dict:
    df = df_presenze.copy()
    df["_dt"] = pd.to_datetime(df["Data"], format="%d/%m/%Y", errors="coerce")
    df = df.dropna(subset=["_dt"])
    df["_presenza_num"] = df["In Presenza"].apply(a_float_it)
    df["_zoom_num"] = df["Su Zoom"].apply(a_float_it)
    if "Totale" in df.columns:
        df["_totale_num"] = df["Totale"].apply(a_float_it)
    else:
        df["_totale_num"] = df["_presenza_num"] + df["_zoom_num"]
    df = df[df["Tipo Adunanza"] == tipo_adunanza]

    mesi = []
    medie_valide = []
    for mese in S88_MESI_ORDINE_SERVIZIO:
        anno_calendario = anno_teocratico if mese >= 9 else anno_teocratico + 1
        sotto = df[(df["_dt"].dt.year == anno_calendario) & (df["_dt"].dt.month == mese)]
        n = len(sotto)
        totale = sotto["_totale_num"].sum()
        media = (totale / n) if n else None
        mesi.append({"numero": n, "totale": totale, "media": media})
        if media is not None:
            medie_valide.append(media)
    media_finale = (sum(medie_valide) / len(medie_valide)) if medie_valide else None
    return {"mesi": mesi, "media_finale": media_finale}


def genera_pdf_s88(df_presenze: pd.DataFrame) -> bytes:
    oggi = datetime.now()
    anno_corrente = anno_teocratico_di(f"{oggi.year}-{oggi.month:02d}")
    anno_precedente = anno_corrente - 1
    etichetta_corrente = f"{anno_corrente}-{anno_corrente + 1}"
    etichetta_precedente = f"{anno_precedente}-{anno_precedente + 1}"

    mappa_blocchi = {
        1: ("Infrasettimanale", anno_precedente, etichetta_precedente),
        2: ("Infrasettimanale", anno_corrente, etichetta_corrente),
        3: ("Fine settimana", anno_precedente, etichetta_precedente),
        4: ("Fine settimana", anno_corrente, etichetta_corrente),
    }

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=(S88_PAGE_W, S88_PAGE_H))
    for blocco, (tipo, anno, etichetta_anno) in mappa_blocchi.items():
        rect_anno = S88_CAMPI_RECT.get(f"Service Year_{blocco}")
        if rect_anno:
            _s88_testo_centrato(c, etichetta_anno, rect_anno, "Helvetica-Bold", 10.0)

        dati = _s88_calcola_dati(df_presenze, tipo, anno)
        for i, dati_mese in enumerate(dati["mesi"], start=1):
            if dati_mese["numero"] > 0:
                _s88_testo_centrato(c, str(dati_mese["numero"]), S88_CAMPI_RECT[f"{blocco}-Meeting_{i}"])
                _s88_testo_centrato(c, str(int(dati_mese["totale"])), S88_CAMPI_RECT[f"{blocco}-Attendance_{i}"])
                _s88_testo_centrato(c, formatta_numero_it(dati_mese["media"]), S88_CAMPI_RECT[f"{blocco}-Average_{i}"])
        if dati["media_finale"] is not None:
            _s88_testo_centrato(c, formatta_numero_it(dati["media_finale"]),
                                 S88_CAMPI_RECT[f"{blocco}-Average_Total"], "Helvetica-Bold", 9.5)
    c.save()
    buf.seek(0)

    overlay_reader = PdfReader(buf)
    template_reader = PdfReader(PERCORSO_MODULO_S88)
    writer = PdfWriter()

    pagina = template_reader.pages[0]
    pagina.merge_page(overlay_reader.pages[0])
    if "/Annots" in pagina:
        del pagina["/Annots"]
    writer.add_page(pagina)

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


# ─────────────────────────────────────────────────────────────────
def _presenze_campi_form(chiave_prefix: str, data_default, tipo_default: str,
                         presenza_default: int, zoom_default: int, giorni_per_tipo: dict,
                         disabled: bool = False):
    giorni_validi = giorni_adunanze_tutti(giorni_per_tipo)
    chiave_data = f"{chiave_prefix}_data"
    chiave_tipo = f"{chiave_prefix}_tipo"

    if chiave_data not in st.session_state:
        st.session_state[chiave_data] = data_default
    if chiave_tipo not in st.session_state:
        st.session_state[chiave_tipo] = tipo_default if tipo_default in TIPI_ADUNANZA else TIPI_ADUNANZA[0]

    data_valida, data_proposta = _prossima_data_valida_precedente(
        st.session_state[chiave_data], giorni_validi)
    if not data_valida and data_proposta is not None:
        st.session_state[chiave_data] = data_proposta
        data_valida = True

    chiave_ultima_data = f"{chiave_prefix}_ultima_data_vista"
    if st.session_state.get(chiave_ultima_data) != st.session_state[chiave_data]:
        st.session_state[chiave_ultima_data] = st.session_state[chiave_data]
        tipo_suggerito = tipo_adunanza_del_giorno(
            GIORNI_SETTIMANA_IT[st.session_state[chiave_data].weekday()], giorni_per_tipo)
        if tipo_suggerito:
            st.session_state[chiave_tipo] = tipo_suggerito

    data_scelta = st.date_input("Data", format="DD/MM/YYYY", key=chiave_data, disabled=disabled)

    if not data_valida and not giorni_validi:
        st.warning("Nessun giorno di adunanza configurato — impostalo nella card ⚙️ Impostazioni "
                   "in Home per attivare il controllo sulla data.")

    tipo_adunanza = st.selectbox("Tipo di adunanza", TIPI_ADUNANZA, key=chiave_tipo, disabled=disabled)

    col_p, col_z = st.columns(2)
    with col_p:
        in_presenza = st.number_input("In presenza", min_value=0, step=1, value=presenza_default,
                                       key=f"{chiave_prefix}_presenza", disabled=disabled)
    with col_z:
        su_zoom = st.number_input("Su Zoom", min_value=0, step=1, value=zoom_default,
                                   key=f"{chiave_prefix}_zoom", disabled=disabled)

    totale = int(in_presenza) + int(su_zoom)
    st.metric("Totale (calcolato)", totale)

    return data_scelta, tipo_adunanza, in_presenza, su_zoom, totale, data_valida


def _form_modifica_presenza(dati_selezione: dict):
    riga = dati_selezione["riga"]
    numero_riga_foglio = dati_selezione["numero_riga_foglio"]

    st.markdown("#### ✏️ Modifica presenza")

    try:
        data_default = datetime.strptime(str(riga.get("Data", "")), "%d/%m/%Y").date()
    except Exception:
        data_default = datetime.now().date()

    giorni_per_tipo = leggi_giorni_adunanze_per_tipo(workbook)
    chiave_prefix = f"presenze_mod_{numero_riga_foglio}"
    data_scelta, tipo_adunanza, in_presenza, su_zoom, totale, data_valida = _presenze_campi_form(
        chiave_prefix, data_default, riga.get("Tipo Adunanza", ""),
        int(a_float_it(riga.get("In Presenza", "0"))), int(a_float_it(riga.get("Su Zoom", "0"))),
        giorni_per_tipo, disabled=sola_lettura(),
    )

    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        invia = st.button("✔ Salva", type="primary", use_container_width=True,
                          disabled=(not data_valida) or sola_lettura(), key=f"{chiave_prefix}_salva")
    with col_btn2:
        annulla = st.button("✖ Annulla", use_container_width=True, key=f"{chiave_prefix}_annulla")

    if annulla:
        st.session_state.presenze_modifica = None
        st.session_state.presenze_tabella_versione = st.session_state.get("presenze_tabella_versione", 0) + 1
        st.rerun()

    if invia:
        valori = {
            "Data": data_scelta.strftime("%d/%m/%Y"),
            "Tipo Adunanza": tipo_adunanza,
            "In Presenza": str(int(in_presenza)),
            "Su Zoom": str(int(su_zoom)),
            "Totale": str(int(totale)),
        }
        ok, err_salva = salva_riga_foglio(
            workbook, NOME_FOGLIO_PRESENZE, RIGA_INTESTAZIONE_PRESENZE,
            valori, riga_da_aggiornare=numero_riga_foglio
        )
        if ok:
            st.cache_data.clear()
            st.session_state.presenze_modifica = None
            st.session_state.presenze_tabella_versione = st.session_state.get("presenze_tabella_versione", 0) + 1
            st.success("✔ Modificato correttamente.")
            st.rerun()
        else:
            st.error(err_salva)


def _form_nuova_presenza():
    st.markdown("#### ➕ Nuova presenza")

    giorni_per_tipo = leggi_giorni_adunanze_per_tipo(workbook)
    chiave_prefix = "presenze_nuovo"
    data_scelta, tipo_adunanza, in_presenza, su_zoom, totale, data_valida = _presenze_campi_form(
        chiave_prefix, datetime.now().date(), TIPI_ADUNANZA[0], 0, 0, giorni_per_tipo,
        disabled=sola_lettura(),
    )

    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        invia = st.button("✔ Salva", type="primary", use_container_width=True,
                          disabled=(not data_valida) or sola_lettura(), key=f"{chiave_prefix}_salva")
    with col_btn2:
        annulla = st.button("✖ Annulla", use_container_width=True, key=f"{chiave_prefix}_annulla")

    if annulla:
        st.session_state.presenze_form_nuovo_aperto = False
        for suffisso in ("_data", "_tipo", "_presenza", "_zoom", "_ultima_data_vista"):
            st.session_state.pop(f"{chiave_prefix}{suffisso}", None)
        st.rerun()

    if invia:
        valori = {
            "Data": data_scelta.strftime("%d/%m/%Y"),
            "Tipo Adunanza": tipo_adunanza,
            "In Presenza": str(int(in_presenza)),
            "Su Zoom": str(int(su_zoom)),
            "Totale": str(int(totale)),
        }
        ok, err_salva = salva_riga_foglio(
            workbook, NOME_FOGLIO_PRESENZE, RIGA_INTESTAZIONE_PRESENZE, valori
        )
        if ok:
            st.cache_data.clear()
            st.session_state.presenze_form_nuovo_aperto = False
            for suffisso in ("_data", "_tipo", "_presenza", "_zoom", "_ultima_data_vista"):
                st.session_state.pop(f"{chiave_prefix}{suffisso}", None)
            st.success("✔ Presenza inserita correttamente.")
            st.rerun()
        else:
            st.error(err_salva)


# ─────────────────────────────────────────────────────────────────
# FUNZIONE PRINCIPALE PAGINA PRESENZE
# ─────────────────────────────────────────────────────────────────
def mostra_presenze_adunanze():
    st.title("🙌 Presenti alle adunanze")

    is_modalita_ristretta = (st.query_params.get("page") == "presenze" or st.query_params.get("modalita") == "presenze")

    if not is_modalita_ristretta:
        st.button("🏠 Torna alla Home", key="home_da_presenze", use_container_width=True,
                  on_click=vai_a, args=("home",))

    if not collegato:
        st.warning("⚠️ Nessun foglio dati collegato.")
        return

    if st.button("➕ Inserisci presenti alle adunanze", key="apri_nuova_presenza", use_container_width=True):
        st.session_state.presenze_form_nuovo_aperto = not st.session_state.get(
            "presenze_form_nuovo_aperto", False)
        st.session_state.presenze_modifica = None
    if st.session_state.get("presenze_form_nuovo_aperto"):
        _form_nuova_presenza()

    if st.session_state.get("presenze_modifica"):
        _form_modifica_presenza(st.session_state.presenze_modifica)

    df, err = leggi_foglio_come_df(workbook, NOME_FOGLIO_PRESENZE, RIGA_INTESTAZIONE_PRESENZE)
    if err:
        st.error(err)
        return

    if df.empty:
        st.info("Nessuna presenza registrata ancora.")
        return

    if not os.path.exists(PERCORSO_MODULO_S88):
        st.warning("Modulo S-88 non trovato: metti il file «S-88_I.pdf» nella stessa cartella di app.py "
                   "per poterlo generare.")
    else:
        if st.button("📄 Genera modulo S-88", key="genera_s88", use_container_width=True):
            with st.spinner("Genero il modulo S-88…"):
                st.session_state.s88_pdf_pronto = genera_pdf_s88(df)

        if st.session_state.get("s88_pdf_pronto"):
            st.download_button(
                "⬇️ Scarica modulo S-88 (PDF)",
                data=st.session_state.s88_pdf_pronto,
                file_name="Modulo_S-88.pdf",
                mime="application/pdf",
                key="download_s88",
                use_container_width=True,
                on_click=lambda: st.session_state.pop("s88_pdf_pronto", None),
            )

    df_prep = df.copy()
    df_prep["_dt"] = pd.to_datetime(df_prep["Data"], format="%d/%m/%Y", errors="coerce")
    df_prep = df_prep.dropna(subset=["_dt"])
    df_prep["_anno_mese"] = df_prep["_dt"].dt.strftime("%Y-%m")
    df_prep["_riga_foglio"] = RIGA_INTESTAZIONE_PRESENZE + 1 + df_prep.index

    mesi_disponibili = sorted(df_prep["_anno_mese"].unique(), reverse=True)

    if not mesi_disponibili:
        st.warning("Nessuna data valida trovata nel foglio.")
        return

    st.markdown("### 📅 Seleziona Mese/Anno")
    mese_selezionato = st.selectbox(
        "Mese/Anno",
        options=mesi_disponibili,
        label_visibility="collapsed",
        key="select_anno_mese"
    )

    if "presenze_tabella_versione" not in st.session_state:
        st.session_state.presenze_tabella_versione = 0

    if st.session_state.get("presenze_ultimo_mese_visto") != mese_selezionato:
        st.session_state.presenze_ultimo_mese_visto = mese_selezionato
        st.session_state.presenze_tabella_versione += 1

    df_mese = df_prep[df_prep["_anno_mese"] == mese_selezionato].copy()

    df_mese["_presenza_num"] = df_mese["In Presenza"].apply(a_float_it)
    df_mese["_zoom_num"] = df_mese["Su Zoom"].apply(a_float_it)

    if "Totale" in df_mese.columns:
        df_mese["_totale_num"] = df_mese["Totale"].apply(a_float_it)
    else:
        df_mese["_totale_num"] = df_mese["_presenza_num"] + df_mese["_zoom_num"]

    st.markdown("#### 📊 Riepilogo")

    dati_riepilogo = []
    tipi_list = ["Infrasettimanale", "Fine settimana"]

    for tipo in tipi_list:
        sotto = df_mese[df_mese["Tipo Adunanza"] == tipo]

        settimane = len(sotto)
        totale = sotto["_totale_num"].sum()
        tot_presenza = sotto["_presenza_num"].sum()
        tot_zoom = sotto["_zoom_num"].sum()

        media = (totale / settimane) if settimane > 0 else 0
        perc_zoom = (tot_zoom / totale * 100) if totale > 0 else 0
        perc_presenza = (tot_presenza / totale * 100) if totale > 0 else 0

        dati_riepilogo.append({
            "Tipo Adunanza": tipo,
            "Sett.": settimane,
            "Media": f"{media:,.2f}".replace(".", ","),
            "Totale": int(totale),
            "% Zoom": f"{perc_zoom:.2f}%".replace(".", ","),
            "% Pres.": f"{perc_presenza:.2f}%".replace(".", ","),
        })

    st.dataframe(
        pd.DataFrame(dati_riepilogo),
        hide_index=True,
        use_container_width=True
    )

    st.markdown(f"#### 📋 Dettaglio adunanze per {mese_selezionato}")

    colonne_visibili = [c for c in df_mese.columns if not c.startswith("_")]

    config_colonne = {
        col: st.column_config.Column(alignment="center")
        for col in colonne_visibili
    }

    df_mese_reset = df_mese.reset_index(drop=True)
    chiave_tabella_dettaglio = f"presenze_tabella_dettaglio_{st.session_state.presenze_tabella_versione}"
    evento_dettaglio = st.dataframe(
        df_mese_reset[colonne_visibili],
        hide_index=True,
        use_container_width=True,
        on_select="rerun",
        selection_mode="single-row",
        key=chiave_tabella_dettaglio,
        column_config=config_colonne,
    )

    righe_sel_dett = evento_dettaglio.selection.rows if evento_dettaglio and evento_dettaglio.selection else []
    righe_sel_dett = [i for i in righe_sel_dett if i < len(df_mese_reset)]

    if righe_sel_dett:
        riga_scelta = df_mese_reset.loc[righe_sel_dett[0]]
        numero_riga_foglio = int(riga_scelta["_riga_foglio"])

        col_mod, col_elim = st.columns(2)
        with col_mod:
            if st.button("✏️ Modifica riga selezionata", key="presenze_btn_mod", use_container_width=True):
                st.session_state.presenze_form_nuovo_aperto = False
                st.session_state.presenze_modifica = {
                    "riga": riga_scelta.to_dict(),
                    "numero_riga_foglio": numero_riga_foglio,
                }
                st.rerun()
        with col_elim:
            if st.button("🗑️ Elimina riga selezionata", key="presenze_btn_elim", use_container_width=True,
                         disabled=sola_lettura()):
                st.session_state.presenze_conferma_elimina = numero_riga_foglio
                st.rerun()

        if st.session_state.get("presenze_conferma_elimina") == numero_riga_foglio:
            st.warning("Sei sicuro di voler eliminare questa riga? L'operazione non è reversibile.")
            col_si, col_no = st.columns(2)
            with col_si:
                if st.button("✔ Sì, elimina", key="presenze_conf_si", type="primary", use_container_width=True):
                    ok, err_elim = elimina_riga_foglio(workbook, NOME_FOGLIO_PRESENZE, numero_riga_foglio)
                    if ok:
                        st.cache_data.clear()
                        st.session_state.presenze_conferma_elimina = None
                        st.session_state.presenze_tabella_versione += 1
                        st.success("✔ Riga eliminata.")
                        st.rerun()
                    else:
                        st.error(err_elim)
            with col_no:
                if st.button("No, annulla", key="presenze_conf_no", use_container_width=True):
                    st.session_state.presenze_conferma_elimina = None
                    st.rerun()
    else:
        st.session_state.presenze_conferma_elimina = None
# ─────────────────────────────────────────────────────────────────
# PAGINA: IMPORTA DA S-21 (Proclamatore trasferito o nuovo)
# ─────────────────────────────────────────────────────────────────

def normalizza_mese_anno(valore):
    """
    Estrae e formatta una o più date nel formato YYYY-MM come TESTO PURO.
    Garantisce che mesi ad una sola cifra (es. 2026-5) diventino '2026-05'.
    Usata SOLO nella pagina Importa S-21: a differenza di
    '_normalizza_data_report' (usata per il controllo rapporti in Home),
    questa può ritornare una LISTA se nel testo sono presenti più date.
    """
    if pd.isna(valore) or valore is None:
        return ""

    valore_str = str(valore).strip()
    if not valore_str:
        return ""

    if len(valore_str) >= 10 and valore_str[4] == '-' and valore_str[7] == '-':
        valore_str = valore_str[:7]

    pattern = r'(\d{4})-(\d{1,2})'
    matches = re.findall(pattern, valore_str)

    if matches:
        date_formattate = [f"{anno}-{int(mese):02d}" for anno, mese in matches]
        return date_formattate[0] if len(date_formattate) == 1 else date_formattate

    return valore_str


# ─────────────────────────────────────────────────────────────────
# PAGINA: IMPORTA DA S-21
# ─────────────────────────────────────────────────────────────────
def mostra_importa_s21():
    st.title("📥 Importa da S-21")

    col_home, col_manuale = st.columns([1, 1])
    with col_home:
        st.button("🏠 Torna alla Home", key="home_da_importa_s21", use_container_width=True,
                  on_click=vai_a_home_reset_importa_s21)

    with col_manuale:
        if st.button("➕ Inserisci storico manualmente", key="s21_apri_manuale", use_container_width=True, type="primary"):
            st.session_state.s21_form_manuale_aperto = not st.session_state.get("s21_form_manuale_aperto", False)

    if not collegato:
        st.warning("⚠️ Nessun foglio dati collegato.")
        return

    if st.session_state.get("s21_form_manuale_aperto", False):
        st.markdown("---")
        st.subheader("📝 Inserimento/Modifica Storico S-21 Manuale")

        df_anagrafica, err_anag = leggi_foglio_come_df(workbook, NOME_FOGLIO_ANAGRAFICA, RIGA_INTESTAZIONE_ANAGRAFICA)
        if err_anag:
            st.error(err_anag)
            return

        df_tutti, err_tutti = leggi_foglio_tutti(workbook)
        if err_tutti:
            st.error(err_tutti)
            return

        opzione_nuova = "➕ Nuova persona (non ancora in Anagrafica)"
        opzioni_proclamatori = [""] + [opzione_nuova]
        mappa_nomi_reali = {}
        mappa_dati_anagrafici = {}

        for _, riga in df_anagrafica.iterrows():
            nome = str(riga.get("Cognome e Nome", "")).strip()
            if not nome:
                continue
            stato_cod = str(riga.get("Attivi / Inattivi", "")).strip().upper()

            if stato_cod in ("I", "INATTIVO"):
                etichetta = f"⛔ {nome} (Inattivo)"
            elif stato_cod in ("TR", "TRASFERITO"):
                etichetta = f"🚚 {nome} (Trasferito)"
            else:
                etichetta = f"🟢 {nome}"

            opzioni_proclamatori.append(etichetta)
            mappa_nomi_reali[etichetta] = nome
            mappa_dati_anagrafici[nome] = {
                "gruppo": str(riga.get("Gruppo", "")).strip(),
                "tipo": str(riga.get("Tipo", "")).strip()
            }

        scelta_etichetta = st.selectbox("Abbina al Proclamatore:", opzioni_proclamatori, key="s21_man_persona")
        nuova_persona_flag = (scelta_etichetta == opzione_nuova)

        if nuova_persona_flag:
            nome_finale = st.text_input("Nome e Cognome nuovo proclamatore:", key="s21_man_nome_nuovo")
            default_tipo = ""
            default_gruppo = ""
        else:
            nome_finale = mappa_nomi_reali.get(scelta_etichetta, "")
            dati_rec = mappa_dati_anagrafici.get(nome_finale, {})
            default_tipo = dati_rec.get("tipo", "")
            default_gruppo = dati_rec.get("gruppo", "")

        opzioni_tipo_servizio = ["", "Proclamatore", "Pioniere Regolare", "Pioniere Speciale", "Missionario sul campo"]
        idx_tipo = opzioni_tipo_servizio.index(default_tipo) if default_tipo in opzioni_tipo_servizio else 0

        tipo_servizio_scelto = st.selectbox(
            "Tipo Servizio (opzionale):",
            opzioni_tipo_servizio,
            index=idx_tipo,
            key=f"s21_man_tipo_servizio_{nome_finale}"
        )

        pion_aus_disabilitato = (tipo_servizio_scelto != "Proclamatore")

        gruppi_disponibili = [""] + sorted({str(g).strip() for g in df_anagrafica.get("Gruppo", pd.Series(dtype=str)) if str(g).strip()})
        idx_gruppo = gruppi_disponibili.index(default_gruppo) if default_gruppo in gruppi_disponibili else 0

        col_grp, col_st = st.columns(2)
        with col_grp:
            sorvegliante_gruppo = st.selectbox(
                "Sorvegliante del gruppo:",
                gruppi_disponibili,
                index=idx_gruppo,
                key=f"s21_man_gruppo_{nome_finale}"
            )
        with col_st:
            st.text_input("Stato", value="Attivo", disabled=True, key="s21_man_stato_vis")

        campi_compilati = (
            bool(scelta_etichetta) and
            bool(nome_finale.strip()) and
            bool(tipo_servizio_scelto) and
            bool(sorvegliante_gruppo)
        )

        if not campi_compilati:
            st.info("📌 Per abilitare e compilare la griglia **Dati di Servizio**, completa prima tutti i campi sopra (Proclamatore, Tipo Servizio e Sorvegliante del gruppo).")
            if st.button("✖ Annulla", use_container_width=True, key="s21_man_annulla_bloc"):
                st.session_state.s21_form_manuale_aperto = False
                st.rerun()
        else:
            st.markdown("##### 📅 Dati di Servizio (Schema S-21)")

            anno_corrente = datetime.now().year
            anno_servizio = st.number_input("Anno di Servizio:", min_value=2000, max_value=2099, value=anno_corrente, step=1, key="s21_man_anno")

            mesi_s21 = [
                f"{anno_servizio-1}-09", f"{anno_servizio-1}-10", f"{anno_servizio-1}-11", f"{anno_servizio-1}-12",
                f"{anno_servizio}-01", f"{anno_servizio}-02", f"{anno_servizio}-03", f"{anno_servizio}-04",
                f"{anno_servizio}-05", f"{anno_servizio}-06", f"{anno_servizio}-07", f"{anno_servizio}-08"
            ]

            mappa_esistenti = {}
            if not df_tutti.empty and nome_finale.strip():
                col_nome_tutti = None
                for c in df_tutti.columns:
                    if str(c).strip().lower() in ("cognome e nome", "nome", "proclamatore"):
                        col_nome_tutti = c
                        break
                if not col_nome_tutti and len(df_tutti.columns) > 1:
                    col_nome_tutti = df_tutti.columns[1]

                col_mese_tutti = None
                for c in df_tutti.columns:
                    if str(c).strip().lower() in ("mese", "mese/anno"):
                        col_mese_tutti = c
                        break
                if not col_mese_tutti and len(df_tutti.columns) > 2:
                    col_mese_tutti = df_tutti.columns[2]

                if col_nome_tutti and col_mese_tutti:
                    df_persona = df_tutti[df_tutti[col_nome_tutti].astype(str).str.strip().str.lower() == nome_finale.strip().lower()]
                    for idx, riga in df_persona.iterrows():
                        m_a = str(riga.get(col_mese_tutti, "")).strip()
                        m_a_norm = normalizza_mese_anno(m_a)
                        if isinstance(m_a_norm, list):
                            m_a_norm = m_a_norm[0]

                        if m_a_norm:
                            valore_serv = str(riga.get("Servizio", riga.get("Ha partecipato al ministero", ""))).strip().lower()
                            mappa_esistenti[m_a_norm] = {
                                "index_df": idx,
                                "partecipato": valore_serv in ("si", "sì", "1", "true"),
                                "tipo": str(riga.get("Hai servito come ?", riga.get("Tipo Servizio", ""))).strip(),
                                "ore": pd.to_numeric(riga.get("Ore", 0), errors="coerce"),
                                "studi": int(pd.to_numeric(riga.get("Studi Biblici", 0), errors="coerce") or 0),
                                "osservazioni": str(riga.get("Commenti:", riga.get("Osservazioni", ""))).strip()
                            }

            righe_griglia = []
            for m in mesi_s21:
                m_norm = normalizza_mese_anno(m)
                if isinstance(m_norm, list):
                    m_norm = m_norm[0]

                if m_norm in mappa_esistenti:
                    dati_m = mappa_esistenti[m_norm]
                    righe_griglia.append({
                        "Mese/Anno": m_norm,
                        "Partecipato": dati_m["partecipato"],
                        "Ore": dati_m["ore"] if pd.notna(dati_m["ore"]) else 0,
                        "Studi Biblici": dati_m["studi"],
                        "Pioniere Ausiliario": False if pion_aus_disabilitato else (dati_m["tipo"] == "Pioniere Ausiliario"),
                        "Osservazioni": dati_m["osservazioni"]
                    })
                else:
                    righe_griglia.append({
                        "Mese/Anno": m_norm,
                        "Partecipato": False,
                        "Ore": 0,
                        "Studi Biblici": 0,
                        "Pioniere Ausiliario": False,
                        "Osservazioni": ""
                    })

            df_iniziale = pd.DataFrame(righe_griglia)

            tabella_manuale = st.data_editor(
                df_iniziale,
                hide_index=True,
                use_container_width=True,
                height=480,
                num_rows="dynamic",
                key=f"s21_manuale_editor_{nome_finale}_{anno_servizio}_{tipo_servizio_scelto}",
                disabled=sola_lettura(),
                column_config={
                    "Mese/Anno": st.column_config.TextColumn("Mese/Anno (AAAA-MM)", width="small", alignment="center"),
                    "Partecipato": st.column_config.CheckboxColumn("Ministero", width="small"),
                    "Ore": st.column_config.NumberColumn("Ore", width="small", min_value=0, step=1, alignment="center"),
                    "Studi Biblici": st.column_config.NumberColumn("Studi", width="small", min_value=0, step=1, alignment="center"),
                    "Pioniere Ausiliario": st.column_config.CheckboxColumn("Pion. Aus.", width="small", disabled=pion_aus_disabilitato),
                    "Osservazioni": st.column_config.TextColumn("Commenti:", width="large")
                }
            )

            col_salva, col_annulla = st.columns(2)
            with col_salva:
                btn_salva = st.button("✔ Salva", type="primary", use_container_width=True,
                                       key="s21_man_salva", disabled=sola_lettura())
            with col_annulla:
                btn_annulla = st.button("✖ Annulla", use_container_width=True, key="s21_man_annulla")

            if btn_annulla:
                st.session_state.s21_form_manuale_aperto = False
                st.rerun()

            if btn_salva:
                nome_pulito = nome_finale.strip()
                if not nome_pulito:
                    st.error("⚠️ Inserisci o seleziona un proclamatore valido.")
                else:
                    nomi_esistenti_anag = [str(x).strip().lower() for x in df_anagrafica.get("Cognome e Nome", pd.Series(dtype=str)) if str(x).strip()]

                    if nuova_persona_flag or (nome_pulito.lower() not in nomi_esistenti_anag):
                        if nome_pulito.lower() in nomi_esistenti_anag:
                            st.warning(f"⚠️ **{nome_pulito}** è già presente in Anagrafica. Inserimento anagrafico saltato.")
                        else:
                            id_massimo = 0
                            if "ID" in df_anagrafica.columns:
                                ids_numerici = pd.to_numeric(df_anagrafica["ID"], errors="coerce").dropna()
                                if not ids_numerici.empty:
                                    id_massimo = int(ids_numerici.max())
                            nuovo_id = id_massimo + 1

                            nuovo_rec = {
                                "ID": nuovo_id,
                                "Cognome e Nome": nome_pulito,
                                "Tipo": tipo_servizio_scelto,
                                "Gruppo": sorvegliante_gruppo,
                                "Attivi / Inattivi": "A"
                            }

                            try:
                                salva_riga_anagrafica(workbook, nuovo_rec, None)
                            except TypeError:
                                salva_riga_anagrafica(workbook, nuovo_rec)

                    aggiornati = 0
                    inseriti = 0

                    data_ora_consegna = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

                    for _, r in tabella_manuale.iterrows():
                        if not r["Partecipato"]:
                            continue

                        m_a_raw = r["Mese/Anno"]
                        m_a = normalizza_mese_anno(m_a_raw)
                        if isinstance(m_a, list):
                            m_a = m_a[0]
                        if not m_a:
                            continue

                        part = "Si"
                        is_pion_aus = False if pion_aus_disabilitato else bool(r["Pioniere Ausiliario"])
                        tipo = "Pioniere Ausiliario" if is_pion_aus else tipo_servizio_scelto

                        if tipo == "Proclamatore":
                            valore_ore = ""
                        else:
                            valore_ore = str(int(r["Ore"])) if pd.notna(r["Ore"]) and r["Ore"] > 0 else ""

                        valore_studi = str(int(r["Studi Biblici"])) if pd.notna(r["Studi Biblici"]) else "0"

                        valori_riga = {
                            "Informazioni cronologiche": data_ora_consegna,
                            "Cognome e Nome": nome_pulito,
                            "Mese": str(m_a),
                            "Hai servito come ?": tipo,
                            "Servizio": part,
                            "Ore": valore_ore,
                            "Cred. Ore": "",
                            "Studi Biblici": valore_studi,
                            "Commenti:": str(r["Osservazioni"]).strip(),
                            "Sorvegliante del gruppo": sorvegliante_gruppo,
                            "ND": "ND"
                        }

                        if m_a in mappa_esistenti:
                            riga_idx = mappa_esistenti[m_a]["index_df"]
                            riga_excel_numero = RIGA_INTESTAZIONE_TUTTI + 1 + riga_idx
                            try:
                                ok, _ = salva_riga_foglio(workbook, NOME_FOGLIO_TUTTI, RIGA_INTESTAZIONE_TUTTI, valori_riga, riga_excel_numero)
                            except TypeError:
                                try:
                                    ok, _ = salva_riga_foglio(workbook, NOME_FOGLIO_TUTTI, RIGA_INTESTAZIONE_TUTTI, valori_riga, riga_numero=riga_excel_numero)
                                except TypeError:
                                    ok, _ = salva_riga_foglio(workbook, NOME_FOGLIO_TUTTI, RIGA_INTESTAZIONE_TUTTI, valori_riga)
                            if ok:
                                aggiornati += 1
                        else:
                            try:
                                ok, _ = salva_riga_foglio(workbook, NOME_FOGLIO_TUTTI, RIGA_INTESTAZIONE_TUTTI, valori_riga, None)
                            except TypeError:
                                try:
                                    ok, _ = salva_riga_foglio(workbook, NOME_FOGLIO_TUTTI, RIGA_INTESTAZIONE_TUTTI, valori_riga, riga_numero=None)
                                except TypeError:
                                    ok, _ = salva_riga_foglio(workbook, NOME_FOGLIO_TUTTI, RIGA_INTESTAZIONE_TUTTI, valori_riga)
                            if ok:
                                inseriti += 1

                    st.cache_data.clear()
                    st.success("✔️ S-21 Salvati correttamente")
                    st.session_state.s21_form_manuale_aperto = False
                    st.rerun()

        st.markdown("---")

    st.caption("Carica la S-21 ricevuta da un'altra congregazione per un Proclamatore trasferito (o mai "
               "appartenuto qui): provo a leggere automaticamente ore, studi e mesi dal PDF, ma "
               "**niente viene scritto in archivio finché non confermi tu**, riga per riga.")

    file_caricato = st.file_uploader("Carica il PDF della S-21 ricevuta", type=["pdf"], key="importa_s21_file")
    if file_caricato is None:
        return

    chiave_file = f"{file_caricato.name}_{file_caricato.size}"
    if st.session_state.get("importa_s21_chiave_file") != chiave_file:
        with st.spinner("Leggo il PDF…"):
            try:
                st.session_state.importa_s21_dati = _s21_estrai_dati_pdf(io.BytesIO(file_caricato.getvalue()))
                st.session_state.importa_s21_chiave_file = chiave_file
                st.session_state.pop("importa_s21_persona_scelta", None)
                st.session_state.pop("importa_s21_tabella_editor", None)
            except Exception as e:
                st.error(f"Non sono riuscito a leggere questo PDF: {e}")
                st.session_state.pop("importa_s21_dati", None)
                return

    dati_estratti = st.session_state.get("importa_s21_dati")
    if not dati_estratti:
        return

    if not dati_estratti["testo_rilevato"]:
        st.warning("⚠️ Non ho trovato testo leggibile nelle caselle dati di questo PDF — probabilmente "
                   "è una scansione o una foto di un modulo compilato a mano. Non tento la lettura "
                   "automatica della scrittura a mano (troppo rischiosa per dati ufficiali di servizio): "
                   "apri il PDF a fianco su un'altra finestra, leggi i valori a occhio e inseriscili "
                   "tu nella tabella qui sotto (puoi aggiungere righe con il tasto '+' in fondo).")
    else:
        st.success("✔ Testo letto correttamente dal PDF. Controlla comunque ogni valore prima di "
                   "confermare — soprattutto \"Tipo Servizio\", che è solo una proposta automatica.")

    st.subheader("1. Di chi è questa S-21?")
    nomi_letti = [p["nome"].strip() for p in dati_estratti["pannelli"] if p["nome"].strip()]
    nome_suggerito = nomi_letti[0] if nomi_letti else ""
    if nome_suggerito:
        st.caption(f"Nome letto dal modulo: **{nome_suggerito}**")
    else:
        st.caption("Nessun nome letto dal modulo (probabile scansione): scegli o cerca la persona a mano.")

    df_anagrafica, err = leggi_foglio_come_df(workbook, NOME_FOGLIO_ANAGRAFICA, RIGA_INTESTAZIONE_ANAGRAFICA)
    if err:
        st.error(err)
        return

    candidati = _importa_s21_cerca_persona(df_anagrafica, nome_suggerito)
    opzione_nuova = "➕ Nuova persona (non ancora in Anagrafica)"
    opzioni = [opzione_nuova] + candidati["Cognome e Nome"].tolist()
    if not df_anagrafica.empty:
        altri = [n for n in df_anagrafica["Cognome e Nome"].tolist() if n not in opzioni]
        opzioni += altri
    scelta_nome = st.selectbox("Abbina al Proclamatore:", opzioni, key="importa_s21_persona_scelta")

    gruppi_disponibili = sorted({g.strip() for g in df_anagrafica.get("Gruppo", pd.Series(dtype=str))
                                  if g and g.strip() and g.strip().lower() != "trasferiti"})

    riga_anagrafica_esistente = None
    if scelta_nome != opzione_nuova:
        corrispondenza = df_anagrafica[df_anagrafica["Cognome e Nome"] == scelta_nome]
        if not corrispondenza.empty:
            riga_anagrafica_esistente = corrispondenza.iloc[0].to_dict()

    nome_persona = scelta_nome
    aggiornamento_anagrafica = None

    if scelta_nome == opzione_nuova:
        st.caption("Questa persona non è ancora in Anagrafica: verrà creata con i dati sotto.")
        nome_persona = st.text_input("Nome completo", value=nome_suggerito, key="importa_s21_nuovo_nome")
        col_a, col_b = st.columns(2)
        with col_a:
            gruppo_nuovo = st.selectbox("Gruppo di servizio", gruppi_disponibili,
                                         key="importa_s21_nuovo_gruppo") if gruppi_disponibili else \
                st.text_input("Gruppo di servizio", key="importa_s21_nuovo_gruppo_testo")
        with col_b:
            st.caption("Stato")
            st.write("Attivo (nuova persona)")
        if nome_persona.strip():
            pannello_rif = dati_estratti["pannelli"][-1] if dati_estratti["pannelli"] else {}

            id_massimo = 0
            if "ID" in df_anagrafica.columns:
                ids_numerici = pd.to_numeric(df_anagrafica["ID"], errors="coerce").dropna()
                if not ids_numerici.empty:
                    id_massimo = int(ids_numerici.max())
            nuovo_id = id_massimo + 1

            aggiornamento_anagrafica = {
                "ID": nuovo_id,
                "Cognome e Nome": nome_persona.strip(),
                "Gruppo": gruppo_nuovo,
                "Attivi / Inattivi": "A",
                "Data Nascita": pannello_rif.get("data_nascita", ""),
                "Data Battesimo": pannello_rif.get("data_battesimo", ""),
                "Sesso": pannello_rif.get("sesso", ""),
                "A/U": pannello_rif.get("classe_spirituale", ""),
                "Incarico": pannello_rif.get("incarico", ""),
                "Tipo": pannello_rif.get("tipo", ""),
            }
    elif riga_anagrafica_esistente is not None:
        stato_attuale = categoria_stato_proclamatore(riga_anagrafica_esistente.get("Attivi / Inattivi", ""))
        if stato_attuale in ("TR", "I"):
            etichetta_stato = "Trasferito" if stato_attuale == "TR" else "Inattivo"
            st.info(f"📌 **{nome_persona}** risulta attualmente **{etichetta_stato}** in Anagrafica. "
                    f"Importando questi dati, lo riporto **Attivo** e gli assegno il gruppo scelto qui sotto.")
            gruppo_attuale = (riga_anagrafica_esistente.get("Gruppo") or "").strip()
            indice_default = (gruppi_disponibili.index(gruppo_attuale)
                               if gruppo_attuale in gruppi_disponibili else 0)
            gruppo_scelto = st.selectbox("Nuovo gruppo di servizio", gruppi_disponibili,
                                          index=indice_default if gruppi_disponibili else 0,
                                          key="importa_s21_gruppo_riattivazione") if gruppi_disponibili else \
                st.text_input("Nuovo gruppo di servizio", value=gruppo_attuale,
                               key="importa_s21_gruppo_riattivazione_testo")
            aggiornamento_anagrafica = dict(riga_anagrafica_esistente)
            aggiornamento_anagrafica["Gruppo"] = gruppo_scelto
            aggiornamento_anagrafica["Attivi / Inattivi"] = "A"
        else:
            st.caption(f"**{nome_persona}** è già Attivo in Anagrafica — nessuna modifica necessaria lì, "
                       "importo solo i mesi mancanti.")

    if not nome_persona or not nome_persona.strip():
        st.warning("Inserisci il nome della persona per continuare.")
        return

    st.subheader("2. Controlla i mesi da importare")

    df_tutti, err_tutti = leggi_foglio_tutti(workbook)
    if err_tutti:
        st.error(err_tutti)
        return
    mesi_gia_presenti = set()
    if not df_tutti.empty:
        col_nome_tutti = next((c for c in df_tutti.columns if str(c).strip().lower() in ("cognome e nome", "nome", "proclamatore")), df_tutti.columns[1] if len(df_tutti.columns) > 1 else None)
        col_mese_tutti = next((c for c in df_tutti.columns if str(c).strip().lower() in ("mese", "mese/anno")), df_tutti.columns[2] if len(df_tutti.columns) > 2 else None)

        if col_nome_tutti and col_mese_tutti:
            righe_persona_tutti = df_tutti[df_tutti[col_nome_tutti].astype(str).str.strip().str.lower() == nome_persona.strip().lower()]
            for m in righe_persona_tutti[col_mese_tutti].dropna():
                m_norm = normalizza_mese_anno(m)
                if isinstance(m_norm, list):
                    mesi_gia_presenti.update(m_norm)
                elif m_norm:
                    mesi_gia_presenti.add(m_norm)

    righe_proposte = _s21_costruisci_righe_import(dati_estratti, mesi_gia_presenti)

    for riga in righe_proposte:
        if "Mese/Anno" in riga:
            m_norm = normalizza_mese_anno(riga["Mese/Anno"])
            riga["Mese/Anno"] = m_norm[0] if isinstance(m_norm, list) else m_norm

    if not righe_proposte:
        st.info("Nessun mese con dati trovato nel PDF. Se è una scansione, aggiungi righe a mano con "
               "il tasto '+' della tabella qui sotto.")
        righe_proposte = [{
            "Mese (dal modulo)": "", "Anno servizio letto": "(manuale)", "Mese/Anno": "",
            "Tipo Servizio": "Proclamatore", "Ha partecipato al ministero": "Si",
            "Ore": "", "Studi Biblici": "", "Osservazioni": "",
            "gia_presente": False, "Importa": True,
        }]

    df_proposte = pd.DataFrame(righe_proposte)
    n_gia_presenti = int(df_proposte["gia_presente"].sum())
    if n_gia_presenti:
        st.caption(f"ℹ️ {n_gia_presenti} mese/i già presenti in archivio sono stati esclusi in automatico "
                   "(casella 'Importa' deselezionata) — puoi comunque riattivarli se necessario.")

    colonne_editor = ["Importa", "Mese/Anno", "Tipo Servizio", "Ha partecipato al ministero",
                       "Ore", "Studi Biblici", "Osservazioni", "Anno servizio letto", "Mese (dal modulo)"]

    tabella_modificata = st.data_editor(
        df_proposte[colonne_editor],
        hide_index=True,
        use_container_width=True,
        height=480,
        num_rows="dynamic",
        key="importa_s21_tabella_editor",
        disabled=sola_lettura(),
        column_config={
            "Importa": st.column_config.CheckboxColumn("Importa", width="small"),
            "Mese/Anno": st.column_config.TextColumn("Mese/Anno (AAAA-MM)", width="small", alignment="center",
                                                     help="Formato AAAA-MM, es. 2024-09"),
            "Tipo Servizio": st.column_config.SelectboxColumn(
                "Tipo Servizio", width="medium",
                options=["Proclamatore", "Pioniere Ausiliario", "Pioniere Regolare",
                         "Pioniere Speciale", "Missionario sul campo"]),
            "Ha partecipato al ministero": st.column_config.SelectboxColumn(
                "Ministero", width="small", options=["Si", ""]),
            "Ore": st.column_config.TextColumn(width="small", alignment="center"),
            "Studi Biblici": st.column_config.TextColumn("Studi", width="small", alignment="center"),
            "Osservazioni": st.column_config.TextColumn("Commenti:", width="large"),
            "Anno servizio letto": st.column_config.TextColumn("Anno letto sul modulo", width="small",
                                                               disabled=True, alignment="center"),
            "Mese (dal modulo)": st.column_config.TextColumn("Mese sul modulo", width="small", disabled=True, alignment="center"),
        },
    )

    st.subheader("3. Conferma")
    righe_da_importare = tabella_modificata[tabella_modificata["Importa"] == True]
    righe_da_importare = righe_da_importare[righe_da_importare["Mese/Anno"].astype(str).str.strip() != ""]
    st.caption(f"Righe selezionate per l'importazione: **{len(righe_da_importare)}**")

    if st.button(f"✅ Importa {len(righe_da_importare)} mese/i per {nome_persona}",
                 use_container_width=True,
                 disabled=(righe_da_importare.empty or sola_lettura())):
        errori = []
        importate = 0

        if aggiornamento_anagrafica is not None:
            riga_numero = None
            if riga_anagrafica_esistente is not None:
                df_con_indice = df_anagrafica.reset_index(drop=True)
                corrispondenza_idx = df_con_indice.index[
                    df_con_indice["Cognome e Nome"] == scelta_nome].tolist()
                if corrispondenza_idx:
                    riga_numero = RIGA_INTESTAZIONE_ANAGRAFICA + 1 + corrispondenza_idx[0]

            try:
                ok, err_salva = salva_riga_anagrafica(workbook, aggiornamento_anagrafica, riga_numero)
            except TypeError:
                ok, err_salva = salva_riga_anagrafica(workbook, aggiornamento_anagrafica)

            if not ok:
                errori.append(f"Anagrafica: {err_salva}")

        data_ora_consegna_pdf = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        mesi_presenti_ora = set(mesi_gia_presenti)

        for _, riga in righe_da_importare.iterrows():
            mese_anno = normalizza_mese_anno(riga["Mese/Anno"])
            if isinstance(mese_anno, list):
                mese_anno = mese_anno[0]

            if mese_anno in mesi_presenti_ora:
                errori.append(f"{mese_anno}: saltato, già presente in archivio.")
                continue

            valori_pdf = {
                "Informazioni cronologiche": data_ora_consegna_pdf,
                "Cognome e Nome": nome_persona.strip(),
                "Mese": str(mese_anno),
                "Hai servito come ?": riga["Tipo Servizio"],
                "Servizio": riga["Ha partecipato al ministero"],
                "Ore": riga["Ore"],
                "Cred. Ore": "",
                "Studi Biblici": riga["Studi Biblici"],
                "Commenti:": str(riga["Osservazioni"]).strip(),
                "Sorvegliante del gruppo": aggiornamento_anagrafica.get("Gruppo", "") if aggiornamento_anagrafica else "",
                "ND": "ND"
            }

            try:
                ok, err_salva = salva_riga_foglio(workbook, NOME_FOGLIO_TUTTI, RIGA_INTESTAZIONE_TUTTI, valori_pdf, None)
            except TypeError:
                try:
                    ok, err_salva = salva_riga_foglio(workbook, NOME_FOGLIO_TUTTI, RIGA_INTESTAZIONE_TUTTI, valori_pdf, riga_numero=None)
                except TypeError:
                    ok, err_salva = salva_riga_foglio(workbook, NOME_FOGLIO_TUTTI, RIGA_INTESTAZIONE_TUTTI, valori_pdf)

            if ok:
                importate += 1
                mesi_presenti_ora.add(mese_anno)
            else:
                errori.append(f"{mese_anno}: {err_salva}")

        st.cache_data.clear()

        if importate > 0:
            st.success(f"✔️ Importati correttamente {importate} mese/i per {nome_persona}!")
        if errori:
            st.warning("Alcune note o errori durante l'importazione:\n- " + "\n- ".join(errori))

        for k in ("importa_s21_dati", "importa_s21_chiave_file", "importa_s21_persona_scelta", "importa_s21_tabella_editor"):
            st.session_state.pop(k, None)
        st.rerun()
import streamlit.components.v1 as components

# ─────────────────────────────────────────────────────────────────
# PAGINA: IMPOSTAZIONI
# ─────────────────────────────────────────────────────────────────
def mostra_impostazioni():
    st.title("⚙️ Impostazioni")
    st.button("🏠 Torna alla Home", key="home_da_impostazioni", use_container_width=True,
              on_click=vai_a, args=("home",))

    if not collegato:
        st.warning("⚠️ Nessun foglio dati collegato.")
        return

    st.subheader("📅 Giorni delle adunanze")
    st.caption("Usati in «Presenti alle adunanze» per controllare la data inserita e proporre in automatico "
               "il tipo di adunanza giusto per quel giorno. Se cambiano in futuro, aggiornali qui — non "
               "serve toccare il codice del programma. Puoi scegliere anche più giorni per tipo.")

    giorni_attuali = leggi_giorni_adunanze_per_tipo(workbook)
    giorni_scelti = {}
    for tipo in TIPI_ADUNANZA:
        giorni_scelti[tipo] = st.multiselect(f"Giorni — {tipo}", GIORNI_SETTIMANA_IT,
                                             default=giorni_attuali.get(tipo, []),
                                             key=f"impostazioni_giorni_{tipo}",
                                             disabled=sola_lettura())

    tutti_vuoti = not any(giorni_scelti.values())
    if st.button("✔ Salva impostazione", type="primary", use_container_width=True,
                 disabled=tutti_vuoti or sola_lettura()):
        ok, err_salva = salva_giorni_adunanze_per_tipo(workbook, giorni_scelti)
        if ok:
            st.cache_data.clear()
            st.success("✔ Giorni delle adunanze aggiornati.")
        else:
            st.error(err_salva)

    if tutti_vuoti:
        st.info("Seleziona almeno un giorno per almeno un tipo di adunanza.")

    st.markdown("---")

    st.subheader("🔗 Link di Accesso Rapido Presenze")
    st.info("Condividi questo link con chi deve registrare solo le presenze. Chi lo apre vedrà **esclusivamente** la schermata di inserimento, senza poter accedere al resto del programma:")

    url_app = "https://gestioneseg.streamlit.app/?page=presenze"

    html_copia_link = f"""
    <div style="font-family: sans-serif; font-size: 14px;">
        <div style="background-color: #f0f2f6; padding: 10px; border-radius: 8px; border: 1px solid #d6d8db; margin-bottom: 10px; word-break: break-all;">
            <code style="color: #31333F; font-size: 13px;">{url_app}</code>
        </div>
        <button id="btnCopia" onclick="copiaLink()" style="
            background-color: #ff4b4b;
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 600;
            font-size: 14px;
            display: flex;
            align-items: center;
            gap: 6px;
        ">
            📋 Copia link
        </button>
        <div id="messaggioCopiato" style="display: none; color: #0f5132; background-color: #d1e7dd; border: 1px solid #badbcc; padding: 8px 12px; border-radius: 6px; margin-top: 10px; font-weight: 600;">
            ✅ Il link è stato copiato negli appunti!
        </div>
    </div>

    <script>
    function copiaLink() {{
        navigator.clipboard.writeText("{url_app}").then(function() {{
            var msg = document.getElementById("messaggioCopiato");
            msg.style.display = "block";
            setTimeout(function() {{
                msg.style.display = "none";
            }}, 3500);
        }}).catch(function(err) {{
            alert("Errore nella copia: " + err);
        }});
    }}
    </script>
    """

    components.html(html_copia_link, height=140)


# ─────────────────────────────────────────────────────────────────
# PAGINA: ACCESSI / GESTIONE UTENTI (solo Amministratore)
# ─────────────────────────────────────────────────────────────────
def _form_utente(editor: dict, df_utenti: pd.DataFrame):
    modo = editor.get("modo")
    e = editor.get("riga", {}) if modo == "modifica" else {}
    chiave = editor.get("numero_riga_foglio", "nuovo")

    if modo == "modifica":
        st.markdown(f"#### ✏️ Modifica accesso — {e.get('Utente', '')}")
        id_mostrato = e.get("ID", "") or "—"
    else:
        st.markdown("#### ➕ Nuovo accesso")
        id_mostrato = str(prossimo_id_anagrafica(df_utenti))

    with st.form(f"form_utente_{chiave}", clear_on_submit=False):
        st.text_input("ID", value=id_mostrato, disabled=True,
                      help="Assegnato automaticamente, non modificabile.")

        nome_utente = st.text_input("Nome e Cognome *", value=e.get("Utente", ""))
        email_utente = st.text_input(
            "Indirizzo email (Google) *", value=e.get("Indirizzo", ""),
            help="Deve corrispondere esattamente all'account Google con cui la persona farà l'accesso.")

        ruolo_corrente = e.get("Ruolo", "") or OPZIONI_RUOLO_UTENTE[0]
        if ruolo_corrente not in OPZIONI_RUOLO_UTENTE:
            ruolo_corrente = OPZIONI_RUOLO_UTENTE[0]
        ruolo_scelto = st.selectbox("Ruolo", OPZIONI_RUOLO_UTENTE,
                                     index=OPZIONI_RUOLO_UTENTE.index(ruolo_corrente))

        id_telegram = st.text_input("ID Telegram (opzionale)", value=e.get("Id telegram", ""),
                                    help="Verrà usato in futuro per l'invio di notifiche.")

        col_salva, col_annulla, col_elimina = st.columns(3)
        with col_salva:
            invia = st.form_submit_button("✔ Salva", type="primary", use_container_width=True)
        with col_annulla:
            annulla = st.form_submit_button("✖ Annulla", use_container_width=True)
        with col_elimina:
            elimina = st.form_submit_button("🗑️ Elimina", use_container_width=True,
                                            disabled=(modo != "modifica"))

    if annulla:
        st.session_state.utenti_editor = None
        st.rerun()

    if elimina and modo == "modifica":
        st.session_state.utenti_conferma_elimina = editor
        st.rerun()

    if invia:
        nome_pulito = nome_utente.strip()
        email_pulita = email_utente.strip().lower()

        email_gia_presenti = set()
        if not df_utenti.empty and "Indirizzo" in df_utenti.columns:
            email_gia_presenti = set(df_utenti["Indirizzo"].astype(str).str.strip().str.lower())
        if modo == "modifica":
            email_gia_presenti.discard((e.get("Indirizzo", "") or "").strip().lower())

        if not nome_pulito or not email_pulita:
            st.error("Nome e indirizzo email sono obbligatori.")
        elif not REGEX_EMAIL_VALIDA.match(email_pulita):
            st.error("L'indirizzo email non è in un formato valido (es. nome.cognome@gmail.com).")
        elif email_pulita in email_gia_presenti:
            st.error(f"Esiste già un utente con l'indirizzo «{email_pulita}».")
        else:
            valori = {
                "ID": id_mostrato,
                "Utente": nome_pulito,
                "Indirizzo": email_pulita,
                "Ruolo": ruolo_scelto,
                "Id telegram": id_telegram.strip(),
            }
            numero_riga = editor.get("numero_riga_foglio") if modo == "modifica" else None
            ok, err_salva = salva_riga_foglio(workbook, NOME_FOGLIO_UTENTI, RIGA_INTESTAZIONE_UTENTI,
                                               valori, riga_da_aggiornare=numero_riga)
            if ok:
                st.cache_data.clear()
                st.session_state.utenti_editor = None
                st.session_state.utenti_tabella_versione = st.session_state.get("utenti_tabella_versione", 0) + 1
                st.success(f"✔ «{nome_pulito}» salvato correttamente.")
                st.rerun()
            else:
                st.error(err_salva)

    conferma = st.session_state.get("utenti_conferma_elimina")
    if conferma and modo == "modifica" and conferma.get("numero_riga_foglio") == editor.get("numero_riga_foglio"):
        st.warning(f"Confermi l'eliminazione dell'accesso di «{e.get('Utente', '')}»? "
                   "L'operazione non è reversibile: la persona non potrà più accedere all'app.")
        col_si, col_no = st.columns(2)
        with col_si:
            if st.button("✔ Sì, elimina", key="utenti_conf_si", type="primary", use_container_width=True):
                ok, err_elim = elimina_riga_foglio(workbook, NOME_FOGLIO_UTENTI,
                                                    editor["numero_riga_foglio"])
                if ok:
                    st.cache_data.clear()
                    st.session_state.utenti_editor = None
                    st.session_state.utenti_conferma_elimina = None
                    st.session_state.utenti_tabella_versione = st.session_state.get("utenti_tabella_versione", 0) + 1
                    st.success("✔ Accesso eliminato.")
                    st.rerun()
                else:
                    st.error(err_elim)
        with col_no:
            if st.button("No, annulla", key="utenti_conf_no", use_container_width=True):
                st.session_state.utenti_conferma_elimina = None
                st.rerun()


def mostra_gestione_utenti():
    st.title("🔐 Accessi")
    st.button("🏠 Torna alla Home", key="home_da_utenti", use_container_width=True,
              on_click=vai_a, args=("home",))

    if st.session_state.get("ruolo") != "amministratore":
        st.warning("⚠️ Questa pagina è riservata agli amministratori.")
        return

    if not collegato:
        st.warning("⚠️ Nessun foglio dati collegato.")
        return

    if "utenti_tabella_versione" not in st.session_state:
        st.session_state.utenti_tabella_versione = 0

    contenitore_azioni = st.container()

    df_utenti, err = leggi_foglio_come_df(workbook, NOME_FOGLIO_UTENTI, RIGA_INTESTAZIONE_UTENTI)
    if err:
        st.error(err)
        return

    st.caption(f"Persone abilitate ad accedere all'app (foglio «{NOME_FOGLIO_UTENTI}»). "
               "Il ruolo stabilisce cosa possono vedere e modificare: Amministratore (accesso "
               "completo), Utente (sola lettura su tutte le pagine), Presenze (solo la pagina "
               "Presenti alle adunanze, con permessi pieni lì). Tocca una riga per selezionarla.")

    df_utenti_reset = df_utenti.reset_index(drop=True)
    idx_sel = None

    if df_utenti_reset.empty:
        st.info("Nessun utente presente ancora nel foglio.")
    else:
        colonne_mostrate = [c for c in ["ID", "Utente", "Indirizzo", "Ruolo", "Id telegram"]
                             if c in df_utenti_reset.columns]
        chiave_tabella = f"utenti_tabella_{st.session_state.utenti_tabella_versione}"
        evento = st.dataframe(
            df_utenti_reset[colonne_mostrate] if colonne_mostrate else df_utenti_reset,
            hide_index=True,
            use_container_width=True,
            on_select="rerun",
            selection_mode="single-row",
            key=chiave_tabella,
        )
        righe_sel = evento.selection.rows if evento and evento.selection else []
        if righe_sel and righe_sel[0] < len(df_utenti_reset):
            idx_sel = righe_sel[0]

    with contenitore_azioni:
        col_agg, col_mod = st.columns(2)
        with col_agg:
            if st.button("➕ Aggiungi nuovo accesso", key="utenti_apri_nuovo", use_container_width=True):
                st.session_state.utenti_editor = {"modo": "nuovo"}
                st.session_state.utenti_conferma_elimina = None
        with col_mod:
            if idx_sel is not None:
                if st.button("✏️ Modifica Accesso", key="utenti_apri_modifica", use_container_width=True):
                    numero_riga_foglio = RIGA_INTESTAZIONE_UTENTI + 1 + idx_sel
                    st.session_state.utenti_editor = {
                        "modo": "modifica",
                        "riga": df_utenti_reset.loc[idx_sel].to_dict(),
                        "numero_riga_foglio": numero_riga_foglio,
                    }
                    st.session_state.utenti_conferma_elimina = None

        editor = st.session_state.get("utenti_editor")
        if editor:
            _form_utente(editor, df_utenti)


# ─────────────────────────────────────────────────────────────────
# PAGINA: RAPPORTO PER LA FILIALE
# ─────────────────────────────────────────────────────────────────
CATEGORIE_FILIALE = [
    ("Proclamatore", "proclamatore"),
    ("Pioniere Ausiliario", "pioniere ausiliario"),
    ("Pioniere Regolare", "pioniere regolare"),
    ("Pioniere Speciale", "pioniere speciale"),
    ("Missionario sul campo", "missionario|rappresentante"),
]
COLORI_FILIALE = {
    "Proclamatore": "#1a1a1a",
    "Pioniere Ausiliario": "#1F77B4",
    "Pioniere Regolare": "#2E8B57",
    "Pioniere Speciale": "#9B30A0",
    "Missionario sul campo": "#B8860B",
}


def _filiale_mesi_disponibili(df_tutti: pd.DataFrame) -> list:
    mesi = set()
    if not df_tutti.empty and "Mese/Anno" in df_tutti.columns:
        for m in df_tutti["Mese/Anno"].dropna().unique():
            try:
                a, mm = str(m).split("-")
                mesi.add((int(a), int(mm)))
            except Exception:
                continue
    return sorted(mesi, reverse=True)


def _filiale_calcola_dati(df_tutti: pd.DataFrame, anno: int, mese: int):
    def _stesso_mese(mese_anno):
        try:
            a, m = str(mese_anno).split("-")
            return int(a) == anno and int(m) == mese
        except Exception:
            return False

    df_mese = df_tutti[df_tutti["Mese/Anno"].apply(_stesso_mese)] if not df_tutti.empty else df_tutti

    righe = []
    tot_rapporti = tot_ministero = 0
    tot_ore = tot_studi = 0.0
    for etichetta, parola in CATEGORIE_FILIALE:
        df_cat = df_mese[df_mese["Tipo Servizio"].str.lower().str.contains(parola, na=False, regex=True)] \
            if not df_mese.empty else df_mese
        n_rapporti = len(df_cat)
        if n_rapporti == 0:
            continue
        n_ministero = int(df_cat["Ha partecipato al ministero"].astype(bool).sum())
        ore = sum(a_float_it(v) for v in df_cat.get("Ore", []))
        studi = sum(a_float_it(v) for v in df_cat.get("Studi Biblici", []))
        righe.append({"tipo": etichetta, "rapporti": n_rapporti, "ministero": n_ministero,
                       "ore": ore, "studi": studi})
        tot_rapporti += n_rapporti
        tot_ministero += n_ministero
        tot_ore += ore
        tot_studi += studi

    totale = {"rapporti": tot_rapporti, "ministero": tot_ministero, "ore": tot_ore, "studi": tot_studi}
    return righe, totale


def mostra_rapporto_filiale():
    st.title("🏢 Rapporto per la Filiale")
    st.button("🏠 Torna alla Home", key="home_da_filiale", use_container_width=True,
              on_click=vai_a, args=("home",))
    st.caption("Dati statistici mensili (tipo modulo S-10), calcolati dal foglio Tutti.")

    if not collegato:
        st.warning("⚠️  Nessun foglio dati collegato.")
        return

    df_tutti, err_tutti = leggi_foglio_tutti(workbook)
    if err_tutti:
        st.error(err_tutti)
        return

    df_anagrafica, err_ana = leggi_foglio_come_df(workbook, NOME_FOGLIO_ANAGRAFICA, RIGA_INTESTAZIONE_ANAGRAFICA)
    if err_ana:
        st.error(err_ana)
        return

    mesi_disponibili = _filiale_mesi_disponibili(df_tutti)
    if not mesi_disponibili:
        st.info("Nessun rapporto trovato nel foglio Tutti.")
        return

    oggi = datetime.now()
    mese_precedente = oggi.month - 1 or 12
    anno_mese_precedente = oggi.year if oggi.month > 1 else oggi.year - 1
    indice_default = mesi_disponibili.index((anno_mese_precedente, mese_precedente)) \
        if (anno_mese_precedente, mese_precedente) in mesi_disponibili else 0
    scelta_mese = st.selectbox(
        "Mese",
        mesi_disponibili,
        index=indice_default,
        format_func=lambda am: f"{MESI_ITALIANI[am[1]]} {am[0]}",
        key="filiale_mese_scelto",
    )
    anno_scelto, mese_scelto = scelta_mese

    righe, totale = _filiale_calcola_dati(df_tutti, anno_scelto, mese_scelto)

    righe_html = ""
    for i, r in enumerate(righe):
        colore = COLORI_FILIALE.get(r["tipo"], "#1a1a1a")
        sfondo = "#EAF4FB" if i % 2 == 0 else "#FFFFFF"
        righe_html += (
            f"<tr style='color:{colore};background:{sfondo};'>"
            f"<td style='padding:6px 10px;font-weight:600;'>{r['tipo']}</td>"
            f"<td style='padding:6px 10px;text-align:center;'>{r['rapporti']}</td>"
            f"<td style='padding:6px 10px;text-align:center;'>{r['ministero']}</td>"
            f"<td style='padding:6px 10px;text-align:center;'>{formatta_numero_it(r['ore'])}</td>"
            f"<td style='padding:6px 10px;text-align:center;'>{formatta_numero_it(r['studi'])}</td>"
            f"</tr>"
        )

    tabella_html = f"""
    <table style="border-collapse:collapse;width:100%;border:2px dashed #999;">
        <tr style="background:#1B6FA8;color:#FFFFFF;">
            <th style="padding:8px 10px;text-align:left;">Tipo</th>
            <th style="padding:8px 10px;">Rapporti<br>Registrati</th>
            <th style="padding:8px 10px;">Forma<br>Minist</th>
            <th style="padding:8px 10px;">Ore</th>
            <th style="padding:8px 10px;">Studi</th>
        </tr>
        {righe_html}
        <tr style="color:#1B6FA8;font-weight:700;border-top:2px solid #1B6FA8;">
            <td style="padding:6px 10px;">Totale</td>
            <td style="padding:6px 10px;text-align:center;">{totale['rapporti']}</td>
            <td style="padding:6px 10px;text-align:center;">{totale['ministero']}</td>
            <td style="padding:6px 10px;text-align:center;">{formatta_numero_it(totale['ore'])}</td>
            <td style="padding:6px 10px;text-align:center;">{formatta_numero_it(totale['studi'])}</td>
        </tr>
    </table>
    """
    st.markdown(tabella_html, unsafe_allow_html=True)

    if not righe:
        st.info("Nessun rapporto trovato per questo mese nel foglio Tutti.")

    if "Attivi / Inattivi" in df_anagrafica.columns:
        n_attivi_anagrafica = int(
            (df_anagrafica["Attivi / Inattivi"].apply(categoria_stato_proclamatore) == "A").sum()
        )
    else:
        n_attivi_anagrafica = len(df_anagrafica)

    st.markdown(
        f"<div style='margin-top:18px;color:#2E8B57;font-weight:600;'>"
        f"N.ro proclamatori attivi in archivio utenti: {n_attivi_anagrafica}</div>",
        unsafe_allow_html=True,
    )


def _form_modifica_rapporto_tutti(dati_selezione: dict):
    nome = dati_selezione["nome"]
    mese_leggibile = dati_selezione["mese_leggibile"]
    riga_foglio = dati_selezione["riga_foglio"]
    grezza = dati_selezione["grezza"]
    bloccato = sola_lettura()

    st.title("Modifica rapporto")
    st.caption(f"{nome} — {mese_leggibile} (foglio «{NOME_FOGLIO_TUTTI}», riga {riga_foglio})")

    with st.form("form_modifica_tutti", clear_on_submit=False):
        mese_anno = st.text_input("Mese/Anno (formato AAAA-MM)", value=grezza[0], disabled=bloccato)

        opzioni_tipo = list(OPZIONI_HAI_SERVITO)
        valore_tipo = grezza[1]
        if valore_tipo and valore_tipo not in opzioni_tipo:
            opzioni_tipo = [valore_tipo] + opzioni_tipo
        indice_tipo = opzioni_tipo.index(valore_tipo) if valore_tipo in opzioni_tipo else 0
        tipo_servizio = st.selectbox("Ha servito come", opzioni_tipo, index=indice_tipo, disabled=bloccato)

        opzioni_ministero = ["Si", "No"]
        valore_ministero = grezza[2] if grezza[2] in opzioni_ministero else "No"
        ministero = st.selectbox("Ha partecipato al ministero", opzioni_ministero,
                                  index=opzioni_ministero.index(valore_ministero), disabled=bloccato)

        ore = st.number_input("Ore", value=a_float_it(grezza[4]), step=1.0, disabled=bloccato)
        cred_ore = st.number_input("Cred. Ore", value=a_float_it(grezza[5]), step=1.0, disabled=bloccato)
        studi = st.text_input("Studi Biblici", value=grezza[6], disabled=bloccato)
        osservazioni = st.text_area("Osservazioni", value=grezza[7], disabled=bloccato)

        col_btn1, col_btn2 = st.columns([1, 4])
        with col_btn1:
            invia = st.form_submit_button("✔ Salva", type="primary", use_container_width=True,
                                          disabled=bloccato)
        with col_btn2:
            annulla = st.form_submit_button("Annulla", use_container_width=True)

    if annulla:
        st.session_state.storico_modifica = None
        st.rerun()

    if invia:
        nuova_grezza = list(grezza)
        nuova_grezza[0] = "'" + mese_anno.strip()
        nuova_grezza[1] = tipo_servizio
        nuova_grezza[2] = ministero
        nuova_grezza[4] = formatta_numero_it(ore)
        nuova_grezza[5] = formatta_numero_it(cred_ore)
        nuova_grezza[6] = studi.strip()
        nuova_grezza[7] = osservazioni.strip()

        ok, err = salva_riga_tutti(workbook, riga_foglio, nuova_grezza)
        if ok:
            st.cache_data.clear()
            st.session_state.storico_modifica = None
            st.success("✔ Salvato correttamente.")
            st.rerun()
        else:
            st.error(err)

# ─────────────────────────────────────────────────────────────────
# Pagina: Storico rapporti consegnati
# ─────────────────────────────────────────────────────────────────
def mostra_storico_proclamatori():
    st.title("Storico rapporti consegnati")
    st.button("🏠 Torna alla Home", key="home_da_storico", use_container_width=True,
              on_click=vai_a, args=("home",))
    st.caption(f"Rapporti storici letti dal foglio «{NOME_FOGLIO_TUTTI}» "
               f"(intestazione riga {RIGA_INTESTAZIONE_TUTTI}).")

    if not collegato:
        st.warning("⚠️  Nessun foglio dati collegato.")
        return

    if "storico_modifica" not in st.session_state:
        st.session_state.storico_modifica = None

    if st.session_state.storico_modifica is not None:
        _form_modifica_rapporto_tutti(st.session_state.storico_modifica)
        return

    df_anagrafica, err_anagrafica = leggi_foglio_come_df(
        workbook, NOME_FOGLIO_ANAGRAFICA, RIGA_INTESTAZIONE_ANAGRAFICA)
    if err_anagrafica:
        st.error(err_anagrafica)
        return

    df_tutti, err_tutti = leggi_foglio_tutti(workbook)
    if err_tutti:
        st.error(err_tutti)
        return

    anni_presenti = anni_teocratici_per_menu(df_tutti)

    def formatta_anno_teocratico(valore):
        try:
            val = int(float(valore))
            return f"{val} – {val + 1} (set {val} → ago {val + 1})"
        except (ValueError, TypeError):
            return str(valore)

    col_anno, col_ricerca = st.columns([1, 3])
    with col_anno:
        anno_scelto = st.selectbox(
            "Anno teocratico",
            anni_presenti,
            format_func=formatta_anno_teocratico,
        )
    with col_ricerca:
        ricerca = st_keyup("🔍 Cerca per nome", placeholder="Digita per filtrare…", key="ricerca_storico_proclamatori")

    if df_anagrafica.empty or "Cognome e Nome" not in df_anagrafica.columns:
        st.info("Nessun Proclamatore trovato in Anagrafica.")
        return

    def stato_valido_anagrafica(valore: str) -> bool:
        v = (valore or "").strip().upper()
        return v.startswith("A") or v.startswith("I")

    colonna_stato = "Attivi / Inattivi" if "Attivi / Inattivi" in df_anagrafica.columns else None
    colonna_gruppo = "Gruppo" if "Gruppo" in df_anagrafica.columns else None

    stato_anagrafica_per_nome = {}
    gruppo_per_nome = {}
    for _, riga in df_anagrafica.iterrows():
        n = str(riga.get("Cognome e Nome", "")).strip()
        if not n:
            continue
        if colonna_stato:
            stato_anagrafica_per_nome[n] = str(riga.get(colonna_stato, "")).strip()
        if colonna_gruppo:
            gruppo_per_nome[n] = str(riga.get(colonna_gruppo, "")).strip()

    nomi = sorted(n for n in df_anagrafica["Cognome e Nome"].astype(str).str.strip().unique() if n)

    if colonna_stato:
        nomi = [n for n in nomi if stato_valido_anagrafica(stato_anagrafica_per_nome.get(n, ""))]

    testo_ricerca = ricerca.strip().lower()
    if testo_ricerca:
        nomi = [n for n in nomi if testo_ricerca in n.lower()]

    def calcola_stato_proclamatore(nome: str) -> str:
        st_anag = stato_anagrafica_per_nome.get(nome, "").strip().upper()

        if st_anag.startswith("I"):
            return "inattivo"

        col_partecipazione = "Ha partecipato al ministero"
        righe_p = df_tutti[df_tutti["Nome"].astype(str).str.strip().str.lower() == nome.lower()]

        if not righe_p.empty and col_partecipazione in righe_p.columns:
            righe_anno_corrente = righe_p[
                righe_p["Mese/Anno"].apply(anno_teocratico_di) == anno_scelto
            ].sort_values("Mese/Anno")

            if not righe_anno_corrente.empty:
                valori_no = [
                    str(v).strip().lower() in ["no", "false", "0"]
                    for v in righe_anno_corrente[col_partecipazione]
                ]

                consecutivi_no = 0
                for e_no in reversed(valori_no):
                    if e_no:
                        consecutivi_no += 1
                    else:
                        break

                if consecutivi_no >= 6:
                    return "inattivo"
                elif any(valori_no):
                    return "irregolare"

        return "attivo"

    mappa_stati = {n: calcola_stato_proclamatore(n) for n in nomi}

    tot_attivi = sum(1 for s in mappa_stati.values() if s == "attivo")
    tot_inattivi = sum(1 for s in mappa_stati.values() if s == "inattivo")
    tot_irregolari = sum(1 for s in mappa_stati.values() if s == "irregolare")

    scelta_stato = st.radio(
        "Filtra per stato:",
        options=["Tutti", "Attivi", "Inattivi", "Irregolari"],
        format_func=lambda op: {
            "Tutti": f"👥 Tutti ({len(nomi)})",
            "Attivi": f"🟢 Attivi ({tot_attivi})",
            "Inattivi": f"🔺 Inattivi ({tot_inattivi})",
            "Irregolari": f"⚠️ Irregolari ({tot_irregolari})"
        }[op],
        horizontal=True,
        key="filtro_stato_radio_storico"
    )

    if scelta_stato == "Attivi":
        nomi = [n for n in nomi if mappa_stati[n] == "attivo"]
    elif scelta_stato == "Inattivi":
        nomi = [n for n in nomi if mappa_stati[n] == "inattivo"]
    elif scelta_stato == "Irregolari":
        nomi = [n for n in nomi if mappa_stati[n] == "irregolare"]

    if not nomi:
        st.info("Nessun Proclamatore corrisponde ai criteri di ricerca.")
        return

    gruppi = {}
    for n in nomi:
        g = gruppo_per_nome.get(n, "") or "(Senza gruppo)"
        gruppi.setdefault(g, []).append(n)
    for g in gruppi:
        gruppi[g].sort()

    def _riga_proclamatore(nome: str):
        st_proc = mappa_stati.get(nome, "attivo")
        indicatore = "🔺 " if st_proc == "inattivo" else "⚠️ " if st_proc == "irregolare" else "🟢 "
        etichetta = f"{indicatore}{nome}"

        with st.expander(etichetta):
            righe_persona = df_tutti[df_tutti["Nome"].str.strip().str.lower() == nome.strip().lower()]
            righe_persona = righe_persona[
                righe_persona["Mese/Anno"].apply(anno_teocratico_di) == anno_scelto
            ]
            colonne_tabella = ["Anno di servizio", "Ha partecipato al ministero", "Studi Biblici",
                                "Pioniere ausiliario", "Ore", "Cred. Ore", "Osservazioni"]
            if righe_persona.empty:
                st.caption("Nessun rapporto trovato per l'anno teocratico selezionato.")
            else:
                righe_persona = righe_persona.sort_values("Mese/Anno")
                totale_ore = sum(a_float_it(v) for v in righe_persona["Ore"])
                totale_cred = sum(a_float_it(v) for v in righe_persona["Cred. Ore"])
                riga_totale = pd.DataFrame([{
                    "Anno di servizio": "Totale",
                    "Ha partecipato al ministero": None,
                    "Studi Biblici": "",
                    "Pioniere ausiliario": None,
                    "Ore": formatta_numero_it(totale_ore),
                    "Cred. Ore": formatta_numero_it(totale_cred),
                    "Osservazioni": "",
                }])
                tabella_completa = pd.concat(
                    [righe_persona[colonne_tabella], riga_totale[colonne_tabella]], ignore_index=True
                )
                evento_tabella = st.dataframe(
                    tabella_completa,
                    hide_index=True,
                    use_container_width=True,
                    on_select="rerun",
                    selection_mode="single-row",
                    key=f"storico_tabella_{nome}",
                    column_config={
                        "Anno di servizio": st.column_config.TextColumn(width="small"),
                        "Ha partecipato al ministero": st.column_config.CheckboxColumn(
                            "Ha partecipato al ministero", width="small", disabled=True),
                        "Studi Biblici": st.column_config.TextColumn(width="small"),
                        "Pioniere ausiliario": st.column_config.CheckboxColumn(
                            "Pioniere ausiliario", width="small", disabled=True),
                        "Ore": st.column_config.TextColumn(width="small"),
                        "Cred. Ore": st.column_config.TextColumn(width="small"),
                        "Osservazioni": st.column_config.TextColumn(width="large"),
                    },
                )

                righe_sel = evento_tabella.selection.rows if evento_tabella and evento_tabella.selection else []
                if righe_sel:
                    posizione = righe_sel[0]
                    if posizione < len(righe_persona):
                        idx_originale = righe_persona.index[posizione]
                        riga_dati = df_tutti.loc[idx_originale]
                        mese_leggibile = riga_dati["Anno di servizio"]
                        if st.button(f"✏️ Modifica «{mese_leggibile}»", key=f"storico_modifica_btn_{nome}_{posizione}"):
                            st.session_state.storico_modifica = {
                                "nome": nome,
                                "mese_leggibile": mese_leggibile,
                                "riga_foglio": int(riga_dati["RigaFoglio"]),
                                "grezza": list(riga_dati["_grezza"]),
                            }
                            st.rerun()

    for gruppo in sorted(gruppi.keys()):
        if gruppi[gruppo]:
            if gruppo == "(Senza gruppo)":
                st.markdown(
                    """
                    <div style="display: flex; align-items: center; gap: 10px; margin: 15px 0 10px 0;">
                        <span style="background-color: #f0f2f6; padding: 6px 12px; border-radius: 8px; font-size: 1.2rem;">📂</span>
                        <h4 style="margin: 0; color: #555;">Senza gruppo</h4>
                    </div>
                    """,
                    unsafe_allow_html=True
                )
            else:
                st.markdown(
                    f"""
                    <div style="
                        display: flex;
                        align-items: center;
                        justify-content: space-between;
                        background: linear-gradient(135deg, #1e293b 0%, #334155 100%);
                        color: white;
                        padding: 10px 18px;
                        border-radius: 10px;
                        margin: 20px 0 12px 0;
                        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
                    ">
                        <div style="display: flex; align-items: center; gap: 12px;">
                            <span style="background: rgba(255, 255, 255, 0.15); padding: 6px 10px; border-radius: 8px; font-size: 1.1rem;">👨‍💼</span>
                            <span style="font-size: 1.15rem; font-weight: 600; letter-spacing: 0.3px;">Gruppo: {gruppo}</span>
                        </div>
                        <span style="background: rgba(255, 255, 255, 0.2); padding: 3px 10px; border-radius: 12px; font-size: 0.85rem; font-weight: 500;">
                            {len(gruppi[gruppo])} proclamatori
                        </span>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

            for nome in gruppi[gruppo]:
                _riga_proclamatore(nome)
            st.divider()

# ─────────────────────────────────────────────────────────────────
# CONTROLLO ACCESSO RISTRETTO (DA RUOLO O LINK DIRETTO)
# ─────────────────────────────────────────────────────────────────
query_params = st.query_params
modalita_solo_presenze = (
    st.session_state.get("ruolo") == "presenze" or
    query_params.get("page") == "presenze" or
    query_params.get("modalita") == "presenze"
)

if modalita_solo_presenze:
    mostra_presenze_adunanze()
    st.stop()


# ─────────────────────────────────────────────────────────────────
# ROUTING COMPLETO — Accessibile solo per Amministratori
# ─────────────────────────────────────────────────────────────────
if st.session_state.pagina == "registrazioni":
    mostra_registrazioni()
elif st.session_state.pagina == "anagrafiche":
    mostra_anagrafiche()
elif st.session_state.pagina == "storico":
    mostra_storico_proclamatori()
elif st.session_state.pagina == "cartoline":
    mostra_cartoline_registrazione()
elif st.session_state.pagina == "riepilogo_statistiche":
    mostra_riepilogo_attivita()
elif st.session_state.pagina == "gruppi":
    mostra_gruppi_servizio()
elif st.session_state.pagina == "filiale":
    mostra_rapporto_filiale()
elif st.session_state.pagina == "presenze":
    mostra_presenze_adunanze()
elif st.session_state.pagina == "importa_s21":
    mostra_importa_s21()
elif st.session_state.pagina == "impostazioni":
    mostra_impostazioni()
elif st.session_state.pagina == "utenti":
    mostra_gestione_utenti()
else:
    mostra_home()

